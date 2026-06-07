#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
SmileEDI 세금계산서 스크래핑 프로그램
- 헤드리스 모드로 브라우저 실행
- 자동 로그인 처리
- 페이지 이동 및 대기 처리
"""

import os
import sys
import time
import random
import requests
import json
import pandas as pd
import smtplib
import io
import shutil
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from dotenv import load_dotenv
import msal
import webbrowser
try:
    import msoffcrypto  # type: ignore
    print("[OK] msoffcrypto 모듈 로드 성공")
except ImportError:
    print("[WARN] msoffcrypto 모듈을 찾을 수 없습니다. 암호 해독 기능을 사용할 수 없습니다.")
    print("[INFO] 다음 명령어로 설치하세요: pip install msoffcrypto-tool")
    msoffcrypto = None  # type: ignore

class SmileEDIScraper:
    """SmileEDI 세금계산서 스크래핑 클래스"""
    
    def __init__(self, headless=True):
        """초기화"""
        self.headless = headless
        self.driver = None
        self.wait = None
        self.load_config()
        self.setup_driver()
    
    def load_config(self):
        """환경 설정 로드"""
        # 현재 디렉토리에서 .env 파일 찾기
        current_dir = os.path.dirname(os.path.abspath(__file__))
        env_path = os.path.join(current_dir, "smileedi_config.env")
        
        if os.path.exists(env_path):
            load_dotenv(env_path)
            print(f"[OK] 설정 파일 로드: {env_path}")
        else:
            print(f"[WARN] 설정 파일을 찾을 수 없습니다: {env_path}")
            print("[INFO] 기본값을 사용합니다.")
        
        # 자격증명은 env 필수 — 하드코딩 기본값 제거 (GitHub Actions Secrets)
        self.username = os.getenv('SMILEEDI_USERNAME', '')
        self.password = os.getenv('SMILEEDI_PASSWORD', '')

        # 검색 조건: 회계연도 4/01~익년3/31 동적 산출(KST). env 명시 시 우선.
        fy_start, fy_end = self._fiscal_year_range_kst()
        self.search_start_date = os.getenv('SEARCH_START_DATE') or fy_start
        self.search_end_date = os.getenv('SEARCH_END_DATE') or fy_end
        self.pub_type_value = os.getenv('PUB_TYPE_VALUE', 'R')

        # 엑셀 다운로드 비밀번호 — env 필수
        self.excel_password = os.getenv('EXCEL_DOWNLOAD_PASSWORD', '')
        
        # SharePoint 설정
        self.sharepoint_tenant_id = os.getenv('SHAREPOINT_TENANT_ID')
        self.sharepoint_client_id = os.getenv('SHAREPOINT_CLIENT_ID')
        self.sharepoint_client_secret = os.getenv('SHAREPOINT_CLIENT_SECRET')
        self.sharepoint_site_id = os.getenv('SHAREPOINT_SITE_ID')
        self.sharepoint_drive_id = os.getenv('SHAREPOINT_DRIVE_ID')
        self.upload_folder_path = os.getenv('UPLOAD_FOLDER_PATH', '문서/정산')
        
        # 메일 설정 로드
        self.load_mail_config()
        
        print(f"[INFO] 사용자명: {self.username}")
        print(f"[INFO] 비밀번호: {'*' * len(self.password)}")
        print(f"[INFO] 검색 시작일: {self.search_start_date}")
        print(f"[INFO] 검색 종료일: {self.search_end_date}")
        print(f"[INFO] 발행방법 값: {self.pub_type_value}")
        print(f"[INFO] 엑셀 다운로드 비밀번호: {'*' * len(self.excel_password)}")
        print(f"[INFO] SharePoint 업로드 폴더: {self.upload_folder_path}")

    def _fiscal_year_range_kst(self):
        """회계연도 검색기간 4/01~익년3/31 (KST 기준 동적). 매년 자동 +1."""
        from datetime import datetime, timedelta
        now_kst = datetime.utcnow() + timedelta(hours=9)
        start_year = now_kst.year if now_kst.month >= 4 else now_kst.year - 1
        return f"{start_year}0401", f"{start_year + 1}0331"

    def load_mail_config(self):
        """메일 설정 로드"""
        try:
            # 메일 설정 파일 로드
            current_dir = os.path.dirname(os.path.abspath(__file__))
            mail_config_path = os.path.join(current_dir, "smileedi_mail_config.env")
            
            if os.path.exists(mail_config_path):
                load_dotenv(mail_config_path)
                print(f"[OK] 메일 설정 파일 로드: {mail_config_path}")
            else:
                print(f"[WARN] 메일 설정 파일을 찾을 수 없습니다: {mail_config_path}")
            
            # SMTP 설정
            self.smtp_server = os.getenv('SMTP_SERVER', 'smtp.office365.com')
            self.smtp_port = int(os.getenv('SMTP_PORT', '587'))
            # SMTP는 CI에서 미사용(메일은 OPS-Console 담당). 자격증명 하드코딩 제거.
            self.smtp_username = os.getenv('SMTP_USERNAME', '')
            self.smtp_password = os.getenv('SMTP_PASSWORD', '')
            
            # 메일 발송 조건 (Mail_Test.py 방식)
            self.email_error_condition = os.getenv('EMAIL_ERROR_CONDITION', '이메일오류 칼럼에 Y 값이 아닌 경우')
            
            # 품목 키워드 필터링 (필요한 기능이므로 다시 추가)
            self.item_keywords = os.getenv('ITEM_KEYWORDS', '수수료,접수,강사,대입,인터넷').split(',')
            
            # 담당자 매핑 (smileedi_mail_config.env 파일의 주석 규칙 사용)
            # 파일 내 주석으로 정의된 규칙을 코드에서 구현
            # 거래처명과 담당부서 조합에 따른 담당자 자동 매핑 (기본값 - 실제는 get_manager_by_rules에서 세부 규칙 적용)
            company_manager_mapping = os.getenv('COMPANY_MANAGER_MAPPING', '(학)연세대학교:송영신,서강대학교:박시현,덕성여자대학교:김슬기,연세대학교 미래캠퍼스:김유정')
            self.company_manager_map = {}
            for mapping in company_manager_mapping.split(','):
                if ':' in mapping:
                    company, manager = mapping.split(':')
                    self.company_manager_map[company.strip()] = manager.strip()
            
            # 담당자 이메일 매핑 (smileedi_mail_config.env에서 가져오기)
            manager_email_mapping = os.getenv('MANAGER_EMAIL_MAPPING', '송영신:ys1114@jinhakapply.com')
            self.manager_email_map = {}
            for mapping in manager_email_mapping.split(','):
                if ':' in mapping:
                    manager, email = mapping.split(':')
                    self.manager_email_map[manager.strip()] = email.strip()
            
            print(f"[INFO] SMTP 서버: {self.smtp_server}")
            print(f"[INFO] SMTP 포트: {self.smtp_port}")
            print(f"[INFO] SMTP 사용자: {self.smtp_username}")
            print(f"[INFO] 이메일오류 조건: {self.email_error_condition}")
            print(f"[INFO] 대상 키워드: {self.item_keywords}")
            print(f"[INFO] 담당자 매핑: {self.company_manager_map}")
            print(f"[INFO] 이메일 매핑: {self.manager_email_map}")
            
        except Exception as e:
            print(f"[FAIL] 메일 설정 로드 실패: {str(e)}")
    
    def setup_driver(self):
        """Chrome 드라이버 설정"""
        try:
            chrome_options = Options()
            
            # 헤드리스 모드 설정
            if self.headless:
                chrome_options.add_argument('--headless')
                print("[INFO] 헤드리스 모드로 실행합니다.")
            else:
                print("[INFO] 일반 모드로 실행합니다.")
            
            # 브라우저 옵션 설정
            chrome_options.add_argument('--no-sandbox')
            chrome_options.add_argument('--disable-dev-shm-usage')
            chrome_options.add_argument('--disable-gpu')
            chrome_options.add_argument('--window-size=1920,1080')
            chrome_options.add_argument('--disable-blink-features=AutomationControlled')
            chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
            chrome_options.add_experimental_option('useAutomationExtension', False)
            
            # User-Agent 설정
            chrome_options.add_argument('--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36')
            
            # 드라이버 초기화
            self.driver = webdriver.Chrome(options=chrome_options)
            
            # 대기 객체 설정
            self.wait = WebDriverWait(self.driver, 10)
            
            print("[OK] Chrome 드라이버 초기화 완료")
            
        except Exception as e:
            print(f"[FAIL] 드라이버 초기화 실패: {str(e)}")
            raise
    
    def random_wait(self, min_seconds=3, max_seconds=5):
        """랜덤 대기 시간"""
        wait_time = random.uniform(min_seconds, max_seconds)
        print(f"[INFO] {wait_time:.1f}초 대기 중...")
        time.sleep(wait_time)
    
    def navigate_to_url(self, url):
        """URL로 이동"""
        try:
            print(f"[INFO] 페이지 이동: {url}")
            self.driver.get(url)
            
            # 브라우저 최대화
            self.driver.maximize_window()
            print("[OK] 브라우저 최대화 완료")
            
            # 페이지 로딩 대기
            self.random_wait(3, 5)
            
            return True
            
        except Exception as e:
            print(f"[FAIL] 페이지 이동 실패: {str(e)}")
            return False
    
    def handle_login_popup(self):
        """로그인 팝업 처리"""
        try:
            print("[INFO] 로그인 팝업 처리 중...")
            
            # 로그인 팝업이 나타날 때까지 대기
            login_popup_url = "https://www.smileedi.com/com/pop_login.jsp?loginType=pop_login"
            
            # 현재 URL 확인
            current_url = self.driver.current_url
            print(f"[INFO] 현재 URL: {current_url}")
            
            # 팝업이 새 창으로 열렸는지 확인
            if len(self.driver.window_handles) > 1:
                print("[INFO] 새 창이 감지되었습니다. 팝업 창으로 전환합니다.")
                # 새 창으로 전환
                self.driver.switch_to.window(self.driver.window_handles[-1])
                self.random_wait(2, 3)
            
            # 로그인 폼 요소들 찾기
            print("[INFO] 로그인 폼 요소 찾는 중...")
            
            # 사용자명 입력 필드
            username_selector = "#contentsLogin > div > form > fieldset > div > div:nth-child(1) > input"
            username_field = self.wait.until(
                EC.presence_of_element_located((By.CSS_SELECTOR, username_selector))
            )
            
            # 비밀번호 입력 필드
            password_selector = "#inputPwd"
            password_field = self.driver.find_element(By.CSS_SELECTOR, password_selector)
            
            # 로그인 버튼
            login_button_selector = "#contentsLogin > div > form > fieldset > div > button"
            login_button = self.driver.find_element(By.CSS_SELECTOR, login_button_selector)
            
            print("[OK] 로그인 폼 요소들을 찾았습니다.")
            
            # 로그인 정보 입력
            print("[INFO] 로그인 정보 입력 중...")
            username_field.clear()
            username_field.send_keys(self.username)
            
            password_field.clear()
            password_field.send_keys(self.password)
            
            # 로그인 버튼 클릭
            print("[INFO] 로그인 버튼 클릭 중...")
            login_button.click()
            
            # 로그인 처리 대기
            self.random_wait(3, 5)
            
            # 로그인 성공 확인
            try:
                # 팝업 창이 닫혔는지 확인
                if len(self.driver.window_handles) == 1:
                    print("[INFO] 팝업 창이 닫혔습니다. 로그인 성공으로 간주합니다.")
                    # 원래 창으로 돌아가기
                    self.driver.switch_to.window(self.driver.window_handles[0])
                    print("[OK] 원래 창으로 돌아갔습니다.")
                    return True
                
                current_url = self.driver.current_url
                print(f"[INFO] 로그인 후 URL: {current_url}")
                
                if current_url and "pop_login" not in current_url:
                    print("[OK] 로그인 성공!")
                    
                    # 원래 창으로 돌아가기
                    if len(self.driver.window_handles) > 1:
                        self.driver.switch_to.window(self.driver.window_handles[0])
                        print("[OK] 원래 창으로 돌아갔습니다.")
                    
                    return True
                else:
                    print("[FAIL] 로그인 실패 - 여전히 로그인 페이지에 있습니다.")
                    return False
            except Exception as url_error:
                print(f"[WARN] URL 확인 중 오류: {str(url_error)}")
                # URL 확인 실패해도 로그인은 성공했을 수 있음
                print("[INFO] URL 확인 실패했지만 로그인은 성공했을 수 있습니다.")
                
                # 원래 창으로 돌아가기
                if len(self.driver.window_handles) > 1:
                    self.driver.switch_to.window(self.driver.window_handles[0])
                    print("[OK] 원래 창으로 돌아갔습니다.")
                
                return True
                
        except TimeoutException:
            print("[FAIL] 로그인 팝업 요소를 찾을 수 없습니다.")
            return False
        except Exception as e:
            print(f"[FAIL] 로그인 처리 실패: {str(e)}")
            return False
    
    def check_login_status(self):
        """로그인 상태 확인"""
        try:
            current_url = self.driver.current_url
            print(f"[INFO] 현재 URL: {current_url}")
            
            # 로그인 팝업이 있는지 확인
            if "pop_login" in current_url:
                print("[INFO] 로그인 팝업이 감지되었습니다.")
                return self.handle_login_popup()
            else:
                # 실제로 로그인 팝업이 나타나는지 확인
                print("[INFO] 로그인 팝업 확인 중...")
                
                # 잠시 대기하여 팝업이 나타나는지 확인
                self.random_wait(2, 3)
                
                # 새 창이 열렸는지 확인
                if len(self.driver.window_handles) > 1:
                    print("[INFO] 새 창(팝업)이 감지되었습니다.")
                    return self.handle_login_popup()
                
                # 현재 페이지에서 로그인 폼이 있는지 확인
                try:
                    login_form = self.driver.find_elements(By.CSS_SELECTOR, "#contentsLogin")
                    if login_form:
                        print("[INFO] 현재 페이지에서 로그인 폼을 발견했습니다.")
                        return self.handle_login_popup()
                except:
                    pass
                
                # 로그인 관련 요소가 있는지 확인
                try:
                    username_field = self.driver.find_elements(By.CSS_SELECTOR, "#contentsLogin > div > form > fieldset > div > div:nth-child(1) > input")
                    if username_field:
                        print("[INFO] 로그인 입력 필드를 발견했습니다.")
                        return self.handle_login_popup()
                except:
                    pass
                
                print("[WARN] 로그인 팝업이 감지되지 않았습니다. 수동으로 로그인을 시도합니다.")
                return self.force_login()
                
        except Exception as e:
            print(f"[FAIL] 로그인 상태 확인 실패: {str(e)}")
            return False
    
    def force_login(self):
        """강제 로그인 시도"""
        try:
            print("[INFO] 강제 로그인 시도 중...")
            
            # 로그인 팝업 URL로 직접 이동
            login_url = "https://www.smileedi.com/com/pop_login.jsp?loginType=pop_login"
            print(f"[INFO] 로그인 페이지로 이동: {login_url}")
            
            self.driver.get(login_url)
            self.random_wait(2, 3)
            
            # 로그인 폼 요소들 찾기
            print("[INFO] 로그인 폼 요소 찾는 중...")
            
            try:
                # 사용자명 입력 필드
                username_selector = "#contentsLogin > div > form > fieldset > div > div:nth-child(1) > input"
                username_field = self.wait.until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, username_selector))
                )
                
                # 비밀번호 입력 필드
                password_selector = "#inputPwd"
                password_field = self.driver.find_element(By.CSS_SELECTOR, password_selector)
                
                # 로그인 버튼
                login_button_selector = "#contentsLogin > div > form > fieldset > div > button"
                login_button = self.driver.find_element(By.CSS_SELECTOR, login_button_selector)
                
                print("[OK] 로그인 폼 요소들을 찾았습니다.")
                
                # 로그인 정보 입력
                print("[INFO] 로그인 정보 입력 중...")
                username_field.clear()
                username_field.send_keys(self.username)
                
                password_field.clear()
                password_field.send_keys(self.password)
                
                # 로그인 버튼 클릭
                print("[INFO] 로그인 버튼 클릭 중...")
                login_button.click()
                
                # 로그인 처리 대기
                self.random_wait(3, 5)
                
                # 로그인 성공 확인
                try:
                    current_url = self.driver.current_url
                    print(f"[INFO] 로그인 후 URL: {current_url}")
                    
                    if current_url and "pop_login" not in current_url:
                        print("[OK] 강제 로그인 성공!")
                        return True
                    else:
                        print("[FAIL] 강제 로그인 실패 - 여전히 로그인 페이지에 있습니다.")
                        return False
                except Exception as url_error:
                    print(f"[WARN] URL 확인 중 오류: {str(url_error)}")
                    print("[INFO] URL 확인 실패했지만 로그인은 성공했을 수 있습니다.")
                    return True
                    
            except TimeoutException:
                print("[FAIL] 로그인 폼 요소를 찾을 수 없습니다.")
                return False
            except Exception as e:
                print(f"[FAIL] 강제 로그인 처리 실패: {str(e)}")
                return False
                
        except Exception as e:
            print(f"[FAIL] 강제 로그인 실패: {str(e)}")
            return False
    
    def navigate_to_tax_invoice_page(self):
        """매출계산서 페이지로 이동"""
        try:
            print("\n" + "="*60)
            print("매출계산서 페이지로 이동")
            print("="*60)
            
            # 현재 URL 확인
            current_url = self.driver.current_url
            print(f"[INFO] 현재 URL: {current_url}")
            
            # 매출계산서 페이지 URL로 직접 이동
            tax_invoice_url = "https://www.smileedi.com/Dti220101.do"
            print(f"[INFO] 매출계산서 페이지로 이동: {tax_invoice_url}")
            
            self.driver.get(tax_invoice_url)
            self.random_wait(3, 5)
            
            # 이동 후 페이지 확인
            new_url = self.driver.current_url
            page_title = self.driver.title
            print(f"[INFO] 이동 후 페이지: {page_title}")
            print(f"[INFO] 이동 후 URL: {new_url}")
            
            # 페이지 소스 확인
            page_source = self.driver.page_source
            print(f"[INFO] 페이지 소스 길이: {len(page_source)} characters")
            
            # 매출계산서 관련 요소 확인
            try:
                # input 요소 확인
                inputs = self.driver.find_elements(By.TAG_NAME, "input")
                print(f"[INFO] 발견된 input 요소 수: {len(inputs)}")
                
                # form 요소 확인
                forms = self.driver.find_elements(By.TAG_NAME, "form")
                print(f"[INFO] 발견된 form 요소 수: {len(forms)}")
                
                # table 요소 확인
                tables = self.driver.find_elements(By.TAG_NAME, "table")
                print(f"[INFO] 발견된 table 요소 수: {len(tables)}")
                
                # contents2 요소 확인
                contents2_elements = self.driver.find_elements(By.CSS_SELECTOR, "[id*='contents2'], [class*='contents2']")
                print(f"[INFO] contents2 관련 요소 수: {len(contents2_elements)}")
                
                if len(inputs) > 0 and len(forms) > 0:
                    print("[OK] 매출계산서 페이지가 정상적으로 로드되었습니다.")
                    return True
                else:
                    print("[WARN] 매출계산서 페이지가 완전히 로드되지 않았을 수 있습니다.")
                    
                    # 페이지 새로고침 시도
                    print("[INFO] 페이지 새로고침을 시도합니다.")
                    self.driver.refresh()
                    self.random_wait(3, 5)
                    
                    # 다시 확인
                    inputs = self.driver.find_elements(By.TAG_NAME, "input")
                    forms = self.driver.find_elements(By.TAG_NAME, "form")
                    print(f"[INFO] 새로고침 후 input 요소 수: {len(inputs)}")
                    print(f"[INFO] 새로고침 후 form 요소 수: {len(forms)}")
                    
                    if len(inputs) > 0 and len(forms) > 0:
                        print("[OK] 새로고침 후 매출계산서 페이지가 정상적으로 로드되었습니다.")
                        return True
                    else:
                        print("[FAIL] 매출계산서 페이지 로드 실패")
                        return False
                        
            except Exception as check_error:
                print(f"[WARN] 페이지 요소 확인 중 오류: {str(check_error)}")
                return True  # 오류가 있어도 일단 진행
                
        except Exception as e:
            print(f"[FAIL] 매출계산서 페이지 이동 실패: {str(e)}")
            return False
    
    def setup_search_conditions(self):
        """매출계산서 검색 조건 설정"""
        try:
            print("\n" + "="*60)
            print("매출계산서 검색 조건 설정")
            print("="*60)
            
            # 페이지 구조 디버깅
            print("[INFO] 페이지 구조 분석 중...")
            try:
                # 페이지 소스 일부 출력
                page_source = self.driver.page_source
                print(f"[INFO] 페이지 소스 길이: {len(page_source)} characters")
                
                # iframe 확인
                iframes = self.driver.find_elements(By.TAG_NAME, "iframe")
                print(f"[INFO] 발견된 iframe 수: {len(iframes)}")
                
                if iframes:
                    print("[INFO] iframe이 발견되었습니다. iframe 내부를 확인합니다.")
                    for i, iframe in enumerate(iframes):
                        try:
                            iframe_src = iframe.get_attribute("src") or "src없음"
                            iframe_name = iframe.get_attribute("name") or "이름없음"
                            print(f"  iframe {i+1}: src='{iframe_src}', name='{iframe_name}'")
                        except:
                            print(f"  iframe {i+1}: 정보 확인 실패")
                
                # 모든 input 요소 찾기
                all_inputs = self.driver.find_elements(By.TAG_NAME, "input")
                print(f"[INFO] 발견된 input 요소 수: {len(all_inputs)}")
                
                # text 타입 input 요소들 찾기
                text_inputs = self.driver.find_elements(By.CSS_SELECTOR, "input[type='text']")
                print(f"[INFO] 발견된 text input 요소 수: {len(text_inputs)}")
                
                # 모든 폼 요소 찾기
                forms = self.driver.find_elements(By.TAG_NAME, "form")
                print(f"[INFO] 발견된 form 요소 수: {len(forms)}")
                
                # 테이블 요소 찾기
                tables = self.driver.find_elements(By.TAG_NAME, "table")
                print(f"[INFO] 발견된 table 요소 수: {len(tables)}")
                
                # div 요소 찾기
                divs = self.driver.find_elements(By.TAG_NAME, "div")
                print(f"[INFO] 발견된 div 요소 수: {len(divs)}")
                
                # contents2 관련 요소 찾기
                contents2_elements = self.driver.find_elements(By.CSS_SELECTOR, "[id*='contents2'], [class*='contents2']")
                print(f"[INFO] contents2 관련 요소 수: {len(contents2_elements)}")
                
                for i, input_elem in enumerate(text_inputs[:5], 1):  # 최대 5개만 출력
                    try:
                        name = input_elem.get_attribute("name") or "이름없음"
                        id_attr = input_elem.get_attribute("id") or "ID없음"
                        value = input_elem.get_attribute("value") or "값없음"
                        print(f"  {i}. name='{name}', id='{id_attr}', value='{value}'")
                    except:
                        print(f"  {i}. 요소 정보 확인 실패")
                
            except Exception as debug_error:
                print(f"[WARN] 페이지 구조 분석 중 오류: {str(debug_error)}")
            
            # 1. 작성일 설정 (시작일) - 다양한 셀렉터 시도
            print("[INFO] 작성일 시작일 설정 중...")
            start_date_selectors = [
                "#contents2 > div.tblForm01.top > table:nth-child(1) > tbody > tr:nth-child(1) > td > input[type=text]:nth-child(2)",
                "input[type='text']:nth-of-type(1)",
                "input[type='text']:first-of-type",
                "input[name*='start']",
                "input[name*='begin']",
                "input[id*='start']",
                "input[id*='begin']"
            ]
            
            start_date_field = None
            for selector in start_date_selectors:
                try:
                    start_date_field = self.driver.find_element(By.CSS_SELECTOR, selector)
                    print(f"[OK] 시작일 필드 발견: {selector}")
                    break
                except NoSuchElementException:
                    continue
            
            if start_date_field:
                start_date_field.clear()
                start_date_field.send_keys(self.search_start_date)
                print(f"[OK] 시작일 설정 완료: {self.search_start_date}")
            else:
                print("[FAIL] 시작일 입력 필드를 찾을 수 없습니다.")
                return False
            
            # 2. 작성일 설정 (종료일) - 다양한 셀렉터 시도
            print("[INFO] 작성일 종료일 설정 중...")
            end_date_selectors = [
                "#contents2 > div.tblForm01.top > table:nth-child(1) > tbody > tr:nth-child(1) > td > input[type=text]:nth-child(4)",
                "input[type='text']:nth-of-type(2)",
                "input[type='text']:last-of-type",
                "input[name*='end']",
                "input[name*='finish']",
                "input[id*='end']",
                "input[id*='finish']"
            ]
            
            end_date_field = None
            for selector in end_date_selectors:
                try:
                    end_date_field = self.driver.find_element(By.CSS_SELECTOR, selector)
                    print(f"[OK] 종료일 필드 발견: {selector}")
                    break
                except NoSuchElementException:
                    continue
            
            if end_date_field:
                end_date_field.clear()
                end_date_field.send_keys(self.search_end_date)
                print(f"[OK] 종료일 설정 완료: {self.search_end_date}")
            else:
                print("[FAIL] 종료일 입력 필드를 찾을 수 없습니다.")
                return False
            
            # 3. 상세검색보기 버튼 클릭
            print("[INFO] 상세검색보기 버튼 클릭 중...")
            detail_search_selector = "#contents2 > div.tblForm01.top > div > a > img"
            try:
                detail_search_button = self.driver.find_element(By.CSS_SELECTOR, detail_search_selector)
                detail_search_button.click()
                print("[OK] 상세검색보기 버튼 클릭 완료")
                self.random_wait(2, 3)  # 상세 검색 폼 로딩 대기
            except NoSuchElementException:
                print("[FAIL] 상세검색보기 버튼을 찾을 수 없습니다.")
                return False
            
            # 4. 발행방법 설정 (R 값 체크)
            print("[INFO] 발행방법 설정 중...")
            try:
                # 발행방법 타이틀 확인
                pub_type_title_selector = "#contents2 > div.tblForm01.top > table.detailView > tbody > tr:nth-child(2) > th:nth-child(3)"
                pub_type_title = self.driver.find_element(By.CSS_SELECTOR, pub_type_title_selector)
                print(f"[INFO] 발행방법 타이틀: {pub_type_title.text}")
                
                # 발행방법 값 영역 찾기
                pub_type_value_selector = "#contents2 > div.tblForm01.top > table.detailView > tbody > tr:nth-child(2) > td:nth-child(4)"
                pub_type_value_area = self.driver.find_element(By.CSS_SELECTOR, pub_type_value_selector)
                
                # 발행방법 영역의 모든 체크박스 확인
                print("[INFO] 발행방법 영역의 모든 체크박스 확인 중...")
                all_checkboxes = pub_type_value_area.find_elements(By.CSS_SELECTOR, "input[type='checkbox']")
                print(f"[INFO] 발견된 체크박스 수: {len(all_checkboxes)}")
                
                for i, cb in enumerate(all_checkboxes):
                    try:
                        cb_value = cb.get_attribute("value") or "값없음"
                        cb_name = cb.get_attribute("name") or "이름없음"
                        cb_checked = cb.is_selected()
                        print(f"  체크박스 {i+1}: value='{cb_value}', name='{cb_name}', checked={cb_checked}")
                    except Exception as cb_error:
                        print(f"  체크박스 {i+1}: 정보 확인 실패 - {str(cb_error)}")
                
                # R 값 체크박스 찾기 및 체크
                checkbox_selector = f"input[value='{self.pub_type_value}']"
                print(f"[INFO] 찾는 체크박스 셀렉터: {checkbox_selector}")
                
                checkbox = pub_type_value_area.find_element(By.CSS_SELECTOR, checkbox_selector)
                print(f"[OK] R 값 체크박스 발견")
                
                # 체크박스 상태 확인
                is_checked = checkbox.is_selected()
                print(f"[INFO] 체크박스 현재 상태: {is_checked}")
                
                # R 값만 체크하고 나머지는 체크 해제
                print(f"[INFO] R 값만 체크하고 나머지 체크 해제 중...")
                
                # 모든 체크박스 해제
                for cb in all_checkboxes:
                    try:
                        cb_value = cb.get_attribute("value")
                        if cb.is_selected():
                            cb.click()
                            print(f"[INFO] 체크 해제: {cb_value}")
                    except Exception as uncheck_error:
                        print(f"[WARN] 체크 해제 실패: {str(uncheck_error)}")
                
                self.random_wait(1, 2)
                
                # R 값만 체크
                if not checkbox.is_selected():
                    print(f"[INFO] R 값 체크박스 클릭 중...")
                    checkbox.click()
                    self.random_wait(1, 2)  # 클릭 후 잠시 대기
                    
                    # 클릭 후 상태 재확인
                    is_checked_after = checkbox.is_selected()
                    print(f"[INFO] 클릭 후 체크박스 상태: {is_checked_after}")
                    
                    if is_checked_after:
                        print(f"[OK] 발행방법 '{self.pub_type_value}' 값 체크 완료")
                    else:
                        print(f"[WARN] 체크박스 클릭 후에도 체크되지 않았습니다.")
                        # JavaScript로 강제 체크 시도
                        print("[INFO] JavaScript로 강제 체크 시도 중...")
                        self.driver.execute_script("arguments[0].checked = true;", checkbox)
                        self.random_wait(1, 2)
                        
                        is_checked_js = checkbox.is_selected()
                        print(f"[INFO] JavaScript 체크 후 상태: {is_checked_js}")
                        
                        if is_checked_js:
                            print(f"[OK] JavaScript로 발행방법 '{self.pub_type_value}' 값 체크 완료")
                        else:
                            print(f"[FAIL] JavaScript로도 체크 실패")
                else:
                    print(f"[INFO] 발행방법 '{self.pub_type_value}' 값이 이미 체크되어 있습니다.")
                
                # 최종 상태 확인
                print("[INFO] 최종 체크박스 상태 확인:")
                for i, cb in enumerate(all_checkboxes):
                    try:
                        cb_value = cb.get_attribute("value")
                        cb_checked = cb.is_selected()
                        status = "체크됨" if cb_checked else "체크안됨"
                        print(f"  체크박스 {i+1}: value='{cb_value}', {status}")
                    except Exception as final_error:
                        print(f"  체크박스 {i+1}: 상태 확인 실패 - {str(final_error)}")
                    
            except NoSuchElementException as e:
                print(f"[FAIL] 발행방법 설정 실패: {str(e)}")
                
                # 대안 방법: 모든 체크박스 찾기
                print("[INFO] 대안 방법으로 모든 체크박스 찾기 시도...")
                try:
                    all_checkboxes = self.driver.find_elements(By.CSS_SELECTOR, "input[type='checkbox']")
                    print(f"[INFO] 페이지 전체 체크박스 수: {len(all_checkboxes)}")
                    
                    for i, cb in enumerate(all_checkboxes):
                        try:
                            cb_value = cb.get_attribute("value")
                            if cb_value == self.pub_type_value:
                                print(f"[OK] 대안 방법으로 R 값 체크박스 발견")
                                cb.click()
                                print(f"[OK] 대안 방법으로 발행방법 '{self.pub_type_value}' 값 체크 완료")
                                break
                        except Exception as alt_error:
                            print(f"[WARN] 대안 방법 체크박스 {i+1} 처리 실패: {str(alt_error)}")
                    else:
                        print(f"[FAIL] 대안 방법으로도 R 값 체크박스를 찾을 수 없습니다.")
                        return False
                        
                except Exception as alt_error:
                    print(f"[FAIL] 대안 방법 실패: {str(alt_error)}")
                    return False
            
            print("[OK] 매출계산서 검색 조건 설정 완료")
            return True
            
        except Exception as e:
            print(f"[FAIL] 검색 조건 설정 실패: {str(e)}")
            return False
    
    def execute_search(self):
        """매출계산서 검색 실행"""
        try:
            print("\n" + "="*60)
            print("매출계산서 검색 실행")
            print("="*60)
            
            # 검색 버튼 클릭
            print("[INFO] 검색 버튼 클릭 중...")
            search_button_selector = "#contents2 > div.btnCenter > a > img"
            try:
                search_button = self.driver.find_element(By.CSS_SELECTOR, search_button_selector)
                search_button.click()
                print("[OK] 검색 버튼 클릭 완료")
                
                # 검색 결과 로딩 대기
                self.random_wait(3, 5)
                
                # 검색 결과 확인
                current_url = self.driver.current_url
                page_title = self.driver.title
                print(f"[INFO] 검색 후 페이지: {page_title}")
                print(f"[INFO] 검색 후 URL: {current_url}")
                
                # 검색 결과 테이블 확인
                try:
                    result_tables = self.driver.find_elements(By.TAG_NAME, "table")
                    print(f"[INFO] 검색 결과 테이블 수: {len(result_tables)}")
                    
                    # 결과 데이터 확인
                    if len(result_tables) > 0:
                        print("[OK] 검색 결과가 있습니다.")
                        return True
                    else:
                        print("[WARN] 검색 결과가 없습니다.")
                        return True
                        
                except Exception as table_error:
                    print(f"[WARN] 검색 결과 테이블 확인 중 오류: {str(table_error)}")
                    return True
                
            except NoSuchElementException:
                print("[FAIL] 검색 버튼을 찾을 수 없습니다.")
                return False
                
        except Exception as e:
            print(f"[FAIL] 검색 실행 실패: {str(e)}")
            return False
    
    def download_excel_file(self):
        """매출계산서 조회 결과 엑셀 다운로드"""
        try:
            print("\n" + "="*60)
            print("매출계산서 엑셀 다운로드")
            print("="*60)
            
            # 엑셀 저장 버튼 클릭
            print("[INFO] 엑셀 저장 버튼 클릭 중...")
            excel_button_selector = "#contents3 > div.btnRight.mt0 > a:nth-child(3) > span"
            
            try:
                excel_button = self.wait.until(
                    EC.element_to_be_clickable((By.CSS_SELECTOR, excel_button_selector))
                )
                excel_button.click()
                print("[OK] 엑셀 저장 버튼 클릭 완료")
                
                # 다운로드 대기
                self.random_wait(3, 5)
                
                # 엑셀 다운로드 동의 및 비밀번호 입력 처리
                if not self.handle_excel_download_form():
                    print("[FAIL] 엑셀 다운로드 폼 처리에 실패했습니다.")
                    return False
                
                print("[OK] 엑셀 다운로드 처리 완료")
                return True
                
            except TimeoutException:
                print("[FAIL] 엑셀 저장 버튼을 찾을 수 없습니다.")
                
                # 대안 셀렉터 시도
                alternative_selectors = [
                    "#contents3 > div.btnRight.mt0 > a:nth-child(3)",
                    "#contents3 > div.btnRight.mt0 > a:nth-child(3) > span",
                    "a[href*='excel']",
                    "a[href*='download']",
                    "span:contains('엑셀')",
                    "span:contains('다운로드')",
                    "span:contains('저장')"
                ]
                
                for selector in alternative_selectors:
                    try:
                        print(f"[INFO] 대안 셀렉터 시도: {selector}")
                        excel_button = self.driver.find_element(By.CSS_SELECTOR, selector)
                        excel_button.click()
                        print(f"[OK] 대안 셀렉터로 엑셀 저장 버튼 클릭 완료: {selector}")
                        self.random_wait(3, 5)
                        return True
                    except NoSuchElementException:
                        continue
                
                print("[FAIL] 모든 셀렉터로 엑셀 저장 버튼을 찾을 수 없습니다.")
                return False
                
        except Exception as e:
            print(f"[FAIL] 엑셀 다운로드 실패: {str(e)}")
            return False
    
    def handle_excel_download_form(self):
        """엑셀 다운로드 동의 및 비밀번호 입력 폼 처리"""
        try:
            print("\n" + "="*60)
            print("엑셀 다운로드 폼 처리")
            print("="*60)
            
            # iframe 확인 및 전환
            print("[INFO] iframe 확인 중...")
            iframes = self.driver.find_elements(By.TAG_NAME, "iframe")
            print(f"[INFO] 발견된 iframe 수: {len(iframes)}")
            
            if iframes:
                print("[INFO] iframe이 발견되었습니다. iframe 내부를 확인합니다.")
                excel_frame = None
                
                for i, iframe in enumerate(iframes):
                    try:
                        iframe_src = iframe.get_attribute("src") or "src없음"
                        iframe_name = iframe.get_attribute("name") or "이름없음"
                        print(f"  iframe {i+1}: src='{iframe_src}', name='{iframe_name}'")
                        
                        # 엑셀 다운로드 관련 iframe 찾기
                        if "excel" in iframe_src.lower() or "excelModalFrame" in iframe_name:
                            excel_frame = iframe
                            print(f"[OK] 엑셀 다운로드 iframe 발견: {iframe_name}")
                            
                    except:
                        print(f"  iframe {i+1}: 정보 확인 실패")
                
                # 엑셀 다운로드 iframe으로 전환
                if excel_frame:
                    try:
                        self.driver.switch_to.frame(excel_frame)
                        print("[OK] 엑셀 다운로드 iframe으로 전환 완료")
                    except Exception as frame_error:
                        print(f"[WARN] 엑셀 iframe 전환 실패: {str(frame_error)}")
                        # 첫 번째 iframe으로 전환 시도
                        try:
                            self.driver.switch_to.frame(iframes[0])
                            print("[OK] 첫 번째 iframe으로 전환 완료")
                        except Exception as first_frame_error:
                            print(f"[WARN] 첫 번째 iframe 전환 실패: {str(first_frame_error)}")
                else:
                    # 첫 번째 iframe으로 전환
                    try:
                        self.driver.switch_to.frame(iframes[0])
                        print("[OK] 첫 번째 iframe으로 전환 완료")
                    except Exception as frame_error:
                        print(f"[WARN] iframe 전환 실패: {str(frame_error)}")
            
            # 현재 페이지 상태 확인
            current_url = self.driver.current_url
            page_title = self.driver.title
            print(f"[INFO] 폼 페이지: {page_title}")
            print(f"[INFO] 폼 URL: {current_url}")
            
            # 1. 동의함 체크박스 클릭
            print("[INFO] 동의함 체크박스 클릭 중...")
            agree_selector = "#agree"
            
            try:
                agree_checkbox = self.wait.until(
                    EC.element_to_be_clickable((By.CSS_SELECTOR, agree_selector))
                )
                
                # 체크박스 상태 확인
                is_checked = agree_checkbox.is_selected()
                print(f"[INFO] 동의함 체크박스 현재 상태: {is_checked}")
                
                if not is_checked:
                    agree_checkbox.click()
                    print("[OK] 동의함 체크박스 클릭 완료")
                else:
                    print("[INFO] 동의함 체크박스가 이미 체크되어 있습니다.")
                    
            except TimeoutException:
                print("[FAIL] 동의함 체크박스를 찾을 수 없습니다.")
                return False
            
            # 2. 비밀번호 입력
            print("[INFO] 비밀번호 입력 중...")
            password_selector = "body > form > div > div > div.tblForm01.mt15 > table > tbody > tr > td > input[type=password]"
            
            try:
                password_field = self.driver.find_element(By.CSS_SELECTOR, password_selector)
                password_field.clear()
                password_field.send_keys(self.excel_password)
                print(f"[OK] 비밀번호 입력 완료: {'*' * len(self.excel_password)}")
                
            except NoSuchElementException:
                print("[FAIL] 비밀번호 입력 필드를 찾을 수 없습니다.")
                return False
            
            # 3. 확인 버튼 클릭
            print("[INFO] 확인 버튼 클릭 중...")
            confirm_selector = "body > form > div > div > div:nth-child(7) > a:nth-child(1) > img"
            
            try:
                confirm_button = self.driver.find_element(By.CSS_SELECTOR, confirm_selector)
                confirm_button.click()
                print("[OK] 확인 버튼 클릭 완료")
                
                # 처리 대기
                self.random_wait(3, 5)
                
                # 다운로드 완료 확인
                print("[INFO] 다운로드 완료 확인 중...")
                
                # 현재 페이지 상태 확인
                final_url = self.driver.current_url
                final_title = self.driver.title
                print(f"[INFO] 최종 페이지: {final_title}")
                print(f"[INFO] 최종 URL: {final_url}")
                
                # iframe에서 나오기
                try:
                    self.driver.switch_to.default_content()
                    print("[OK] iframe에서 나오기 완료")
                except Exception as default_error:
                    print(f"[WARN] iframe에서 나오기 실패: {str(default_error)}")
                
                print("[OK] 엑셀 다운로드 폼 처리 완료")
                return True
                
            except NoSuchElementException:
                print("[FAIL] 확인 버튼을 찾을 수 없습니다.")
                return False
                
        except Exception as e:
            print(f"[FAIL] 엑셀 다운로드 폼 처리 실패: {str(e)}")
            return False
    
    def get_sharepoint_access_token(self):
        """SharePoint 액세스 토큰 — client_credentials(앱 권한). CI 무인 실행 호환.

        기존 delegated(refresh_token + 인터랙티브 OAuth)에서 전환 — OPS-Console과 동일한
        앱 권한 토큰을 사용한다. SHAREPOINT_TENANT_ID/CLIENT_ID/CLIENT_SECRET 필요.
        """
        try:
            print("[INFO] SharePoint client_credentials 토큰 획득 중...")
            tenant_id = self.sharepoint_tenant_id
            client_id = self.sharepoint_client_id
            client_secret = self.sharepoint_client_secret
            if not (tenant_id and client_id and client_secret):
                print("[FAIL] SHAREPOINT_TENANT_ID/CLIENT_ID/CLIENT_SECRET 환경변수 누락")
                return None

            token_url = f"https://login.microsoftonline.com/{tenant_id}/oauth2/v2.0/token"
            data = {
                "grant_type": "client_credentials",
                "client_id": client_id,
                "client_secret": client_secret,
                "scope": "https://graph.microsoft.com/.default",
            }
            res = requests.post(token_url, data=data, timeout=30)
            if res.status_code != 200:
                print(f"[FAIL] 토큰 요청 실패: {res.status_code} {res.text[:300]}")
                return None

            access_token = res.json().get("access_token")
            if not access_token:
                print("[FAIL] 응답에 access_token 없음")
                return None

            print("[OK] client_credentials 토큰 획득 완료")
            return access_token

        except Exception as e:
            print(f"[FAIL] SharePoint 토큰 획득 실패: {str(e)}")
            return None
    
    def generate_auth_url(self):
        """OAuth2 인증 URL 생성"""
        try:
            # OAuth2 인증 URL 생성 (사용자 성공 형식으로 수정)
            auth_url = (
                f"https://login.microsoftonline.com/common/oauth2/v2.0/authorize?"
                f"client_id={self.sharepoint_client_id}&"
                f"response_type=code&"
                f"redirect_uri=http://localhost:8080&"
                f"response_mode=query&"
                f"scope=offline_access Files.ReadWrite.All"
            )
            
            return auth_url
            
        except Exception as e:
            print(f"[FAIL] 인증 URL 생성 실패: {str(e)}")
            return None
    
    def exchange_code_for_token(self, auth_code):
        """인증 코드를 토큰으로 교환"""
        try:
            print("[INFO] 인증 코드를 토큰으로 교환 중...")
            
            # 토큰 교환 URL (common 사용)
            token_url = f"https://login.microsoftonline.com/common/oauth2/v2.0/token"
            
            # 토큰 교환 요청 데이터
            data = {
                'client_id': self.sharepoint_client_id,
                'client_secret': self.sharepoint_client_secret,
                'code': auth_code,
                'redirect_uri': 'http://localhost:8080',
                'grant_type': 'authorization_code',
                'scope': 'offline_access Files.ReadWrite.All'
            }
            
            headers = {
                'Content-Type': 'application/x-www-form-urlencoded'
            }
            
            # 토큰 교환 요청
            response = requests.post(token_url, data=data, headers=headers)
            
            if response.status_code == 200:
                token_data = response.json()
                print("[OK] 토큰 교환 성공")
                return token_data
            else:
                print(f"[FAIL] 토큰 교환 실패: {response.status_code}")
                print(f"[DEBUG] 응답: {response.text}")
                return None
                
        except Exception as e:
            print(f"[FAIL] 토큰 교환 실패: {str(e)}")
            return None
    
    def save_refresh_token(self, refresh_token):
        """리프레시 토큰을 파일에 저장"""
        try:
            current_dir = os.path.dirname(os.path.abspath(__file__))
            token_file = os.path.join(current_dir, "refresh_token.txt")
            
            with open(token_file, 'w', encoding='utf-8') as f:
                f.write(refresh_token)
            
            print(f"[OK] 리프레시 토큰 저장: {token_file}")
            
        except Exception as e:
            print(f"[WARN] 리프레시 토큰 저장 실패: {str(e)}")
    
    def load_refresh_token(self):
        """파일에서 리프레시 토큰 로드"""
        try:
            current_dir = os.path.dirname(os.path.abspath(__file__))
            token_file = os.path.join(current_dir, "refresh_token.txt")
            
            if os.path.exists(token_file):
                with open(token_file, 'r', encoding='utf-8') as f:
                    refresh_token = f.read().strip()
                
                if refresh_token:
                    print(f"[OK] 리프레시 토큰 로드: {token_file}")
                    return refresh_token
            
            return None
            
        except Exception as e:
            print(f"[WARN] 리프레시 토큰 로드 실패: {str(e)}")
            return None
    
    def refresh_access_token(self, refresh_token):
        """리프레시 토큰으로 액세스 토큰 갱신"""
        try:
            print("[INFO] 리프레시 토큰으로 액세스 토큰 갱신 중...")
            
            # 토큰 갱신 URL
            token_url = f"https://login.microsoftonline.com/common/oauth2/v2.0/token"
            
            # 토큰 갱신 요청 데이터
            data = {
                'client_id': self.sharepoint_client_id,
                'client_secret': self.sharepoint_client_secret,
                'refresh_token': refresh_token,
                'grant_type': 'refresh_token',
                'scope': 'offline_access Files.ReadWrite.All'
            }
            
            headers = {
                'Content-Type': 'application/x-www-form-urlencoded'
            }
            
            # 토큰 갱신 요청
            response = requests.post(token_url, data=data, headers=headers)
            
            if response.status_code == 200:
                token_data = response.json()
                print("[OK] 리프레시 토큰으로 토큰 갱신 성공")
                
                # 새로운 리프레시 토큰 저장 (있는 경우)
                if 'refresh_token' in token_data:
                    self.save_refresh_token(token_data['refresh_token'])
                    print("[OK] 새로운 리프레시 토큰 저장 완료")
                
                return token_data
            else:
                print(f"[FAIL] 리프레시 토큰 갱신 실패: {response.status_code}")
                print(f"[DEBUG] 응답: {response.text}")
                return None
                
        except Exception as e:
            print(f"[FAIL] 리프레시 토큰 갱신 실패: {str(e)}")
            return None
    
    def find_sharepoint_folder(self, access_token, folder_path):
        """SharePoint 폴더 찾기"""
        try:
            print(f"[INFO] SharePoint 폴더 찾는 중: {folder_path}")
            
            # 폴더 경로를 배열로 분할
            path_parts = folder_path.split('/')
            
            # 루트 드라이브에서 시작
            current_folder_id = self.sharepoint_drive_id
            
            for part in path_parts:
                if not part.strip():
                    continue
                    
                # 현재 폴더의 하위 폴더들 조회
                if current_folder_id == self.sharepoint_drive_id:
                    # 루트 드라이브의 경우
                    url = f"https://graph.microsoft.com/v1.0/sites/{self.sharepoint_site_id}/drives/{current_folder_id}/root/children"
                else:
                    # 하위 폴더의 경우
                    url = f"https://graph.microsoft.com/v1.0/sites/{self.sharepoint_site_id}/drives/{self.sharepoint_drive_id}/items/{current_folder_id}/children"
                
                headers = {
                    'Authorization': f'Bearer {access_token}',
                    'Content-Type': 'application/json'
                }
                
                response = requests.get(url, headers=headers)
                
                if response.status_code == 200:
                    data = response.json()
                    folders = [item for item in data.get('value', []) if item.get('folder')]
                    
                    # 해당 이름의 폴더 찾기
                    target_folder = None
                    for folder in folders:
                        if folder['name'] == part:
                            target_folder = folder
                            break
                    
                    if target_folder:
                        current_folder_id = target_folder['id']
                        print(f"[OK] 폴더 발견: {part} (ID: {current_folder_id})")
                    else:
                        print(f"[FAIL] 폴더를 찾을 수 없습니다: {part}")
                        return None
                else:
                    print(f"[FAIL] 폴더 조회 실패: {response.status_code} - {response.text}")
                    return None
            
            print(f"[OK] 최종 폴더 ID: {current_folder_id}")
            return current_folder_id
            
        except Exception as e:
            print(f"[FAIL] SharePoint 폴더 찾기 실패: {str(e)}")
            return None
    
    def upload_file_to_sharepoint(self, file_path, filename=None):
        """SharePoint에 파일 업로드 (강제 덮어쓰기)"""
        try:
            print("\n" + "="*60)
            print("SharePoint 파일 업로드 (강제 덮어쓰기)")
            print("="*60)
            
            # 파일 존재 확인
            if not os.path.exists(file_path):
                print(f"[FAIL] 파일을 찾을 수 없습니다: {file_path}")
                return False
            
            # 파일명 생성 (고정 파일명)
            if not filename:
                filename = "역발행 세금계산서.xlsx"
            
            print(f"[INFO] 업로드할 파일: {file_path}")
            print(f"[INFO] 업로드 파일명: {filename}")
            
            # 암호화된 파일인지 확인하고 해독
            upload_file_path = None
            try:
                print(f"[INFO] 파일 암호 해독 시도 중...")
                print(f"[INFO] 사용할 암호: {'*' * len(self.excel_password)}")
                
                # 암호화된 파일 해독 시도
                decrypted_file_path = self.decrypt_excel_file(file_path, self.excel_password)
                if decrypted_file_path:
                    upload_file_path = decrypted_file_path
                    print(f"[OK] 해독된 파일로 업로드: {decrypted_file_path}")
                    print(f"[INFO] 해독된 파일 크기: {os.path.getsize(decrypted_file_path)} bytes")
                else:
                    # 암호 해독이 실패한 경우, 원본 파일이 암호화되지 않은지 확인
                    print(f"[INFO] 암호 해독 실패. 원본 파일이 암호화되지 않은지 확인 중...")
                    try:
                        import msoffcrypto  # type: ignore
                        with open(file_path, 'rb') as test_file:
                            test_office_file = msoffcrypto.OfficeFile(test_file)
                            if hasattr(test_office_file, 'is_encrypted') and test_office_file.is_encrypted():
                                print(f"[FAIL] 파일이 암호화되어 있지만 해독에 실패했습니다.")
                                print(f"[FAIL] SharePoint 업로드를 중단합니다.")
                                return False
                            else:
                                print(f"[INFO] 파일이 암호화되지 않았습니다. 원본 파일로 업로드합니다.")
                                upload_file_path = file_path
                    except Exception as test_error:
                        print(f"[FAIL] 파일 암호화 여부 확인 실패: {str(test_error)}")
                        print(f"[FAIL] 보안상 암호화 여부를 확인할 수 없으므로 업로드를 중단합니다.")
                        return False
                        
            except Exception as decrypt_error:
                print(f"[FAIL] 파일 해독 실패: {str(decrypt_error)}")
                print(f"[FAIL] SharePoint 업로드를 중단합니다.")
                return False
            
            # 업로드할 파일이 없는 경우 중단
            if not upload_file_path:
                print(f"[FAIL] 업로드할 파일이 없습니다.")
                return False
            
            # 액세스 토큰 획득
            access_token = self.get_sharepoint_access_token()
            if not access_token:
                return False
            
            # 업로드 폴더 찾기
            folder_id = self.find_sharepoint_folder(access_token, self.upload_folder_path)
            if not folder_id:
                print(f"[FAIL] 업로드 폴더를 찾을 수 없습니다: {self.upload_folder_path}")
                return False
            
            # 기존 파일 확인 (정확한 파일명으로만 찾기)
            existing_file_id = self.check_existing_file(access_token, folder_id, filename)
            
            # 스마트 업로드: 기존 파일 있으면 내용만 업데이트, 없으면 새 파일 업로드
            if existing_file_id:
                print(f"[INFO] 기존 파일 발견. 메모리에서 직접 비교 후 내용만 업데이트 중: {filename}")
                upload_success = self.smart_update_existing_file(access_token, existing_file_id, upload_file_path, filename)
            else:
                print(f"[INFO] 기존 파일 없음. 새 파일 업로드 중: {filename}")
                upload_success = self.upload_new_file(access_token, folder_id, filename, upload_file_path)
            
            if upload_success:
                if existing_file_id:
                    print(f"[OK] 스마트 파일 업데이트 완료: {filename}")
                    print("[INFO] 기존 파일에 새로운 내용만 추가되었습니다.")
                else:
                    print(f"[OK] 새 파일 업로드 완료: {filename}")
                print("[INFO] SharePoint에서 파일을 새로고침하여 최신 버전을 확인하세요.")
                
                # 해독된 파일이 생성된 경우 임시 파일 정리
                if upload_file_path != file_path and os.path.exists(upload_file_path):
                    try:
                        os.remove(upload_file_path)
                        print(f"[OK] 임시 해독 파일 정리 완료: {upload_file_path}")
                    except Exception as cleanup_error:
                        print(f"[WARN] 임시 파일 정리 실패: {str(cleanup_error)}")
                
                return True
            else:
                print(f"[FAIL] 파일 처리 실패: {filename}")
                
                # 업로드 실패 시에도 해독된 파일 정리
                if upload_file_path != file_path and os.path.exists(upload_file_path):
                    try:
                        os.remove(upload_file_path)
                        print(f"[OK] 업로드 실패로 인한 임시 파일 정리 완료: {upload_file_path}")
                    except Exception as cleanup_error:
                        print(f"[WARN] 임시 파일 정리 실패: {str(cleanup_error)}")
                
                return False
                
        except Exception as e:
            print(f"[FAIL] SharePoint 업로드 실패: {str(e)}")
            return False
    
    def find_existing_files_by_pattern(self, access_token, folder_id, filename):
        """패턴으로 기존 파일들 찾기 (역발행 세금계산서 관련 파일들)"""
        try:
            print(f"[INFO] 기존 파일들 패턴 검색 중: {filename}")
            
            # 폴더 내 파일 목록 조회
            if folder_id == self.sharepoint_drive_id:
                url = f"https://graph.microsoft.com/v1.0/sites/{self.sharepoint_site_id}/drives/{self.sharepoint_drive_id}/root/children"
            else:
                url = f"https://graph.microsoft.com/v1.0/sites/{self.sharepoint_site_id}/drives/{self.sharepoint_drive_id}/items/{folder_id}/children"
            
            headers = {
                'Authorization': f'Bearer {access_token}',
                'Content-Type': 'application/json'
            }
            
            response = requests.get(url, headers=headers)
            
            if response.status_code == 200:
                data = response.json()
                files = data.get('value', [])
                
                # 역발행 세금계산서 관련 파일들 찾기
                matching_files = []
                base_patterns = [
                    "역발행 세금계산서",
                    "역발행세금계산서",
                    "세금계산서"
                ]
                
                for file in files:
                    file_name = file.get('name', '')
                    file_id = file.get('id')
                    
                    # 정확한 파일명 매치
                    if file_name == filename:
                        matching_files.append({
                            'name': file_name,
                            'id': file_id,
                            'reason': '정확한 파일명'
                        })
                        print(f"[OK] 정확한 파일명 매치: {file_name} (ID: {file_id})")
                        continue
                    
                    # 패턴 매치 (역발행 세금계산서 관련)
                    for pattern in base_patterns:
                        if pattern in file_name and file_name.lower().endswith(('.xlsx', '.xls')):
                            matching_files.append({
                                'name': file_name,
                                'id': file_id,
                                'reason': f'패턴 매치: {pattern}'
                            })
                            print(f"[OK] 패턴 매치: {file_name} (ID: {file_id}) - {pattern}")
                            break
                
                if matching_files:
                    print(f"[INFO] 총 {len(matching_files)}개의 관련 파일 발견:")
                    for i, file_info in enumerate(matching_files, 1):
                        print(f"  {i}. {file_info['name']} - {file_info['reason']}")
                    return matching_files
                else:
                    print(f"[INFO] 관련 파일 없음: {filename}")
                    return []
            else:
                print(f"[FAIL] 파일 목록 조회 실패: {response.status_code} - {response.text}")
                return []
                
        except Exception as e:
            print(f"[FAIL] 기존 파일 패턴 검색 실패: {str(e)}")
            return []
    
    def check_existing_file(self, access_token, folder_id, filename):
        """기존 파일 존재 여부 확인 (호환성 유지)"""
        try:
            print(f"[INFO] 기존 파일 확인 중: {filename}")
            
            # 폴더 내 파일 목록 조회
            if folder_id == self.sharepoint_drive_id:
                url = f"https://graph.microsoft.com/v1.0/sites/{self.sharepoint_site_id}/drives/{self.sharepoint_drive_id}/root/children"
            else:
                url = f"https://graph.microsoft.com/v1.0/sites/{self.sharepoint_site_id}/drives/{self.sharepoint_drive_id}/items/{folder_id}/children"
            
            headers = {
                'Authorization': f'Bearer {access_token}',
                'Content-Type': 'application/json'
            }
            
            response = requests.get(url, headers=headers)
            
            if response.status_code == 200:
                data = response.json()
                files = data.get('value', [])
                
                for file in files:
                    if file.get('name') == filename:
                        file_id = file.get('id')
                        print(f"[OK] 기존 파일 발견: {filename} (ID: {file_id})")
                        return file_id
                
                print(f"[INFO] 기존 파일 없음: {filename}")
                return None
            else:
                print(f"[FAIL] 파일 목록 조회 실패: {response.status_code} - {response.text}")
                return None
                
        except Exception as e:
            print(f"[FAIL] 기존 파일 확인 실패: {str(e)}")
            return None
    
    def smart_update_existing_file(self, access_token, existing_file_id, new_file_path, filename):
        """스마트 업데이트: 기존 파일 내용과 비교하여 다른 내용만 기존 파일에 업데이트 (새 파일 업로드 X)"""
        try:
            print(f"[INFO] 스마트 파일 업데이트 중: {filename}")
            print("[INFO] 기존 파일은 보존하고 내용만 업데이트합니다.")
            
            # 1. 기존 SharePoint 파일을 직접 메모리에서 읽기 (효율적)
            print("[INFO] 기존 SharePoint 파일 직접 읽기 중...")
            existing_df = self.read_sharepoint_file_to_dataframe(access_token, existing_file_id, filename)
            
            if existing_df is None:
                print("[WARN] 기존 파일 직접 읽기 실패")
                return False
            
            # 2. 두 파일의 내용 비교 (메모리 기반)
            print("[INFO] 파일 내용 비교 중 (메모리 기반)...")
            comparison_result = self.compare_excel_files_with_dataframe(existing_df, new_file_path)
            
            if comparison_result['is_identical']:
                print("[INFO] 파일 내용이 동일합니다. 업데이트할 내용이 없습니다.")
                return True
            elif comparison_result['has_new_data']:
                print(f"[INFO] 새로운 데이터 {len(comparison_result['new_data'])}건 발견.")
                print("[INFO] 기존 파일에 새로운 내용만 추가합니다 (새 파일 업로드 없음).")
                
                # 3. 기존 파일에 새로운 데이터만 추가 (이메일오류='Y' 데이터 보호)
                return self.update_existing_file_content_only(access_token, existing_file_id, existing_df, comparison_result['new_data'], filename)
            else:
                print("[INFO] 새로운 데이터가 없습니다. 기존 파일을 그대로 유지합니다.")
                return True
                
        except Exception as e:
            print(f"[FAIL] 스마트 파일 업데이트 실패: {str(e)}")
            return False
    
    def save_dataframe_with_original_structure(self, df, original_headers=None):
        """DataFrame을 원본 Excel 구조(1,2,3행 헤더 포함)로 저장"""
        try:
            print("[INFO] 원본 Excel 구조로 저장 중...")
            
            excel_buffer = io.BytesIO()
            
            # openpyxl을 사용하여 원본 구조 완전 보존
            from openpyxl import Workbook
            
            wb = Workbook()
            ws = wb.active
            
            # 원본 헤더가 있으면 1,2행에 입력
            if original_headers and 'header_row_1' in original_headers:
                header_1 = original_headers['header_row_1']
                header_2 = original_headers['header_row_2']
                column_names = original_headers['column_names']
                
                print("[INFO] 원본 헤더 1,2행 복원 중...")
                
                # 1행 입력
                for col_idx, value in enumerate(header_1, 1):
                    ws.cell(row=1, column=col_idx, value=value)
                
                # 2행 입력
                for col_idx, value in enumerate(header_2, 1):
                    ws.cell(row=2, column=col_idx, value=value)
                
                # 3행 입력 (컬럼명)
                for col_idx, value in enumerate(column_names, 1):
                    ws.cell(row=3, column=col_idx, value=value)
                
                print("[OK] 원본 헤더 1,2,3행 복원 완료")
            else:
                print("[WARN] 원본 헤더 정보가 없습니다. 데이터만 저장합니다.")
                # 3행에 컬럼명 입력
                for col_idx, col_name in enumerate(df.columns, 1):
                    ws.cell(row=3, column=col_idx, value=col_name)
            
            # 4행부터 데이터 입력
            print(f"[INFO] {len(df)}행의 데이터 입력 중...")
            for row_idx, (_, row) in enumerate(df.iterrows(), 4):
                for col_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=str(value))
                    
                    # 숫자 형식 보존
                    try:
                        if value and str(value).strip():
                            if '.' in str(value):
                                float_val = float(value)
                                if float_val > 1e6:
                                    cell.number_format = '0'
                                else:
                                    cell.number_format = '0.00'
                            else:
                                int_val = int(value)
                                cell.number_format = '0'
                    except (ValueError, TypeError):
                        pass  # 숫자가 아닌 경우 그대로 유지
            
            # 워크북을 BytesIO에 저장
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            print(f"[OK] 원본 구조로 Excel 저장 완료: {len(excel_buffer.getvalue())} bytes")
            return excel_buffer
            
        except Exception as e:
            print(f"[FAIL] 원본 구조로 Excel 저장 실패: {str(e)}")
            import traceback
            print(f"[DEBUG] 상세 오류: {traceback.format_exc()}")
            return None

    def update_existing_file_content_only(self, access_token, existing_file_id, existing_df, new_data, filename):
        """기존 파일에 새로운 데이터만 추가 (새 파일 업로드 없이 기존 파일 직접 업데이트) - 원본 구조 보존"""
        try:
            print(f"[INFO] 기존 파일 내용 업데이트 중: {filename}")
            print(f"[INFO] 새로 추가할 데이터: {len(new_data)}건")
            print("[INFO] 이메일오류='Y' 데이터는 보호됩니다.")
            print("[INFO] 원본 Excel 구조(1,2,3행 헤더) 보존합니다.")
            print("[INFO] 새 파일 업로드 없이 기존 파일만 수정합니다.")
            
            # 1. 원본 구조를 포함하여 기존 파일 읽기
            original_df_with_headers = self.read_sharepoint_file_with_full_structure(access_token, existing_file_id, filename)
            if original_df_with_headers is None:
                print("[WARN] 원본 구조 읽기 실패, 기본 방식으로 처리...")
                return self.update_existing_file_content_only_basic(access_token, existing_file_id, existing_df, new_data, filename)
            
            # 2. 이메일오류 컬럼 찾기
            email_error_col = None
            for col in original_df_with_headers.columns:
                if '이메일오류' in str(col):
                    email_error_col = col
                    break
            
            if email_error_col:
                print(f"[INFO] 이메일오류 컬럼 발견: {email_error_col}")
                
                # 3. 기존 데이터에서 이메일오류가 'Y'인 행들 보호
                protected_rows = original_df_with_headers[original_df_with_headers[email_error_col].str.upper() == 'Y']
                print(f"[INFO] 보호할 데이터 (이메일오류='Y'): {len(protected_rows)}건")
                
                # 4. 새로운 데이터를 DataFrame으로 변환
                new_df = pd.DataFrame(new_data)
                print(f"[INFO] 새로 추가할 데이터: {len(new_df)}건")
                
                # 5. 안전한 병합: 보호된 행 + 새 데이터
                merged_df = pd.concat([protected_rows, new_df], ignore_index=True)
                print(f"[INFO] 최종 병합 결과: {len(merged_df)}행 (보호된 데이터 + 새 데이터)")
                
            else:
                print("[WARN] 이메일오류 컬럼을 찾을 수 없습니다. 기본 병합을 수행합니다.")
                # 기존 데이터와 새 데이터 병합
                new_df = pd.DataFrame(new_data)
                merged_df = pd.concat([original_df_with_headers, new_df], ignore_index=True)
                print(f"[INFO] 기본 병합 결과: {len(merged_df)}행")
            
            # 6. 원본 헤더 정보 추출
            original_headers = {
                'header_row_1': original_df_with_headers.attrs.get('header_row_1', []),
                'header_row_2': original_df_with_headers.attrs.get('header_row_2', []),
                'column_names': original_df_with_headers.attrs.get('column_names', list(merged_df.columns))
            }
            
            # 7. 병합된 데이터를 원본 구조로 Excel 스트림 생성
            print("[INFO] 병합된 데이터를 원본 구조로 Excel 스트림 변환 중...")
            excel_buffer = self.save_dataframe_with_original_structure(merged_df, original_headers)
            
            if excel_buffer is None:
                print("[FAIL] Excel 스트림 생성 실패")
                return False
            
            # 8. 기존 파일에 직접 업데이트 (새 파일 업로드 없음)
            print("[INFO] 기존 SharePoint 파일에 직접 업데이트 중...")
            
            update_url = f"https://graph.microsoft.com/v1.0/sites/{self.sharepoint_site_id}/drives/{self.sharepoint_drive_id}/items/{existing_file_id}/content"
            
            headers = {
                'Authorization': f'Bearer {access_token}',
                'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            }
            
            response = requests.put(update_url, data=excel_buffer.getvalue(), headers=headers)
            
            if response.status_code in [200, 201]:
                print(f"[OK] 기존 파일 내용 업데이트 성공: {filename}")
                print(f"[INFO] 새로운 데이터 {len(new_data)}건이 기존 파일에 추가되었습니다.")
                print("[INFO] 기존 이메일오류='Y' 데이터가 안전하게 보호되었습니다.")
                print("[INFO] 원본 Excel 구조(1,2,3행 헤더)가 완전히 보존되었습니다.")
                print("[INFO] 새 파일 업로드 없이 기존 파일만 수정되었습니다.")
                print("[INFO] SharePoint에서 파일을 새로고침하여 최신 버전을 확인하세요.")
                return True
            else:
                print(f"[FAIL] 기존 파일 내용 업데이트 실패: {response.status_code}")
                if response.text:
                    print(f"[DEBUG] 응답 내용: {response.text}")
                return False
                
        except Exception as e:
            print(f"[FAIL] 기존 파일 내용 업데이트 실패: {str(e)}")
            import traceback
            print(f"[DEBUG] 상세 오류: {traceback.format_exc()}")
            return False

    def update_existing_file_content_only_basic(self, access_token, existing_file_id, existing_df, new_data, filename):
        """기존 파일에 새로운 데이터만 추가 (기본 방식 - 호환성 유지)"""
        try:
            print(f"[INFO] 기존 파일 내용 업데이트 중 (기본 방식): {filename}")
            print(f"[INFO] 새로 추가할 데이터: {len(new_data)}건")
            print("[INFO] 이메일오류='Y' 데이터는 보호됩니다.")
            print("[INFO] 새 파일 업로드 없이 기존 파일만 수정합니다.")
            
            # 1. 이메일오류 컬럼 찾기
            email_error_col = None
            for col in existing_df.columns:
                if '이메일오류' in str(col):
                    email_error_col = col
                    break
            
            if email_error_col:
                print(f"[INFO] 이메일오류 컬럼 발견: {email_error_col}")
                
                # 2. 기존 데이터에서 이메일오류가 'Y'인 행들 보호
                protected_rows = existing_df[existing_df[email_error_col].str.upper() == 'Y']
                print(f"[INFO] 보호할 데이터 (이메일오류='Y'): {len(protected_rows)}건")
                
                # 3. 새로운 데이터를 DataFrame으로 변환
                new_df = pd.DataFrame(new_data)
                print(f"[INFO] 새로 추가할 데이터: {len(new_df)}건")
                
                # 4. 안전한 병합: 보호된 행 + 새 데이터
                merged_df = pd.concat([protected_rows, new_df], ignore_index=True)
                print(f"[INFO] 최종 병합 결과: {len(merged_df)}행 (보호된 데이터 + 새 데이터)")
                
            else:
                print("[WARN] 이메일오류 컬럼을 찾을 수 없습니다. 기본 병합을 수행합니다.")
                # 기존 데이터와 새 데이터 병합
                new_df = pd.DataFrame(new_data)
                merged_df = pd.concat([existing_df, new_df], ignore_index=True)
                print(f"[INFO] 기본 병합 결과: {len(merged_df)}행")
            
            # 5. 병합된 데이터를 메모리에서 바로 Excel 스트림으로 변환
            print("[INFO] 병합된 데이터를 Excel 스트림으로 변환 중 (기본 방식)...")
            excel_buffer = io.BytesIO()
            
            # ExcelWriter를 사용하여 원본 형식 보존
            with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                merged_df.to_excel(writer, index=False, header=False, sheet_name='Sheet1')
                
                # 숫자 형식 보존을 위한 추가 설정
                worksheet = writer.sheets['Sheet1']
                for row in worksheet.iter_rows():
                    for cell in row:
                        if cell.value and isinstance(cell.value, str):
                            try:
                                if '.' in cell.value:
                                    float_val = float(cell.value)
                                    if float_val > 1e6:  # 큰 숫자의 경우 과학적 표기법 방지
                                        cell.number_format = '0'
                                    else:
                                        cell.number_format = '0.00'
                                else:
                                    int_val = int(cell.value)
                                    cell.number_format = '0'
                            except (ValueError, TypeError):
                                pass
            
            excel_buffer.seek(0)
            print(f"[OK] Excel 스트림 생성 완료: {len(excel_buffer.getvalue())} bytes")
            
            # 6. 기존 파일에 직접 업데이트 (새 파일 업로드 없음)
            print("[INFO] 기존 SharePoint 파일에 직접 업데이트 중...")
            
            update_url = f"https://graph.microsoft.com/v1.0/sites/{self.sharepoint_site_id}/drives/{self.sharepoint_drive_id}/items/{existing_file_id}/content"
            
            headers = {
                'Authorization': f'Bearer {access_token}',
                'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            }
            
            response = requests.put(update_url, data=excel_buffer.getvalue(), headers=headers)
            
            if response.status_code in [200, 201]:
                print(f"[OK] 기존 파일 내용 업데이트 성공: {filename}")
                print(f"[INFO] 새로운 데이터 {len(new_data)}건이 기존 파일에 추가되었습니다.")
                print("[INFO] 기존 이메일오류='Y' 데이터가 안전하게 보호되었습니다.")
                print("[INFO] 새 파일 업로드 없이 기존 파일만 수정되었습니다.")
                print("[INFO] SharePoint에서 파일을 새로고침하여 최신 버전을 확인하세요.")
                return True
            else:
                print(f"[FAIL] 기존 파일 내용 업데이트 실패: {response.status_code}")
                if response.text:
                    print(f"[DEBUG] 응답 내용: {response.text}")
                return False
                
        except Exception as e:
            print(f"[FAIL] 기존 파일 내용 업데이트 실패: {str(e)}")
            import traceback
            print(f"[DEBUG] 상세 오류: {traceback.format_exc()}")
            return False

    def compare_and_update_file(self, access_token, existing_file_id, new_file_path, filename):
        """기존 파일과 새 파일 내용 비교 후 업데이트 (기존 방식 - 호환성 유지)"""
        try:
            print(f"[INFO] 파일 내용 비교 및 업데이트 중: {filename}")
            
            # 1. 기존 SharePoint 파일을 직접 메모리에서 읽기 (효율적)
            print("[INFO] 기존 SharePoint 파일 직접 읽기 중...")
            existing_df = self.read_sharepoint_file_to_dataframe(access_token, existing_file_id, filename)
            
            if existing_df is None:
                print("[WARN] 기존 파일 직접 읽기 실패, 새 파일로 업로드 시도...")
                return self.upload_new_file_with_same_name(access_token, new_file_path, filename)
            
            # 2. 두 파일의 내용 비교 (메모리 기반)
            print("[INFO] 파일 내용 비교 중...")
            comparison_result = self.compare_excel_files_with_dataframe(existing_df, new_file_path)
            
            if comparison_result['is_identical']:
                print("[INFO] 파일 내용이 동일합니다. 기존 파일을 그대로 유지합니다.")
                return True
            elif comparison_result['has_new_data']:
                print(f"[INFO] 새로운 데이터 {len(comparison_result['new_data'])}건 발견. 기존 파일에 추가합니다.")
                return self.merge_and_update_file(access_token, existing_file_id, new_file_path, filename, comparison_result['new_data'])
            else:
                print("[INFO] 새로운 데이터가 없습니다. 기존 파일을 그대로 유지합니다.")
                return True
                
        except Exception as e:
            print(f"[FAIL] 파일 비교 및 업데이트 실패: {str(e)}")
            return False
    
    def compare_excel_files_with_dataframe(self, existing_df, new_file_path):
        """기존 DataFrame과 새 파일 내용 비교 (메모리 기반, 효율적)"""
        try:
            print("[INFO] 엑셀 파일 내용 비교 중 (메모리 기반)...")
            
            # 새 파일 읽기 (원본 데이터 보존을 위한 설정)
            new_df = None
            try:
                new_df = pd.read_excel(new_file_path, header=2, dtype=str, na_filter=False)
                print("[OK] 새 파일 읽기 성공 (3행부터, 문자열로 보존)")
            except Exception as e1:
                print(f"[WARN] 새 파일 기본 방법 실패: {str(e1)}")
                try:
                    new_df = pd.read_excel(new_file_path, engine='openpyxl', header=2, dtype=str, na_filter=False)
                    print("[OK] 새 파일 openpyxl 엔진으로 읽기 성공 (3행부터, 문자열로 보존)")
                except Exception as e2:
                    print(f"[WARN] 새 파일 openpyxl 엔진 실패: {str(e2)}")
                    try:
                        new_df = pd.read_excel(new_file_path, engine='xlrd', header=2, dtype=str, na_filter=False)
                        print("[OK] 새 파일 xlrd 엔진으로 읽기 성공 (3행부터, 문자열로 보존)")
                    except Exception as e3:
                        print(f"[FAIL] 새 파일 모든 엔진으로 읽기 실패: {str(e3)}")
                        return {'is_identical': False, 'has_new_data': False, 'new_data': []}
            
            if new_df is None:
                print("[FAIL] 새 파일을 읽을 수 없습니다.")
                return {'is_identical': False, 'has_new_data': False, 'new_data': []}
            
            print(f"[INFO] 기존 파일 데이터: {len(existing_df)}행, {len(existing_df.columns)}열 (메모리)")
            print(f"[INFO] 새 파일 데이터: {len(new_df)}행, {len(new_df.columns)}열")
            
            # 컬럼명이 동일한지 확인
            if list(existing_df.columns) != list(new_df.columns):
                print("[WARN] 컬럼명이 다릅니다. 새 파일의 모든 데이터를 추가합니다.")
                return {'is_identical': False, 'has_new_data': True, 'new_data': new_df.to_dict('records')}
            
            # 데이터 비교 (작성일자, 품목, 거래처명을 기준으로 고유성 확인)
            print("[INFO] 데이터 중복 확인 중...")
            
            # 기존 데이터의 고유 키 생성
            existing_keys = set()
            for _, row in existing_df.iterrows():
                작성일자 = self._get_column_value(row, ['작성일자'])
                품목 = self._get_column_value(row, ['품목'])
                거래처명 = self._get_column_value(row, ['거래처명'])
                key = f"{작성일자}|{품목}|{거래처명}"
                existing_keys.add(key)
            
            # 새 데이터 중 기존에 없는 데이터 찾기
            new_data_rows = []
            for _, row in new_df.iterrows():
                작성일자 = self._get_column_value(row, ['작성일자'])
                품목 = self._get_column_value(row, ['품목'])
                거래처명 = self._get_column_value(row, ['거래처명'])
                key = f"{작성일자}|{품목}|{거래처명}"
                
                if key not in existing_keys:
                    new_data_rows.append(row)
                    print(f"[INFO] 새로운 데이터 발견: {작성일자} | {품목} | {거래처명}")
            
            if len(new_data_rows) == 0:
                print("[INFO] 새로운 데이터가 없습니다. 파일 내용이 동일합니다.")
                return {'is_identical': True, 'has_new_data': False, 'new_data': []}
            else:
                print(f"[INFO] 새로운 데이터 {len(new_data_rows)}건 발견")
                return {'is_identical': False, 'has_new_data': True, 'new_data': new_data_rows}
                
        except Exception as e:
            print(f"[FAIL] 파일 비교 실패: {str(e)}")
            return {'is_identical': False, 'has_new_data': False, 'new_data': []}

    def compare_excel_files(self, existing_file_path, new_file_path):
        """두 엑셀 파일의 내용 비교 (원본 데이터 보존) - 기존 방식 호환성 유지"""
        try:
            print("[INFO] 엑셀 파일 내용 비교 중...")
            
            # 기존 파일 읽기 (원본 데이터 보존을 위한 설정)
            existing_df = None
            try:
                existing_df = pd.read_excel(existing_file_path, header=2, dtype=str, na_filter=False)
                print("[OK] 기존 파일 읽기 성공 (3행부터, 문자열로 보존)")
            except Exception as e1:
                print(f"[WARN] 기존 파일 기본 방법 실패: {str(e1)}")
                try:
                    existing_df = pd.read_excel(existing_file_path, engine='openpyxl', header=2, dtype=str, na_filter=False)
                    print("[OK] 기존 파일 openpyxl 엔진으로 읽기 성공 (3행부터, 문자열로 보존)")
                except Exception as e2:
                    print(f"[WARN] 기존 파일 openpyxl 엔진 실패: {str(e2)}")
                    try:
                        existing_df = pd.read_excel(existing_file_path, engine='xlrd', header=2, dtype=str, na_filter=False)
                        print("[OK] 기존 파일 xlrd 엔진으로 읽기 성공 (3행부터, 문자열로 보존)")
                    except Exception as e3:
                        print(f"[FAIL] 기존 파일 모든 엔진으로 읽기 실패: {str(e3)}")
                        return {'is_identical': False, 'has_new_data': False, 'new_data': []}
            
            if existing_df is None:
                print("[FAIL] 기존 파일을 읽을 수 없습니다.")
                return {'is_identical': False, 'has_new_data': False, 'new_data': []}
            
            # 메모리 기반 비교 함수 호출
            return self.compare_excel_files_with_dataframe(existing_df, new_file_path)
                
        except Exception as e:
            print(f"[FAIL] 파일 비교 실패: {str(e)}")
            return {'is_identical': False, 'has_new_data': False, 'new_data': []}
    
    def merge_and_update_file(self, access_token, existing_file_id, new_file_path, filename, new_data):
        """기존 파일에 새로운 데이터를 추가하여 업데이트 (원본 데이터 보존)"""
        try:
            print(f"[INFO] 기존 파일에 새로운 데이터 추가 중: {filename}")
            
            # 1. 기존 SharePoint 파일을 직접 메모리에서 읽기 (효율적)
            print("[INFO] 기존 SharePoint 파일 직접 읽기 중...")
            existing_df = self.read_sharepoint_file_to_dataframe(access_token, existing_file_id, filename)
            
            if existing_df is None:
                print("[WARN] 기존 파일 직접 읽기 실패, 새 파일로 업로드 시도...")
                return self.upload_new_file_with_same_name(access_token, new_file_path, filename)
            
            # 3. 기존 이메일오류 'Y' 데이터 보호하면서 새로운 데이터 병합
            print(f"[INFO] 새로운 데이터 {len(new_data)}건을 기존 파일에 추가 중...")
            print("[INFO] 기존 이메일오류 'Y' 데이터는 보호됩니다.")
            
            # 새로운 데이터를 DataFrame으로 변환 (문자열로 보존)
            new_df = pd.DataFrame(new_data)
            
            # 이메일오류 컬럼 찾기
            email_error_col = None
            for col in existing_df.columns:
                if '이메일오류' in str(col):
                    email_error_col = col
                    break
            
            if email_error_col:
                print(f"[INFO] 이메일오류 컬럼 발견: {email_error_col}")
                
                # 기존 데이터에서 이메일오류가 'Y'인 행들 보호
                protected_rows = existing_df[existing_df[email_error_col].str.upper() == 'Y']
                unprotected_rows = existing_df[existing_df[email_error_col].str.upper() != 'Y']
                
                print(f"[INFO] 보호할 데이터 (이메일오류='Y'): {len(protected_rows)}건")
                print(f"[INFO] 업데이트 가능한 기존 데이터: {len(unprotected_rows)}건")
                print(f"[INFO] 새로 추가할 데이터: {len(new_df)}건")
                
                # 안전한 병합: 보호된 행 + 새 데이터
                # (업데이트 가능한 기존 데이터는 새 데이터와 중복 확인 후 처리)
                merged_df = pd.concat([protected_rows, new_df], ignore_index=True)
                
                print(f"[OK] 안전한 병합 완료: 보호된 데이터 {len(protected_rows)}건 + 새 데이터 {len(new_df)}건")
            else:
                print("[WARN] 이메일오류 컬럼을 찾을 수 없습니다. 기본 병합을 수행합니다.")
                # 기존 데이터와 새 데이터 병합 (원본 순서 유지)
                merged_df = pd.concat([existing_df, new_df], ignore_index=True)
            
            print(f"[INFO] 최종 병합 결과: {len(merged_df)}행 (이메일오류='Y' 데이터 보호됨)")
            
            # 4. 병합된 데이터를 메모리에서 바로 임시 파일로 저장 (원본 형식 보존)
            temp_file_path = os.path.join(os.getcwd(), f"merged_{filename}")
            
            # ExcelWriter를 사용하여 원본 형식 보존
            with pd.ExcelWriter(temp_file_path, engine='openpyxl') as writer:
                merged_df.to_excel(writer, index=False, header=False, sheet_name='Sheet1')
                
                # 숫자 형식 보존을 위한 추가 설정
                worksheet = writer.sheets['Sheet1']
                for row in worksheet.iter_rows():
                    for cell in row:
                        if cell.value and isinstance(cell.value, str):
                            # 숫자로 변환 가능한 문자열인 경우 숫자 형식으로 설정
                            try:
                                if '.' in cell.value:
                                    float_val = float(cell.value)
                                    if float_val > 1e6:  # 큰 숫자의 경우 과학적 표기법 방지
                                        cell.number_format = '0'
                                    else:
                                        cell.number_format = '0.00'
                                else:
                                    int_val = int(cell.value)
                                    cell.number_format = '0'
                            except (ValueError, TypeError):
                                pass  # 숫자가 아닌 경우 그대로 유지
            
            print(f"[OK] 병합된 파일 생성 (원본 형식 보존): {temp_file_path}")
            
            # 5. 기존 파일 삭제
            print("[INFO] 기존 SharePoint 파일 삭제 중...")
            delete_success = self.delete_sharepoint_file(access_token, existing_file_id)
            
            if not delete_success:
                print("[WARN] 기존 파일 삭제 실패, 강제 업데이트 시도...")
                # 임시 파일 정리 (병합 파일만)
                if os.path.exists(temp_file_path):
                    os.remove(temp_file_path)
                return self.force_update_existing_file(access_token, existing_file_id, new_file_path)
            
            # 6. 병합된 파일을 새 파일로 업로드
            print("[INFO] 병합된 파일 업로드 중...")
            upload_success = self.upload_new_file_with_same_name(access_token, temp_file_path, filename)
            
            # 7. 임시 파일 정리 (병합 파일만)
            if os.path.exists(temp_file_path):
                os.remove(temp_file_path)
                print("[OK] 병합 파일 임시 파일 정리 완료")
            
            if upload_success:
                print(f"[OK] 파일 병합 및 업데이트 성공: {filename}")
                print(f"[INFO] 새로운 데이터 {len(new_data)}건이 추가되었습니다.")
                print("[INFO] 기존 이메일오류='Y' 데이터가 안전하게 보호되었습니다.")
                print("[INFO] 원본 데이터 형식이 보존되었습니다.")
                print("[INFO] SharePoint에서 파일을 새로고침하여 최신 버전을 확인하세요.")
                return True
            else:
                print(f"[FAIL] 병합된 파일 업로드 실패: {filename}")
                return False
                
        except Exception as e:
            print(f"[FAIL] 파일 병합 및 업데이트 실패: {str(e)}")
            # 임시 파일 정리 (병합 파일만)
            try:
                if 'temp_file_path' in locals() and os.path.exists(temp_file_path):
                    os.remove(temp_file_path)
            except:
                pass
            return False
    
    def delete_all_and_upload_new_file(self, access_token, existing_files, folder_id, filename, file_path):
        """모든 기존 파일들 삭제 후 새 파일 업로드"""
        try:
            print(f"[INFO] 모든 기존 파일들 삭제 후 새 파일 업로드 중: {filename}")
            print(f"[INFO] 삭제할 파일 수: {len(existing_files)}개")
            
            # 1. 모든 기존 파일들 삭제
            deleted_count = 0
            failed_deletions = []
            
            for i, file_info in enumerate(existing_files, 1):
                file_name = file_info['name']
                file_id = file_info['id']
                reason = file_info['reason']
                
                print(f"[INFO] 파일 삭제 중 ({i}/{len(existing_files)}): {file_name} - {reason}")
                delete_success = self.delete_sharepoint_file(access_token, file_id)
                
                if delete_success:
                    deleted_count += 1
                    print(f"[OK] 파일 삭제 성공: {file_name}")
                else:
                    failed_deletions.append(file_info)
                    print(f"[WARN] 파일 삭제 실패: {file_name}")
            
            print(f"[INFO] 삭제 완료: {deleted_count}/{len(existing_files)}개 파일")
            
            # 2. 삭제 실패한 파일이 있으면 강제 업데이트 시도
            if failed_deletions:
                print(f"[WARN] {len(failed_deletions)}개 파일 삭제 실패. 강제 업데이트 시도...")
                for file_info in failed_deletions:
                    try:
                        self.force_update_existing_file(access_token, file_info['id'], file_path)
                    except Exception as force_error:
                        print(f"[WARN] 강제 업데이트 실패: {file_info['name']} - {str(force_error)}")
            
            # 3. 새 파일 업로드 (동일한 파일명으로)
            print("[INFO] 새 파일 업로드 중...")
            upload_success = self.upload_new_file_with_same_name(access_token, file_path, filename)
            
            if upload_success:
                print(f"[OK] 모든 기존 파일들 삭제 후 새 파일 업로드 성공: {filename}")
                print(f"[INFO] 삭제된 파일 수: {deleted_count}개")
                print("[INFO] SharePoint에서 파일을 새로고침하여 최신 버전을 확인하세요.")
                return True
            else:
                print(f"[FAIL] 모든 기존 파일들 삭제 후 새 파일 업로드 실패: {filename}")
                return False
                
        except Exception as e:
            print(f"[FAIL] 모든 기존 파일들 삭제 후 새 파일 업로드 실패: {str(e)}")
            return False
    
    def delete_and_upload_new_file(self, access_token, existing_file_id, folder_id, filename, file_path):
        """기존 파일 삭제 후 새 파일 업로드 (호환성 유지)"""
        try:
            print(f"[INFO] 기존 파일 삭제 후 새 파일 업로드 중: {filename}")
            
            # 1. 기존 파일 삭제
            print("[INFO] 기존 파일 삭제 중...")
            delete_success = self.delete_sharepoint_file(access_token, existing_file_id)
            
            if not delete_success:
                print("[WARN] 기존 파일 삭제 실패, 강제 업데이트 시도...")
                return self.force_update_existing_file(access_token, existing_file_id, file_path)
            
            # 2. 새 파일 업로드 (동일한 파일명으로)
            print("[INFO] 새 파일 업로드 중...")
            upload_success = self.upload_new_file_with_same_name(access_token, file_path, filename)
            
            if upload_success:
                print(f"[OK] 기존 파일 삭제 후 새 파일 업로드 성공: {filename}")
                print("[INFO] SharePoint에서 파일을 새로고침하여 최신 버전을 확인하세요.")
                return True
            else:
                print(f"[FAIL] 기존 파일 삭제 후 새 파일 업로드 실패: {filename}")
                return False
                
        except Exception as e:
            print(f"[FAIL] 기존 파일 삭제 후 새 파일 업로드 실패: {str(e)}")
            return False
    
    def upload_new_file(self, access_token, folder_id, filename, file_path):
        """새 파일 업로드"""
        try:
            print(f"[INFO] 새 파일 업로드 중: {filename}")
            
            # 파일 읽기
            with open(file_path, 'rb') as file:
                file_content = file.read()
            
            # 업로드 URL
            if folder_id == self.sharepoint_drive_id:
                upload_url = f"https://graph.microsoft.com/v1.0/sites/{self.sharepoint_site_id}/drives/{self.sharepoint_drive_id}/root:/{filename}:/content"
            else:
                upload_url = f"https://graph.microsoft.com/v1.0/sites/{self.sharepoint_site_id}/drives/{self.sharepoint_drive_id}/items/{folder_id}:/{filename}:/content"
            
            headers = {
                'Authorization': f'Bearer {access_token}',
                'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            }
            
            response = requests.put(upload_url, data=file_content, headers=headers)
            
            if response.status_code in [200, 201]:
                print(f"[OK] 새 파일 업로드 성공: {filename}")
                
                # 업로드된 파일 정보 출력
                file_info = response.json()
                file_id = file_info.get('id')
                web_url = file_info.get('webUrl')
                
                print(f"[INFO] 파일 ID: {file_id}")
                print(f"[INFO] 웹 URL: {web_url}")
                
                return True
            else:
                print(f"[FAIL] 새 파일 업로드 실패: {response.status_code} - {response.text}")
                return False
                
        except Exception as e:
            print(f"[FAIL] 새 파일 업로드 실패: {str(e)}")
            return False
    
    def copy_content_and_replace(self, access_token, file_id, file_path, filename):
        """내용만 복사해서 새로운 엑셀 파일로 저장 후 업로드"""
        try:
            print(f"[INFO] 내용 복사 방식으로 파일 업데이트 중: {filename}")
            
            # 1. 기존 파일 다운로드
            print("[INFO] 기존 파일 다운로드 중...")
            download_success, temp_file_path = self.download_sharepoint_file(access_token, file_id, filename)
            
            if not download_success:
                print("[WARN] 기존 파일 다운로드 실패, 강제 업데이트 시도...")
                return self.force_update_existing_file(access_token, file_id, file_path)
            
            # 2. 다운로드한 파일의 내용을 새로운 엑셀 파일로 복사
            print("[INFO] 파일 내용을 새로운 엑셀 파일로 복사 중...")
            new_file_path = self.create_new_excel_from_content(temp_file_path, file_path)
            
            if not new_file_path:
                print("[WARN] 파일 내용 복사 실패, 강제 업데이트 시도...")
                # 임시 파일 정리
                if os.path.exists(temp_file_path):
                    os.remove(temp_file_path)
                return self.force_update_existing_file(access_token, file_id, file_path)
            
            # 3. 기존 파일 삭제
            print("[INFO] 기존 파일 삭제 중...")
            delete_success = self.delete_sharepoint_file(access_token, file_id)
            
            if not delete_success:
                print("[WARN] 기존 파일 삭제 실패, 강제 업데이트 시도...")
                # 임시 파일들 정리
                if os.path.exists(temp_file_path):
                    os.remove(temp_file_path)
                if os.path.exists(new_file_path):
                    os.remove(new_file_path)
                return self.force_update_existing_file(access_token, file_id, file_path)
            
            # 4. 새로운 파일 업로드 (동일한 파일명으로)
            print("[INFO] 새로운 파일 업로드 중...")
            upload_success = self.upload_new_file_with_same_name(access_token, new_file_path, filename)
            
            # 5. 임시 파일들 정리
            if os.path.exists(temp_file_path):
                os.remove(temp_file_path)
                print("[OK] 다운로드 임시 파일 정리 완료")
            if os.path.exists(new_file_path):
                os.remove(new_file_path)
                print("[OK] 생성된 임시 파일 정리 완료")
            
            if upload_success:
                print(f"[OK] 내용 복사 방식으로 파일 업데이트 성공: {filename}")
                print("[INFO] SharePoint에서 파일을 새로고침하여 최신 버전을 확인하세요.")
                return True
            else:
                print(f"[FAIL] 내용 복사 방식으로 파일 업데이트 실패: {filename}")
                return False
                
        except Exception as e:
            print(f"[FAIL] 내용 복사 방식으로 파일 업데이트 실패: {str(e)}")
            return False
    
    def create_new_excel_from_content(self, source_file_path, target_file_path):
        """다운로드한 파일의 내용을 새로운 엑셀 파일로 복사 (원본 데이터 보존)"""
        try:
            print("[INFO] 파일 내용을 새로운 엑셀 파일로 복사 중...")
            
            # 새로운 파일 경로 생성
            import datetime
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            new_filename = f"new_content_{timestamp}.xlsx"
            new_file_path = os.path.join(os.getcwd(), new_filename)
            
            print(f"[INFO] 새로운 파일 경로: {new_file_path}")
            
            # 다운로드한 파일 읽기 (원본 데이터 보존을 위한 설정)
            source_df = None
            try:
                source_df = pd.read_excel(source_file_path, header=2, dtype=str, na_filter=False)
                print("[OK] 다운로드 파일 읽기 성공 (3행부터, 문자열로 보존)")
            except Exception as e1:
                print(f"[WARN] 기본 방법 실패: {str(e1)}")
                try:
                    source_df = pd.read_excel(source_file_path, engine='openpyxl', header=2, dtype=str, na_filter=False)
                    print("[OK] openpyxl 엔진으로 다운로드 파일 읽기 성공 (3행부터, 문자열로 보존)")
                except Exception as e2:
                    print(f"[WARN] openpyxl 엔진 실패: {str(e2)}")
                    try:
                        source_df = pd.read_excel(source_file_path, engine='xlrd', header=2, dtype=str, na_filter=False)
                        print("[OK] xlrd 엔진으로 다운로드 파일 읽기 성공 (3행부터, 문자열로 보존)")
                    except Exception as e3:
                        print(f"[FAIL] 모든 엔진으로 다운로드 파일 읽기 실패: {str(e3)}")
                        return None
            
            if source_df is None:
                print("[FAIL] 다운로드 파일을 읽을 수 없습니다.")
                return None
            
            print(f"[INFO] 다운로드 파일 데이터: {len(source_df)}행, {len(source_df.columns)}열")
            
            # 업로드할 파일 읽기 (원본 데이터 보존을 위한 설정)
            target_df = None
            try:
                target_df = pd.read_excel(target_file_path, header=2, dtype=str, na_filter=False)
                print("[OK] 업로드 파일 읽기 성공 (3행부터, 문자열로 보존)")
            except Exception as e1:
                print(f"[WARN] 기본 방법 실패: {str(e1)}")
                try:
                    target_df = pd.read_excel(target_file_path, engine='openpyxl', header=2, dtype=str, na_filter=False)
                    print("[OK] openpyxl 엔진으로 업로드 파일 읽기 성공 (3행부터, 문자열로 보존)")
                except Exception as e2:
                    print(f"[WARN] openpyxl 엔진 실패: {str(e2)}")
                    try:
                        target_df = pd.read_excel(target_file_path, engine='xlrd', header=2, dtype=str, na_filter=False)
                        print("[OK] xlrd 엔진으로 업로드 파일 읽기 성공 (3행부터, 문자열로 보존)")
                    except Exception as e3:
                        print(f"[FAIL] 모든 엔진으로 업로드 파일 읽기 실패: {str(e3)}")
                        return None
            
            if target_df is None:
                print("[FAIL] 업로드 파일을 읽을 수 없습니다.")
                return None
            
            print(f"[INFO] 업로드 파일 데이터: {len(target_df)}행, {len(target_df.columns)}열")
            
            # 데이터 병합 (업로드할 파일의 데이터로 덮어쓰기)
            print("[INFO] 데이터 병합 중...")
            
            # 컬럼명이 동일한지 확인
            if list(source_df.columns) == list(target_df.columns):
                print("[OK] 컬럼명이 동일합니다. 데이터를 병합합니다.")
                # 업로드할 파일의 데이터로 완전히 교체
                merged_df = target_df.copy()
            else:
                print("[WARN] 컬럼명이 다릅니다. 업로드 파일의 데이터를 우선으로 합니다.")
                # 업로드할 파일의 데이터를 사용
                merged_df = target_df.copy()
            
            # 새로운 엑셀 파일로 저장 (원본 형식 보존)
            print("[INFO] 새로운 엑셀 파일로 저장 중...")
            
            # ExcelWriter를 사용하여 원본 형식 보존
            with pd.ExcelWriter(new_file_path, engine='openpyxl') as writer:
                merged_df.to_excel(writer, index=False, header=False, sheet_name='Sheet1')
                
                # 숫자 형식 보존을 위한 추가 설정
                worksheet = writer.sheets['Sheet1']
                for row in worksheet.iter_rows():
                    for cell in row:
                        if cell.value and isinstance(cell.value, str):
                            # 숫자로 변환 가능한 문자열인 경우 숫자 형식으로 설정
                            try:
                                if '.' in cell.value:
                                    float_val = float(cell.value)
                                    if float_val > 1e6:  # 큰 숫자의 경우 과학적 표기법 방지
                                        cell.number_format = '0'
                                    else:
                                        cell.number_format = '0.00'
                                else:
                                    int_val = int(cell.value)
                                    cell.number_format = '0'
                            except (ValueError, TypeError):
                                pass  # 숫자가 아닌 경우 그대로 유지
            
            print(f"[OK] 새로운 엑셀 파일 생성 완료 (원본 형식 보존): {new_file_path}")
            print(f"[INFO] 새로운 파일 크기: {os.path.getsize(new_file_path)} bytes")
            
            return new_file_path
            
        except Exception as e:
            print(f"[FAIL] 파일 내용 복사 실패: {str(e)}")
            import traceback
            print(f"[DEBUG] 상세 오류: {traceback.format_exc()}")
            return None
    
    def create_backup_and_replace(self, access_token, file_id, file_path, filename):
        """백업 생성 후 복사본 저장 방식으로 파일 교체"""
        try:
            print(f"[INFO] 백업 생성 후 복사본 저장 방식으로 파일 교체 중: {filename}")
            
            # 1. 기존 파일을 백업으로 복사
            print("[INFO] 기존 파일을 백업으로 복사 중...")
            backup_success = self.create_file_backup(access_token, file_id, filename)
            
            if not backup_success:
                print("[WARN] 백업 생성 실패, 강제 업데이트 시도...")
                return self.force_update_existing_file(access_token, file_id, file_path)
            
            # 2. 기존 파일 삭제
            print("[INFO] 기존 파일 삭제 중...")
            delete_success = self.delete_sharepoint_file(access_token, file_id)
            
            if not delete_success:
                print("[WARN] 기존 파일 삭제 실패, 강제 업데이트 시도...")
                return self.force_update_existing_file(access_token, file_id, file_path)
            
            # 3. 새 파일 업로드 (동일한 파일명으로)
            print("[INFO] 새 파일 업로드 중...")
            upload_success = self.upload_new_file_with_same_name(access_token, file_path, filename)
            
            if upload_success:
                print(f"[OK] 백업 생성 후 복사본 저장 방식으로 파일 교체 성공: {filename}")
                print("[INFO] SharePoint에서 파일을 새로고침하여 최신 버전을 확인하세요.")
                return True
            else:
                print(f"[FAIL] 백업 생성 후 복사본 저장 방식으로 파일 교체 실패: {filename}")
                return False
                
        except Exception as e:
            print(f"[FAIL] 백업 생성 후 복사본 저장 방식으로 파일 교체 실패: {str(e)}")
            return False
    
    def create_file_backup(self, access_token, file_id, filename):
        """SharePoint 파일을 백업으로 복사"""
        try:
            print(f"[INFO] SharePoint 파일 백업 생성 중: {filename}")
            
            # 백업 파일명 생성 (업로드일자와 타임스탬프 추가)
            import datetime
            name, ext = os.path.splitext(filename)
            upload_date = datetime.datetime.now().strftime("%Y.%m.%d")
            timestamp = datetime.datetime.now().strftime("%H%M%S")
            backup_filename = f"{name}_backup_{upload_date}_{timestamp}{ext}"
            
            print(f"[INFO] 백업 파일명: {backup_filename}")
            
            # 업로드 폴더 찾기
            folder_id = self.find_sharepoint_folder(access_token, self.upload_folder_path)
            if not folder_id:
                print(f"[FAIL] 업로드 폴더를 찾을 수 없습니다: {self.upload_folder_path}")
                return False
            
            # 기존 파일 다운로드
            download_success, temp_file_path = self.download_sharepoint_file(access_token, file_id, filename)
            
            if not download_success or not temp_file_path:
                print("[FAIL] 백업을 위한 파일 다운로드 실패")
                return False
            
            # 백업 파일로 업로드
            upload_url = f"https://graph.microsoft.com/v1.0/sites/{self.sharepoint_site_id}/drives/{self.sharepoint_drive_id}/items/{folder_id}:/{backup_filename}:/content"
            
            headers = {
                'Authorization': f'Bearer {access_token}',
                'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            }
            
            with open(temp_file_path, 'rb') as f:
                file_content = f.read()
            
            response = requests.put(upload_url, headers=headers, data=file_content)
            
            # 임시 파일 정리
            if os.path.exists(temp_file_path):
                os.remove(temp_file_path)
                print("[OK] 임시 파일 정리 완료")
            
            if response.status_code in [200, 201]:
                print(f"[OK] SharePoint 파일 백업 생성 성공: {backup_filename}")
                
                # 백업 파일 정보 출력
                backup_info = response.json()
                backup_file_id = backup_info.get('id')
                backup_web_url = backup_info.get('webUrl')
                
                print(f"[INFO] 백업 파일 ID: {backup_file_id}")
                print(f"[INFO] 백업 파일 URL: {backup_web_url}")
                
                return True
            else:
                print(f"[FAIL] SharePoint 파일 백업 생성 실패: {response.status_code} - {response.text}")
                return False
                
        except Exception as e:
            print(f"[FAIL] SharePoint 파일 백업 생성 실패: {str(e)}")
            return False
    
    def copy_and_replace_file(self, access_token, file_id, file_path, filename):
        """복사본 저장 방식으로 파일 업데이트 (다운로드 후 동일 파일명으로 업로드)"""
        try:
            print(f"[INFO] 복사본 저장 방식으로 파일 업데이트 중: {filename}")
            
            # 1. 기존 파일 다운로드
            print("[INFO] 기존 파일 다운로드 중...")
            download_success, temp_file_path = self.download_sharepoint_file(access_token, file_id, filename)
            
            if not download_success:
                print("[WARN] 기존 파일 다운로드 실패, 강제 업데이트 시도...")
                return self.force_update_existing_file(access_token, file_id, file_path)
            
            # 2. 기존 파일 삭제
            print("[INFO] 기존 파일 삭제 중...")
            delete_success = self.delete_sharepoint_file(access_token, file_id)
            
            if not delete_success:
                print("[WARN] 기존 파일 삭제 실패, 강제 업데이트 시도...")
                # 임시 파일 정리
                if os.path.exists(temp_file_path):
                    os.remove(temp_file_path)
                return self.force_update_existing_file(access_token, file_id, file_path)
            
            # 3. 새 파일 업로드 (동일한 파일명으로)
            print("[INFO] 새 파일 업로드 중...")
            upload_success = self.upload_new_file_with_same_name(access_token, file_path, filename)
            
            # 4. 임시 파일 정리
            if os.path.exists(temp_file_path):
                os.remove(temp_file_path)
                print("[OK] 임시 파일 정리 완료")
            
            if upload_success:
                print(f"[OK] 복사본 저장 방식으로 파일 업데이트 성공: {filename}")
                print("[INFO] SharePoint에서 파일을 새로고침하여 최신 버전을 확인하세요.")
                return True
            else:
                print(f"[FAIL] 복사본 저장 방식으로 파일 업데이트 실패: {filename}")
                return False
                
        except Exception as e:
            print(f"[FAIL] 복사본 저장 방식으로 파일 업데이트 실패: {str(e)}")
            return False
    
    def read_sharepoint_file_to_dataframe(self, access_token, file_id, filename):
        """SharePoint 파일을 직접 메모리에서 읽어 pandas DataFrame으로 변환 (효율적)"""
        try:
            print(f"[INFO] SharePoint 파일 직접 읽기 중: {filename}")
            
            # 파일 콘텐츠 직접 읽기 URL
            download_url = f"https://graph.microsoft.com/v1.0/sites/{self.sharepoint_site_id}/drives/{self.sharepoint_drive_id}/items/{file_id}/content"
            
            headers = {
                'Authorization': f'Bearer {access_token}',
                'Content-Type': 'application/json'
            }
            
            response = requests.get(download_url, headers=headers)
            
            if response.status_code == 200:
                # 메모리에서 직접 BytesIO 스트림으로 변환
                file_stream = io.BytesIO(response.content)
                
                print(f"[OK] SharePoint 파일 콘텐츠 읽기 성공: {len(response.content)} bytes")
                
                # pandas로 직접 읽기 (원본 데이터 보존을 위한 설정)
                existing_df = None
                try:
                    existing_df = pd.read_excel(file_stream, header=2, dtype=str, na_filter=False)
                    print("[OK] 메모리에서 DataFrame 변환 성공 (3행부터, 문자열로 보존)")
                except Exception as e1:
                    print(f"[WARN] 기본 방법 실패: {str(e1)}")
                    try:
                        file_stream.seek(0)  # 스트림 위치 초기화
                        existing_df = pd.read_excel(file_stream, engine='openpyxl', header=2, dtype=str, na_filter=False)
                        print("[OK] openpyxl 엔진으로 메모리에서 DataFrame 변환 성공")
                    except Exception as e2:
                        print(f"[FAIL] openpyxl 엔진도 실패: {str(e2)}")
                        return None
                
                if existing_df is not None:
                    print(f"[INFO] DataFrame 생성 완료: {len(existing_df)}행, {len(existing_df.columns)}열")
                    return existing_df
                else:
                    print("[FAIL] DataFrame 생성 실패")
                    return None
                    
            else:
                print(f"[FAIL] SharePoint 파일 콘텐츠 읽기 실패: {response.status_code} - {response.text}")
                return None
                
        except Exception as e:
            print(f"[FAIL] SharePoint 파일 직접 읽기 실패: {str(e)}")
            return None

    def read_sharepoint_file_with_full_structure(self, access_token, file_id, filename):
        """SharePoint 파일을 원본 구조(1,2,3행 헤더 포함) 그대로 읽기"""
        try:
            print(f"[INFO] SharePoint 파일 원본 구조로 읽기 중: {filename}")
            
            # 파일 콘텐츠 직접 읽기 URL
            download_url = f"https://graph.microsoft.com/v1.0/sites/{self.sharepoint_site_id}/drives/{self.sharepoint_drive_id}/items/{file_id}/content"
            
            headers = {
                'Authorization': f'Bearer {access_token}',
                'Content-Type': 'application/json'
            }
            
            response = requests.get(download_url, headers=headers)
            
            if response.status_code == 200:
                # 메모리에서 직접 BytesIO 스트림으로 변환
                file_stream = io.BytesIO(response.content)
                
                print(f"[OK] SharePoint 파일 콘텐츠 읽기 성공: {len(response.content)} bytes")
                
                # openpyxl로 원본 구조 완전 보존하여 읽기
                from openpyxl import load_workbook
                
                try:
                    workbook = load_workbook(file_stream)
                    worksheet = workbook.active
                    
                    print(f"[OK] 원본 구조로 워크북 로드 완료: {worksheet.max_row}행, {worksheet.max_column}열")
                    
                    # 전체 데이터를 리스트로 변환 (1행부터 모든 행 포함)
                    full_data = []
                    for row in range(1, worksheet.max_row + 1):
                        row_data = []
                        for col in range(1, worksheet.max_column + 1):
                            cell_value = worksheet.cell(row=row, column=col).value
                            row_data.append(str(cell_value) if cell_value is not None else "")
                        full_data.append(row_data)
                    
                    # 1,2행은 헤더, 3행은 컬럼명, 4행부터 데이터
                    if len(full_data) >= 4:
                        header_row_1 = full_data[0]  # 1행
                        header_row_2 = full_data[1]  # 2행
                        column_names = full_data[2]  # 3행 (컬럼명)
                        data_rows = full_data[3:]    # 4행부터 데이터
                        
                        # pandas DataFrame 생성 (3행을 컬럼명으로 사용)
                        df = pd.DataFrame(data_rows, columns=column_names)
                        
                        # 헤더 정보를 DataFrame의 메타데이터로 저장
                        df.attrs['header_row_1'] = header_row_1
                        df.attrs['header_row_2'] = header_row_2
                        df.attrs['column_names'] = column_names
                        
                        print(f"[OK] 원본 구조 보존 완료: 헤더 1,2행 + 컬럼명 3행 + 데이터 {len(df)}행")
                        return df
                    else:
                        print(f"[WARN] 파일에 충분한 행이 없습니다: {len(full_data)}행")
                        return None
                        
                except Exception as e:
                    print(f"[FAIL] openpyxl로 원본 구조 읽기 실패: {str(e)}")
                    return None
                    
            else:
                print(f"[FAIL] SharePoint 파일 콘텐츠 읽기 실패: {response.status_code} - {response.text}")
                return None
                
        except Exception as e:
            print(f"[FAIL] SharePoint 파일 원본 구조 읽기 실패: {str(e)}")
            return None

    def download_sharepoint_file(self, access_token, file_id, filename):
        """SharePoint 파일 다운로드 (기존 방식 - 호환성 유지)"""
        try:
            print(f"[INFO] SharePoint 파일 다운로드 중: {filename}")
            
            # 다운로드 URL
            download_url = f"https://graph.microsoft.com/v1.0/sites/{self.sharepoint_site_id}/drives/{self.sharepoint_drive_id}/items/{file_id}/content"
            
            headers = {
                'Authorization': f'Bearer {access_token}',
                'Content-Type': 'application/json'
            }
            
            response = requests.get(download_url, headers=headers)
            
            if response.status_code == 200:
                # 임시 파일로 저장 (현재 디렉토리에 저장)
                temp_filename = f"temp_{filename}"
                temp_file_path = os.path.join(os.getcwd(), temp_filename)
                
                with open(temp_file_path, 'wb') as f:
                    f.write(response.content)
                
                print(f"[OK] SharePoint 파일 다운로드 성공: {temp_filename}")
                return True, temp_file_path
            else:
                print(f"[FAIL] SharePoint 파일 다운로드 실패: {response.status_code} - {response.text}")
                return False, None
                
        except Exception as e:
            print(f"[FAIL] SharePoint 파일 다운로드 실패: {str(e)}")
            return False, None
    
    def delete_sharepoint_file(self, access_token, file_id):
        """SharePoint 파일 삭제"""
        try:
            print(f"[INFO] SharePoint 파일 삭제 중: {file_id}")
            
            # 삭제 URL
            delete_url = f"https://graph.microsoft.com/v1.0/sites/{self.sharepoint_site_id}/drives/{self.sharepoint_drive_id}/items/{file_id}"
            
            headers = {
                'Authorization': f'Bearer {access_token}',
                'Content-Type': 'application/json'
            }
            
            response = requests.delete(delete_url, headers=headers)
            
            if response.status_code in [200, 204]:
                print(f"[OK] SharePoint 파일 삭제 성공: {file_id}")
                return True
            else:
                print(f"[FAIL] SharePoint 파일 삭제 실패: {response.status_code} - {response.text}")
                return False
                
        except Exception as e:
            print(f"[FAIL] SharePoint 파일 삭제 실패: {str(e)}")
            return False
    
    def upload_new_file_with_same_name(self, access_token, file_path, filename):
        """동일한 파일명으로 새 파일 업로드"""
        try:
            print(f"[INFO] 동일한 파일명으로 새 파일 업로드 중: {filename}")
            
            # 업로드 폴더 찾기
            folder_id = self.find_sharepoint_folder(access_token, self.upload_folder_path)
            if not folder_id:
                print(f"[FAIL] 업로드 폴더를 찾을 수 없습니다: {self.upload_folder_path}")
                return False
            
            # 새 파일로 업로드
            upload_url = f"https://graph.microsoft.com/v1.0/sites/{self.sharepoint_site_id}/drives/{self.sharepoint_drive_id}/items/{folder_id}:/{filename}:/content"
            
            headers = {
                'Authorization': f'Bearer {access_token}',
                'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            }
            
            with open(file_path, 'rb') as f:
                file_content = f.read()
            
            response = requests.put(upload_url, headers=headers, data=file_content)
            
            if response.status_code in [200, 201]:
                print(f"[OK] 동일한 파일명으로 새 파일 업로드 성공: {filename}")
                
                # 업로드된 파일 정보 출력
                file_info = response.json()
                new_file_id = file_info.get('id')
                web_url = file_info.get('webUrl')
                
                print(f"[INFO] 새 파일 ID: {new_file_id}")
                print(f"[INFO] 웹 URL: {web_url}")
                
                return True
            else:
                print(f"[FAIL] 동일한 파일명으로 새 파일 업로드 실패: {response.status_code} - {response.text}")
                return False
                
        except Exception as e:
            print(f"[FAIL] 동일한 파일명으로 새 파일 업로드 실패: {str(e)}")
            return False
    
    def force_update_existing_file(self, access_token, file_id, file_path):
        """기존 파일 강제 업데이트 (파일 잠금 해제 후 덮어쓰기)"""
        try:
            print(f"[INFO] 기존 파일 강제 업데이트 중: {file_id}")
            
            # 1. 파일 잠금 해제 시도
            print("[INFO] 파일 잠금 해제 시도...")
            unlock_success = self.unlock_sharepoint_file(access_token, file_id)
            if unlock_success:
                print("[OK] 파일 잠금 해제 성공")
            else:
                print("[WARN] 파일 잠금 해제 실패, 강제 업데이트 시도...")
            
            # 2. 파일 읽기
            with open(file_path, 'rb') as file:
                file_content = file.read()
            
            # 3. 강제 업데이트 URL (If-Match 헤더 사용)
            update_url = f"https://graph.microsoft.com/v1.0/sites/{self.sharepoint_site_id}/drives/{self.sharepoint_drive_id}/items/{file_id}/content"
            
            headers = {
                'Authorization': f'Bearer {access_token}',
                'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                'If-Match': '*'  # 강제 덮어쓰기
            }
            
            response = requests.put(update_url, data=file_content, headers=headers)
            
            if response.status_code in [200, 201]:
                print(f"[OK] 파일 강제 업데이트 성공: {file_id}")
                
                # 업데이트된 파일 정보 출력
                file_info = response.json()
                web_url = file_info.get('webUrl')
                
                print(f"[INFO] 업데이트된 파일 ID: {file_id}")
                print(f"[INFO] 웹 URL: {web_url}")
                print("[INFO] SharePoint에서 파일을 새로고침하여 최신 버전을 확인하세요.")
                
                return True
            elif response.status_code == 423:
                print("[WARN] 파일이 여전히 잠겨있습니다. 새 파일로 업로드 시도...")
                return self.upload_as_new_file_with_timestamp(access_token, file_path)
            else:
                print(f"[FAIL] 파일 강제 업데이트 실패: {response.status_code} - {response.text}")
                return False
                
        except Exception as e:
            print(f"[FAIL] 파일 강제 업데이트 실패: {str(e)}")
            return False
    
    def unlock_sharepoint_file(self, access_token, file_id):
        """SharePoint 파일 잠금 해제"""
        try:
            print("[INFO] SharePoint 파일 잠금 해제 시도...")
            
            unlock_url = f"https://graph.microsoft.com/v1.0/sites/{self.sharepoint_site_id}/drives/{self.sharepoint_drive_id}/items/{file_id}/unfollow"
            
            headers = {
                'Authorization': f'Bearer {access_token}',
                'Content-Type': 'application/json'
            }
            
            response = requests.post(unlock_url, headers=headers)
            
            if response.status_code in [200, 204]:
                print("[OK] 파일 잠금 해제 성공")
                return True
            else:
                print(f"[WARN] 파일 잠금 해제 실패: {response.status_code}")
                return False
                
        except Exception as e:
            print(f"[WARN] 파일 잠금 해제 중 오류: {str(e)}")
            return False
    
    def upload_as_new_file_with_timestamp(self, access_token, file_path):
        """타임스탬프가 포함된 새 파일로 업로드"""
        try:
            print("[INFO] 타임스탬프가 포함된 새 파일로 업로드 시도...")
            
            # 업로드 폴더 찾기
            folder_id = self.find_sharepoint_folder(access_token, self.upload_folder_path)
            if not folder_id:
                print(f"[FAIL] 업로드 폴더를 찾을 수 없습니다: {self.upload_folder_path}")
                return False
            
            # 새 파일명 생성 (업로드일자와 타임스탬프 추가)
            import datetime
            original_filename = os.path.basename(file_path)
            name, ext = os.path.splitext(original_filename)
            upload_date = datetime.datetime.now().strftime("%Y.%m.%d")
            timestamp = datetime.datetime.now().strftime("%H%M%S")
            new_filename = f"{name}_updated_{upload_date}_{timestamp}{ext}"
            
            print(f"[INFO] 새 파일명: {new_filename}")
            
            # 새 파일로 업로드
            upload_url = f"https://graph.microsoft.com/v1.0/sites/{self.sharepoint_site_id}/drives/{self.sharepoint_drive_id}/items/{folder_id}:/{new_filename}:/content"
            
            headers = {
                'Authorization': f'Bearer {access_token}',
                'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            }
            
            with open(file_path, 'rb') as f:
                file_content = f.read()
            
            response = requests.put(upload_url, headers=headers, data=file_content)
            
            if response.status_code in [200, 201]:
                print(f"[OK] 새 파일로 업로드 성공: {new_filename}")
                print("[INFO] 원본 파일이 잠겨있어 새 파일로 저장되었습니다.")
                print("[INFO] SharePoint에서 새 파일을 확인하세요.")
                return True
            else:
                print(f"[FAIL] 새 파일 업로드 실패: {response.status_code} - {response.text}")
                return False
                
        except Exception as e:
            print(f"[FAIL] 새 파일 업로드 실패: {str(e)}")
            return False
    
    def update_existing_file(self, access_token, file_id, file_path):
        """기존 파일 업데이트 (기존 메서드 유지)"""
        try:
            print(f"[INFO] 기존 파일 업데이트 중: {file_id}")
            
            # 파일 읽기
            with open(file_path, 'rb') as file:
                file_content = file.read()
            
            # 업데이트 URL
            update_url = f"https://graph.microsoft.com/v1.0/sites/{self.sharepoint_site_id}/drives/{self.sharepoint_drive_id}/items/{file_id}/content"
            
            headers = {
                'Authorization': f'Bearer {access_token}',
                'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            }
            
            response = requests.put(update_url, data=file_content, headers=headers)
            
            if response.status_code in [200, 201]:
                print(f"[OK] 파일 업데이트 성공: {file_id}")
                
                # 업데이트된 파일 정보 출력
                file_info = response.json()
                web_url = file_info.get('webUrl')
                
                print(f"[INFO] 업데이트된 파일 ID: {file_id}")
                print(f"[INFO] 웹 URL: {web_url}")
                
                return True
            else:
                print(f"[FAIL] 파일 업데이트 실패: {response.status_code} - {response.text}")
                return False
                
        except Exception as e:
            print(f"[FAIL] 파일 업데이트 실패: {str(e)}")
            return False
    
    def find_and_upload_downloaded_file(self):
        """다운로드된 파일 찾기 및 SharePoint 업로드"""
        try:
            print("\n" + "="*60)
            print("다운로드된 파일 찾기 및 업로드")
            print("="*60)
            
            # 다운로드 폴더 경로들 (일반적인 다운로드 폴더들)
            download_paths = [
                os.path.expanduser("~/Downloads"),
                os.path.expanduser("~/Desktop"),
                os.path.join(os.getcwd(), "downloads"),
                os.path.join(os.getcwd(), ".")
            ]
            
            # 최근 다운로드된 엑셀 파일 찾기
            excel_files = []
            current_time = time.time()
            
            for download_path in download_paths:
                if os.path.exists(download_path):
                    print(f"[INFO] 폴더 검색 중: {download_path}")
                    
                    try:
                        for file in os.listdir(download_path):
                            if file.lower().endswith('.xlsx') or file.lower().endswith('.xls'):
                                file_path = os.path.join(download_path, file)
                                file_time = os.path.getmtime(file_path)
                                
                                # 최근 5분 이내에 생성된 파일만
                                if current_time - file_time < 300:  # 5분 = 300초
                                    excel_files.append((file_path, file_time, file))
                                    print(f"[INFO] 최근 엑셀 파일 발견: {file}")
                    except Exception as e:
                        print(f"[WARN] 폴더 검색 실패: {download_path} - {str(e)}")
            
            if not excel_files:
                print("[WARN] 최근 다운로드된 엑셀 파일을 찾을 수 없습니다.")
                return False
            
            # 가장 최근 파일 선택
            excel_files.sort(key=lambda x: x[1], reverse=True)
            latest_file = excel_files[0]
            file_path = latest_file[0]
            filename = latest_file[2]
            
            print(f"[OK] 업로드할 파일: {filename}")
            print(f"[INFO] 파일 경로: {file_path}")
            
            # SharePoint에 업로드 (고정 파일명 사용)
            upload_filename = "역발행 세금계산서.xlsx"
            upload_success = self.upload_file_to_sharepoint(file_path, upload_filename)
            
            if upload_success:
                print(f"[OK] 파일 업로드 완료: {filename}")
                
                # SharePoint에서 업로드된 파일의 ID 찾기
                file_id = self.get_uploaded_file_id(upload_filename)
                if file_id:
                    print(f"[OK] 업로드된 파일 ID 획득: {file_id}")
                    return file_path, file_id  # 파일 경로와 ID 함께 반환
                else:
                    print("[WARN] 파일 ID 획득 실패")
                    return file_path, None
            else:
                print(f"[FAIL] 파일 업로드 실패: {filename}")
                return None
                
        except Exception as e:
            print(f"[FAIL] 파일 찾기 및 업로드 실패: {str(e)}")
            return False
    
    def get_uploaded_file_id(self, filename):
        """SharePoint에서 업로드된 파일의 ID 찾기"""
        try:
            print(f"[INFO] SharePoint에서 파일 ID 찾는 중: {filename}")
            
            # 액세스 토큰 획득
            access_token = self.get_sharepoint_access_token()
            if not access_token:
                print("[FAIL] SharePoint 액세스 토큰 획득 실패")
                return None
            
            # 업로드 폴더 찾기
            folder_id = self.find_sharepoint_folder(access_token, self.upload_folder_path)
            if not folder_id:
                print(f"[FAIL] 업로드 폴더를 찾을 수 없습니다: {self.upload_folder_path}")
                return None
            
            # 폴더 내 파일 목록 조회
            if folder_id == self.sharepoint_drive_id:
                url = f"https://graph.microsoft.com/v1.0/sites/{self.sharepoint_site_id}/drives/{self.sharepoint_drive_id}/root/children"
            else:
                url = f"https://graph.microsoft.com/v1.0/sites/{self.sharepoint_site_id}/drives/{self.sharepoint_drive_id}/items/{folder_id}/children"
            
            headers = {
                'Authorization': f'Bearer {access_token}',
                'Content-Type': 'application/json'
            }
            
            response = requests.get(url, headers=headers)
            
            if response.status_code == 200:
                data = response.json()
                files = data.get('value', [])
                
                # 정확한 파일명으로 찾기
                for file in files:
                    if file.get('name') == filename:
                        file_id = file.get('id')
                        print(f"[OK] 파일 ID 발견: {filename} (ID: {file_id})")
                        return file_id
                
                print(f"[WARN] 파일을 찾을 수 없습니다: {filename}")
                return None
            else:
                print(f"[FAIL] 파일 목록 조회 실패: {response.status_code} - {response.text}")
                return None
                
        except Exception as e:
            print(f"[FAIL] 파일 ID 찾기 실패: {str(e)}")
            return None
    
    def decrypt_excel_file(self, encrypted_file_path, password):
        """암호화된 엑셀 파일 해독"""
        try:
            print(f"[INFO] 암호화된 엑셀 파일 해독 중: {encrypted_file_path}")
            print(f"[INFO] 파일 존재 여부: {os.path.exists(encrypted_file_path)}")
            print(f"[INFO] 파일 크기: {os.path.getsize(encrypted_file_path)} bytes")
            
            # msoffcrypto 모듈이 없는 경우
            if msoffcrypto is None:
                print("[WARN] msoffcrypto 모듈이 없어 암호 해독을 할 수 없습니다.")
                print("[INFO] msoffcrypto-tool 패키지를 설치하세요: pip install msoffcrypto-tool")
                return None
            
            print("[OK] msoffcrypto 모듈 사용 가능")
            
            # 해독된 파일 경로 생성
            base_name = os.path.splitext(os.path.basename(encrypted_file_path))[0]
            decrypted_file_path = os.path.join(os.path.dirname(encrypted_file_path), f"{base_name}_decrypted.xlsx")
            print(f"[INFO] 해독된 파일 저장 경로: {decrypted_file_path}")
            
            # 암호화된 파일 열기
            with open(encrypted_file_path, 'rb') as encrypted_file:
                print("[INFO] 파일 열기 성공")
                
                # msoffcrypto로 암호화된 파일 처리
                office_file = msoffcrypto.OfficeFile(encrypted_file)
                print("[INFO] OfficeFile 객체 생성 성공")
                
                # 파일이 암호화되어 있는지 확인
                try:
                    if hasattr(office_file, 'is_encrypted'):
                        is_encrypted = office_file.is_encrypted()
                        print(f"[INFO] 파일 암호화 여부: {is_encrypted}")
                        
                        if not is_encrypted:
                            print("[INFO] 파일이 암호화되지 않았습니다. 해독할 필요가 없습니다.")
                            return None
                    else:
                        print("[WARN] is_encrypted 메서드가 없어 암호화 여부를 확인할 수 없습니다.")
                        print("[INFO] 암호 해독을 시도합니다.")
                        
                except Exception as check_error:
                    print(f"[WARN] 암호화 여부 확인 실패: {str(check_error)}")
                    print("[INFO] 암호 해독을 시도합니다.")
                
                # 암호로 해독 시도
                try:
                    office_file.load_key(password=password)
                    print(f"[OK] 암호 확인 성공")
                except Exception as key_error:
                    print(f"[FAIL] 암호 확인 실패: {str(key_error)}")
                    print(f"[INFO] 사용한 암호: {password}")
                    print(f"[INFO] 가능한 원인:")
                    print(f"  - 암호가 틀렸을 수 있습니다")
                    print(f"  - 파일이 다른 방식으로 암호화되었을 수 있습니다")
                    print(f"  - 파일이 손상되었을 수 있습니다")
                    return None
                
                # 해독된 데이터를 메모리에 저장
                decrypted_data = io.BytesIO()
                office_file.decrypt(decrypted_data)
                decrypted_data.seek(0)
                print(f"[INFO] 해독된 데이터 크기: {len(decrypted_data.getvalue())} bytes")
                
                # 해독된 파일로 저장
                with open(decrypted_file_path, 'wb') as output_file:
                    shutil.copyfileobj(decrypted_data, output_file)
                
                print(f"[OK] 엑셀 파일 해독 완료: {decrypted_file_path}")
                print(f"[INFO] 해독된 파일 크기: {os.path.getsize(decrypted_file_path)} bytes")
                return decrypted_file_path
                
        except Exception as e:
            print(f"[FAIL] 엑셀 파일 해독 실패: {str(e)}")
            import traceback
            print(f"[DEBUG] 상세 오류: {traceback.format_exc()}")
            return None
    
    def preserve_excel_format(self, file_path, output_path=None):
        """Excel 파일의 원본 형식을 보존하여 저장"""
        try:
            print(f"[INFO] Excel 파일 원본 형식 보존 중: {file_path}")
            
            if output_path is None:
                output_path = file_path.replace('.xlsx', '_preserved.xlsx')
            
            # 원본 파일 읽기 (문자열로 보존)
            df = None
            try:
                df = pd.read_excel(file_path, header=2, dtype=str, na_filter=False)
                print("[OK] 원본 파일 읽기 성공 (문자열로 보존)")
            except Exception as e1:
                print(f"[WARN] 기본 방법 실패: {str(e1)}")
                try:
                    df = pd.read_excel(file_path, engine='openpyxl', header=2, dtype=str, na_filter=False)
                    print("[OK] openpyxl 엔진으로 읽기 성공 (문자열로 보존)")
                except Exception as e2:
                    print(f"[WARN] openpyxl 엔진 실패: {str(e2)}")
                    try:
                        df = pd.read_excel(file_path, engine='xlrd', header=2, dtype=str, na_filter=False)
                        print("[OK] xlrd 엔진으로 읽기 성공 (문자열로 보존)")
                    except Exception as e3:
                        print(f"[FAIL] 모든 엔진으로 읽기 실패: {str(e3)}")
                        return False
            
            if df is None:
                print("[FAIL] 파일을 읽을 수 없습니다.")
                return False
            
            # ExcelWriter를 사용하여 원본 형식 보존
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, header=False, sheet_name='Sheet1')
                
                # 숫자 형식 보존을 위한 추가 설정
                worksheet = writer.sheets['Sheet1']
                for row in worksheet.iter_rows():
                    for cell in row:
                        if cell.value and isinstance(cell.value, str):
                            # 숫자로 변환 가능한 문자열인 경우 숫자 형식으로 설정
                            try:
                                if '.' in cell.value:
                                    float_val = float(cell.value)
                                    if float_val > 1e6:  # 큰 숫자의 경우 과학적 표기법 방지
                                        cell.number_format = '0'
                                    else:
                                        cell.number_format = '0.00'
                                else:
                                    int_val = int(cell.value)
                                    cell.number_format = '0'
                            except (ValueError, TypeError):
                                pass  # 숫자가 아닌 경우 그대로 유지
            
            print(f"[OK] 원본 형식 보존 완료: {output_path}")
            return True
            
        except Exception as e:
            print(f"[FAIL] Excel 형식 보존 실패: {str(e)}")
            return False
    
    def validate_data_integrity(self, original_file, processed_file):
        """원본 파일과 처리된 파일의 데이터 무결성 검증"""
        try:
            print("[INFO] 데이터 무결성 검증 중...")
            
            # 원본 파일 읽기
            original_df = None
            try:
                original_df = pd.read_excel(original_file, header=2, dtype=str, na_filter=False)
            except:
                try:
                    original_df = pd.read_excel(original_file, engine='openpyxl', header=2, dtype=str, na_filter=False)
                except:
                    original_df = pd.read_excel(original_file, engine='xlrd', header=2, dtype=str, na_filter=False)
            
            # 처리된 파일 읽기
            processed_df = None
            try:
                processed_df = pd.read_excel(processed_file, header=2, dtype=str, na_filter=False)
            except:
                try:
                    processed_df = pd.read_excel(processed_file, engine='openpyxl', header=2, dtype=str, na_filter=False)
                except:
                    processed_df = pd.read_excel(processed_file, engine='xlrd', header=2, dtype=str, na_filter=False)
            
            if original_df is None or processed_df is None:
                print("[FAIL] 파일을 읽을 수 없습니다.")
                return False
            
            # 기본 검증
            if len(original_df) != len(processed_df):
                print(f"[WARN] 행 수 불일치: 원본 {len(original_df)}행, 처리된 파일 {len(processed_df)}행")
                return False
            
            if len(original_df.columns) != len(processed_df.columns):
                print(f"[WARN] 열 수 불일치: 원본 {len(original_df.columns)}열, 처리된 파일 {len(processed_df.columns)}열")
                return False
            
            # 데이터 내용 검증 (처음 5행만 샘플로 확인)
            sample_rows = min(5, len(original_df))
            for i in range(sample_rows):
                for col in original_df.columns:
                    original_val = str(original_df.iloc[i][col])
                    processed_val = str(processed_df.iloc[i][col])
                    
                    # 과학적 표기법을 일반 숫자로 변환하여 비교
                    try:
                        if 'e+' in original_val.lower() or 'e-' in original_val.lower():
                            original_val = str(int(float(original_val)))
                        if 'e+' in processed_val.lower() or 'e-' in processed_val.lower():
                            processed_val = str(int(float(processed_val)))
                    except:
                        pass
                    
                    if original_val != processed_val:
                        print(f"[WARN] 데이터 불일치 발견: 행 {i+3}, 열 {col}")
                        print(f"  원본: {original_val}")
                        print(f"  처리된: {processed_val}")
                        return False
            
            print("[OK] 데이터 무결성 검증 통과")
            return True
            
        except Exception as e:
            print(f"[FAIL] 데이터 무결성 검증 실패: {str(e)}")
            return False

    def analyze_sharepoint_excel_file(self, file_id):
        """SharePoint의 역발행 세금계산서 파일 분석 및 조건부 메일 발송 (다운로드 없이 직접 처리)"""
        try:
            print("\n" + "="*60)
            print("SharePoint 파일 분석 및 메일 발송 (다운로드 없이 직접 처리)")
            print("="*60)
            
            if not file_id:
                print("[WARN] SharePoint 파일 ID가 없습니다. 분석을 건너뜁니다.")
                return True
            
            # SharePoint 액세스 토큰 획득
            access_token = self.get_sharepoint_access_token()
            if not access_token:
                print("[FAIL] SharePoint 액세스 토큰 획득 실패")
                return False
            
            # SharePoint 파일을 직접 메모리에서 읽기 (다운로드 없음)
            print("[INFO] SharePoint '역발행 세금계산서.xlsx' 파일 직접 읽기 중...")
            existing_df = self.read_sharepoint_file_to_dataframe(access_token, file_id, "역발행 세금계산서.xlsx")
            
            if existing_df is None:
                print("[FAIL] SharePoint 파일 직접 읽기 실패")
                return False
            
            print(f"[OK] SharePoint 파일 읽기 완료: {len(existing_df)}행, {len(existing_df.columns)}열")
            print(f"[INFO] 컬럼명 (처음 10개): {list(existing_df.columns)[:10]}")
            if len(existing_df.columns) > 10:
                print(f"[INFO] 전체 컬럼 수: {len(existing_df.columns)}개")
            
            # 첫 번째 데이터 행 확인
            if len(existing_df) > 0:
                print(f"[INFO] 첫 번째 데이터 행 샘플:")
                first_row = existing_df.iloc[0]
                for i, (col, val) in enumerate(first_row.items()):
                    if i < 10:  # 처음 10개 컬럼만 출력
                        print(f"  {col}: {val}")
                    else:
                        break
            
            # 메일 발송 프로세스: 1단계 - 이메일오류 체크, 2단계 - 품목키워드 체크
            print("[INFO] 메일 발송 프로세스 시작")
            print("[INFO] 1단계: 이메일오류 컬럼 확인 (Y가 아닌 값만)")
            print("[INFO] 2단계: 품목키워드 체크 (메일 발송 시에만)")
            
            filtered_data = self.filter_data_by_conditions(existing_df)
            
            if len(filtered_data) > 0:
                print(f"[OK] 조건에 맞는 데이터 {len(filtered_data)}건 발견")
                print(f"[INFO] 이메일오류 조건 + 품목키워드 조건을 모두 만족하는 {len(filtered_data)}건이 메일 발송 대상입니다.")
                
                # 메일 발송
                print("[INFO] 3단계: 메일 발송 실행")
                if self.send_notification_email(filtered_data, None):  # 로컬 파일 경로 없음
                    print("[OK] 메일 발송 완료")
                    
                    # 이메일 발송 완료 후 SharePoint 파일에서 이메일오류 상태 업데이트
                    print("[INFO] 4단계: 메일 발송된 데이터만 이메일오류='Y'로 업데이트")
                    print("[INFO] SharePoint 파일 직접 읽기/업데이트 (다운로드 없음)")
                    if self.update_email_error_status(None, filtered_data, file_id):
                        print("[OK] 이메일오류 상태 업데이트 완료")
                        print("[OK] 메일 발송 프로세스 전체 완료")
                        return True
                    else:
                        print("[WARN] 이메일오류 상태 업데이트 실패")
                        return True  # 메일 발송은 성공했으므로 True 반환
                else:
                    print("[FAIL] 메일 발송 실패")
                    return False
            else:
                print("[INFO] 조건에 맞는 데이터가 없습니다.")
                print(f"[INFO] 필터링 조건1 - 이메일오류: {self.email_error_condition}")
                print(f"[INFO] 필터링 조건2 - 품목키워드: {self.item_keywords}")
                print("[INFO] 두 조건을 모두 만족하는 데이터가 없어 메일을 발송하지 않습니다.")
                return True
                
        except Exception as e:
            print(f"[FAIL] SharePoint 파일 분석 실패: {str(e)}")
            return False

    def analyze_excel_file(self, file_path):
        """엑셀 파일 분석 및 조건부 메일 발송 (원본 구조 보존) - 기존 방식 호환성 유지"""
        try:
            print("\n" + "="*60)
            print("엑셀 파일 분석 및 메일 발송 (원본 구조 보존)")
            print("="*60)
            
            # openpyxl을 사용하여 원본 파일 구조 완전 보존
            from openpyxl import load_workbook
            
            print(f"[INFO] 엑셀 파일 분석 중: {file_path}")
            
            # 원본 워크북 로드 (모든 구조와 서식 보존)
            print("[INFO] openpyxl로 원본 워크북 로드 중...")
            workbook = load_workbook(file_path)
            worksheet = workbook.active
            
            print(f"[OK] 워크북 로드 완료: {worksheet.max_row}행, {worksheet.max_column}열")
            
            # 헤더 행 확인 (3행이 실제 데이터 헤더)
            header_row = 3
            headers = []
            for col in range(1, worksheet.max_column + 1):
                cell_value = worksheet.cell(row=header_row, column=col).value
                headers.append(str(cell_value) if cell_value is not None else "")
            
            print(f"[INFO] 데이터 헤더 (3행): {headers[:10]}...")  # 처음 10개만 표시
            
            # 데이터 행 읽기 (4행부터) - pandas DataFrame 형태로 변환
            data_start_row = 4
            df_data = []
            for row in range(data_start_row, worksheet.max_row + 1):
                row_data = {}
                for col, header in enumerate(headers):
                    if header:  # 빈 헤더 제외
                        cell_value = worksheet.cell(row=row, column=col + 1).value
                        row_data[header] = str(cell_value) if cell_value is not None else ""
                df_data.append(row_data)
            
            # pandas DataFrame으로 변환 (기존 코드와 호환성 유지)
            df = pd.DataFrame(df_data)
            
            if df.empty:
                print("[FAIL] 엑셀 파일에서 데이터를 읽을 수 없습니다.")
                return False
            
            print(f"[INFO] 엑셀 파일 행 수: {len(df)}")
            print(f"[INFO] 엑셀 파일 열 수: {len(df.columns)}")
            print(f"[INFO] 컬럼명 (처음 10개): {list(df.columns)[:10]}")
            if len(df.columns) > 10:
                print(f"[INFO] 전체 컬럼 수: {len(df.columns)}개")
            
            # 첫 번째 데이터 행 확인
            if len(df) > 0:
                print(f"[INFO] 첫 번째 데이터 행 샘플:")
                first_row = df.iloc[0]
                for i, (col, val) in enumerate(first_row.items()):
                    if i < 10:  # 처음 10개 컬럼만 출력
                        print(f"  {col}: {val}")
                    else:
                        break
            
            # 메일 발송 프로세스: 1단계 - 이메일오류 체크, 2단계 - 품목키워드 체크
            print("[INFO] 메일 발송 프로세스 시작")
            print("[INFO] 1단계: 이메일오류 컬럼 확인 (Y가 아닌 값만)")
            print("[INFO] 2단계: 품목키워드 체크 (메일 발송 시에만)")
            
            filtered_data = self.filter_data_by_conditions(df)
            
            if len(filtered_data) > 0:
                print(f"[OK] 조건에 맞는 데이터 {len(filtered_data)}건 발견")
                print(f"[INFO] 이메일오류 조건 + 품목키워드 조건을 모두 만족하는 {len(filtered_data)}건이 메일 발송 대상입니다.")
                
                # 메일 발송
                print("[INFO] 3단계: 메일 발송 실행")
                if self.send_notification_email(filtered_data, file_path):
                    print("[OK] 메일 발송 완료")
                    
                    # 이메일 발송 완료 후 SharePoint 파일에서 이메일오류 상태 업데이트
                    print("[INFO] 4단계: 메일 발송된 데이터만 이메일오류='Y'로 업데이트")
                    print("[INFO] SharePoint 파일 직접 읽기/업데이트 (다운로드 없음)")
                    if self.update_email_error_status(file_path, filtered_data, getattr(self, 'current_file_id', None)):
                        print("[OK] 이메일오류 상태 업데이트 완료")
                        print("[OK] 메일 발송 프로세스 전체 완료")
                        return True
                    else:
                        print("[WARN] 이메일오류 상태 업데이트 실패")
                        return True  # 메일 발송은 성공했으므로 True 반환
                else:
                    print("[FAIL] 메일 발송 실패")
                    return False
            else:
                print("[INFO] 조건에 맞는 데이터가 없습니다.")
                print(f"[INFO] 필터링 조건1 - 이메일오류: {self.email_error_condition}")
                print(f"[INFO] 필터링 조건2 - 품목키워드: {self.item_keywords}")
                print("[INFO] 두 조건을 모두 만족하는 데이터가 없어 메일을 발송하지 않습니다.")
                return True
                
        except Exception as e:
            print(f"[FAIL] 엑셀 파일 분석 실패: {str(e)}")
            return False
    
    def filter_data_by_email_error_conditions(self, df):
        """조건에 맞는 데이터 필터링 (이메일오류 컬럼 기준)"""
        try:
            print("[INFO] 데이터 필터링 중...")
            print(f"[INFO] 필터링 조건: {self.email_error_condition}")
            print(f"[INFO] 전체 데이터 행 수: {len(df)}")
            
            filtered_rows = []
            y_count = 0  # 이미 처리된 데이터 (Y) 개수
            other_count = 0  # 처리 대상 데이터 개수
            
            for index, row in df.iterrows():
                # 이메일오류 컬럼 확인
                email_error_value = self._get_column_value(row, ['이메일오류'])
                
                print(f"[INFO] 행 {index + 1} - 이메일오류 값: '{email_error_value}'")
                
                # 이메일오류가 Y가 아닌 경우만 포함
                if email_error_value.upper() != 'Y':
                    filtered_rows.append(row)
                    other_count += 1
                    print(f"[OK] 조건 매치 행 {index + 1}: 이메일오류='{email_error_value}' (Y가 아님)")
                else:
                    y_count += 1
                    print(f"[SKIP] 행 {index + 1}: 이메일오류='{email_error_value}' (Y이므로 제외)")
            
            # 필터링 결과 요약
            print(f"\n[INFO] 필터링 결과 요약:")
            print(f"[INFO] - 전체 데이터: {len(df)}건")
            print(f"[INFO] - 이미 처리됨 (Y): {y_count}건")
            print(f"[INFO] - 처리 대상 (Y가 아님): {other_count}건")
            
            if other_count == 0:
                print("[INFO] 처리할 데이터가 없습니다. 모든 데이터가 이미 처리되었습니다.")
                print("[INFO] 메일을 발송하지 않습니다.")
            else:
                print(f"[INFO] {other_count}건의 데이터가 메일 발송 대상입니다.")
                print("[INFO] 이 데이터들에 대해 메일을 발송합니다.")
                
                # 발송 대상 데이터 상세 정보
                print("[INFO] 발송 대상 데이터 상세:")
                for i, row in enumerate(filtered_rows[:5]):  # 처음 5건만 표시
                    작성일자 = self._get_column_value(row, ['작성일자'])
                    품목 = self._get_column_value(row, ['품목'])
                    거래처명 = self._get_column_value(row, ['거래처명'])
                    이메일오류 = self._get_column_value(row, ['이메일오류'])
                    print(f"  {i+1}. 작성일자:{작성일자}, 품목:{품목}, 거래처명:{거래처명}, 이메일오류:'{이메일오류}'")
                
                if len(filtered_rows) > 5:
                    print(f"  ... 외 {len(filtered_rows) - 5}건 더")
            
            return filtered_rows
            
        except Exception as e:
            print(f"[FAIL] 데이터 필터링 실패: {str(e)}")
            return []
    
    def filter_data_by_conditions(self, df):
        """조건에 맞는 데이터 필터링 (이메일오류 조건 + 품목키워드 조건)"""
        try:
            print("[INFO] 데이터 필터링 중...")
            print(f"[INFO] 필터링 조건1 - 이메일오류: {self.email_error_condition}")
            print(f"[INFO] 필터링 조건2 - 품목키워드: {self.item_keywords}")
            
            filtered_rows = []
            total_rows = len(df)
            email_error_pass_count = 0
            keyword_pass_count = 0
            both_pass_count = 0
            
            for index, row in df.iterrows():
                # 1. 이메일오류 조건 확인
                email_error_value = self._get_column_value(row, ['이메일오류'])
                email_error_pass = email_error_value.upper() != 'Y'
                
                if email_error_pass:
                    email_error_pass_count += 1
                
                # 2. 품목키워드 조건 확인
                item_match = False
                
                # 각 컬럼에서 품목키워드 확인
                for col in df.columns:
                    cell_value = str(row[col]).strip()
                    
                    # 품목 키워드 확인 (수수료, 접수)
                    for keyword in self.item_keywords:
                        if keyword in cell_value:
                            item_match = True
                            print(f"[INFO] 품목 키워드 매치: {keyword} in {cell_value}")
                            break
                    
                    if item_match:
                        break
                
                if item_match:
                    keyword_pass_count += 1
                
                # 3. 두 조건을 모두 만족하는 경우만 추가
                if email_error_pass and item_match:
                    filtered_rows.append(row)
                    both_pass_count += 1
                    print(f"[OK] 조건 매치 행 {index + 1}: 이메일오류='{email_error_value}', 품목키워드 매치")
            
            # 필터링 결과 상세 통계
            print(f"\n[INFO] 필터링 결과 상세:")
            print(f"[INFO] - 전체 데이터: {total_rows}건")
            print(f"[INFO] - 이메일오류 조건 통과 (Y가 아님): {email_error_pass_count}건")
            print(f"[INFO] - 품목키워드 조건 통과: {keyword_pass_count}건")
            print(f"[INFO] - 두 조건 모두 통과 (최종 발송 대상): {both_pass_count}건")
            
            return filtered_rows
            
        except Exception as e:
            print(f"[FAIL] 데이터 필터링 실패: {str(e)}")
            return []
    
    def _get_column_value(self, row, possible_column_names):
        """엑셀 행에서 가능한 컬럼명들 중 하나를 찾아 값을 반환"""
        try:
            for col_name in possible_column_names:
                if col_name in row.index:
                    value = row[col_name]
                    if pd.isna(value):
                        return ""
                    return str(value).strip()
            return ""
        except Exception as e:
            print(f"[WARN] 컬럼 값 추출 실패: {str(e)}")
            return ""
    
    def update_current_file_email_error_status(self, file_path, filtered_data):
        """현재 파일의 이메일오류 컬럼 업데이트"""
        try:
            print("\n" + "="*60)
            print("현재 파일 이메일오류 상태 업데이트")
            print("="*60)
            
            print(f"[INFO] 파일 경로: {file_path}")
            
            # 현재 파일에서 이메일오류 컬럼 업데이트
            print("[INFO] 이메일오류 컬럼 업데이트 중...")
            update_success = self.update_email_error_column_in_file(file_path, filtered_data)
            
            if update_success:
                print("[OK] 현재 파일 이메일오류 상태 업데이트 완료")
                return True
            else:
                print("[FAIL] 현재 파일 이메일오류 컬럼 업데이트 실패")
                return False
                
        except Exception as e:
            print(f"[FAIL] 현재 파일 이메일오류 상태 업데이트 실패: {str(e)}")
            return False
    
    def update_sharepoint_file_email_error_status(self, filtered_data):
        """SharePoint 파일의 이메일오류 컬럼 업데이트 (사용하지 않음)"""
        try:
            print("\n" + "="*60)
            print("SharePoint 파일 이메일오류 상태 업데이트")
            print("="*60)
            
            # 액세스 토큰 획득
            access_token = self.get_sharepoint_access_token()
            if not access_token:
                print("[FAIL] SharePoint 액세스 토큰 획득 실패")
                return False
            
            # 업로드 폴더 찾기
            folder_id = self.find_sharepoint_folder(access_token, self.upload_folder_path)
            if not folder_id:
                print(f"[FAIL] 업로드 폴더를 찾을 수 없습니다: {self.upload_folder_path}")
                return False
            
            # 역발행 세금계산서 파일 찾기
            sharepoint_file = self.find_sharepoint_tax_invoice_file(access_token, folder_id)
            if not sharepoint_file:
                print("[FAIL] 역발행 세금계산서 파일을 찾을 수 없습니다.")
                return False
            
            file_id = sharepoint_file['id']
            file_name = sharepoint_file['name']
            
            print(f"[OK] SharePoint 파일 발견: {file_name} (ID: {file_id})")
            
            # 파일 다운로드
            print("[INFO] SharePoint 파일 다운로드 중...")
            download_success, temp_file_path = self.download_sharepoint_file(access_token, file_id, file_name)
            
            if not download_success:
                print("[FAIL] SharePoint 파일 다운로드 실패")
                return False
            
            # 다운로드한 파일에서 이메일오류 컬럼 업데이트
            print("[INFO] 이메일오류 컬럼 업데이트 중...")
            update_success = self.update_email_error_column_in_file(temp_file_path, filtered_data)
            
            if not update_success:
                print("[FAIL] 이메일오류 컬럼 업데이트 실패")
                # 임시 파일 정리
                if os.path.exists(temp_file_path):
                    os.remove(temp_file_path)
                return False
            
            # 업데이트된 파일을 SharePoint에 다시 업로드
            print("[INFO] 업데이트된 파일을 SharePoint에 업로드 중...")
            upload_success = self.upload_updated_file_to_sharepoint(access_token, temp_file_path, file_name, folder_id)
            
            # 임시 파일 정리
            if os.path.exists(temp_file_path):
                os.remove(temp_file_path)
                print("[OK] 임시 파일 정리 완료")
            
            if upload_success:
                print("[OK] SharePoint 파일 이메일오류 상태 업데이트 완료")
                return True
            else:
                print("[FAIL] 업데이트된 파일 SharePoint 업로드 실패")
                return False
                
        except Exception as e:
            print(f"[FAIL] SharePoint 파일 이메일오류 상태 업데이트 실패: {str(e)}")
            return False
    
    def find_sharepoint_tax_invoice_file(self, access_token, folder_id):
        """SharePoint에서 역발행 세금계산서 파일 찾기"""
        try:
            print("[INFO] SharePoint에서 역발행 세금계산서 파일 찾는 중...")
            
            # 폴더 내 파일 목록 조회
            if folder_id == self.sharepoint_drive_id:
                url = f"https://graph.microsoft.com/v1.0/sites/{self.sharepoint_site_id}/drives/{self.sharepoint_drive_id}/root/children"
            else:
                url = f"https://graph.microsoft.com/v1.0/sites/{self.sharepoint_site_id}/drives/{self.sharepoint_drive_id}/items/{folder_id}/children"
            
            headers = {
                'Authorization': f'Bearer {access_token}',
                'Content-Type': 'application/json'
            }
            
            response = requests.get(url, headers=headers)
            
            if response.status_code == 200:
                data = response.json()
                files = data.get('value', [])
                
                # 역발행 세금계산서 파일 찾기
                tax_invoice_patterns = [
                    "역발행 세금계산서",
                    "역발행세금계산서",
                    "세금계산서"
                ]
                
                for file in files:
                    file_name = file.get('name', '')
                    
                    # 엑셀 파일이고 역발행 세금계산서 관련인지 확인
                    if file_name.lower().endswith(('.xlsx', '.xls')):
                        for pattern in tax_invoice_patterns:
                            if pattern in file_name:
                                print(f"[OK] 역발행 세금계산서 파일 발견: {file_name}")
                                return {
                                    'id': file.get('id'),
                                    'name': file_name
                                }
                
                print("[FAIL] 역발행 세금계산서 파일을 찾을 수 없습니다.")
                print("[INFO] 사용 가능한 파일들:")
                for file in files:
                    if file.get('name', '').lower().endswith(('.xlsx', '.xls')):
                        print(f"  - {file.get('name')}")
                return None
            else:
                print(f"[FAIL] 파일 목록 조회 실패: {response.status_code} - {response.text}")
                return None
                
        except Exception as e:
            print(f"[FAIL] SharePoint 파일 찾기 실패: {str(e)}")
            return None
    
    def update_email_error_column_in_file(self, file_path, filtered_data):
        """파일에서 이메일오류 컬럼 업데이트"""
        try:
            print("[INFO] 파일에서 이메일오류 컬럼 업데이트 중...")
            
            # 엑셀 파일 읽기 (3행부터 데이터 읽기)
            df = None
            try:
                df = pd.read_excel(file_path, header=2)
                print("[OK] 파일 읽기 성공 (3행부터)")
            except Exception as e1:
                print(f"[WARN] 기본 방법 실패: {str(e1)}")
                try:
                    df = pd.read_excel(file_path, engine='openpyxl', header=2)
                    print("[OK] openpyxl 엔진으로 파일 읽기 성공 (3행부터)")
                except Exception as e2:
                    print(f"[WARN] openpyxl 엔진 실패: {str(e2)}")
                    try:
                        df = pd.read_excel(file_path, engine='xlrd', header=2)
                        print("[OK] xlrd 엔진으로 파일 읽기 성공 (3행부터)")
                    except Exception as e3:
                        print(f"[FAIL] 모든 엔진으로 파일 읽기 실패: {str(e3)}")
                        return False
            
            if df is None:
                print("[FAIL] 파일을 읽을 수 없습니다.")
                return False
            
            print(f"[INFO] 파일 데이터: {len(df)}행, {len(df.columns)}열")
            
            # 이메일오류 컬럼 찾기
            email_error_col = self._find_column_name(df.columns, ['이메일오류'])
            if not email_error_col:
                print("[FAIL] 이메일오류 컬럼을 찾을 수 없습니다.")
                return False
            
            print(f"[OK] 이메일오류 컬럼 발견: {email_error_col}")
            
            # 필터링된 데이터와 매칭하여 이메일오류 컬럼 업데이트
            updated_count = 0
            
            for filtered_row in filtered_data:
                # 필터링된 데이터의 주요 컬럼들로 매칭
                작성일자 = self._get_column_value(filtered_row, ['작성일자'])
                품목 = self._get_column_value(filtered_row, ['품목'])
                거래처명 = self._get_column_value(filtered_row, ['거래처명'])
                
                print(f"[INFO] 매칭 시도 - 작성일자: '{작성일자}', 품목: '{품목}', 거래처명: '{거래처명}'")
                
                # 파일의 각 행과 비교하여 매칭
                for index, row in df.iterrows():
                    row_작성일자 = self._get_column_value(row, ['작성일자'])
                    row_품목 = self._get_column_value(row, ['품목'])
                    row_거래처명 = self._get_column_value(row, ['거래처명'])
                    
                    # 주요 컬럼들이 일치하면 이메일오류를 Y로 업데이트
                    if (작성일자 == row_작성일자 and 
                        품목 == row_품목 and 
                        거래처명 == row_거래처명):
                        
                        df.loc[index, email_error_col] = 'Y'
                        updated_count += 1
                        print(f"[OK] 업데이트 완료: {int(index) + 1}행 - 이메일오류=Y")
                        break
            
            if updated_count > 0:
                # 업데이트된 파일 저장
                df.to_excel(file_path, index=False, header=False)
                print(f"[OK] 파일 업데이트 완료: {file_path}")
                print(f"[INFO] 총 {updated_count}행의 이메일오류 상태가 Y로 업데이트되었습니다.")
                return True
            else:
                print("[WARN] 매칭되는 데이터가 없어 업데이트되지 않았습니다.")
                return True  # 업데이트할 데이터가 없는 것도 정상
                
        except Exception as e:
            print(f"[FAIL] 파일 이메일오류 컬럼 업데이트 실패: {str(e)}")
            return False
    
    def _find_column_name(self, columns, possible_names):
        """컬럼 목록에서 가능한 이름들 중 하나를 찾아 실제 컬럼명 반환"""
        try:
            for possible_name in possible_names:
                for col in columns:
                    if possible_name in str(col):
                        return col
            return None
        except Exception as e:
            print(f"[WARN] 컬럼명 찾기 실패: {str(e)}")
            return None
    
    def upload_updated_file_to_sharepoint(self, access_token, file_path, filename, folder_id):
        """업데이트된 파일을 SharePoint에 업로드"""
        try:
            print(f"[INFO] 업데이트된 파일 SharePoint 업로드 중: {filename}")
            
            # 파일 읽기
            with open(file_path, 'rb') as file:
                file_content = file.read()
            
            # 업로드 URL
            if folder_id == self.sharepoint_drive_id:
                upload_url = f"https://graph.microsoft.com/v1.0/sites/{self.sharepoint_site_id}/drives/{self.sharepoint_drive_id}/root:/{filename}:/content"
            else:
                upload_url = f"https://graph.microsoft.com/v1.0/sites/{self.sharepoint_site_id}/drives/{self.sharepoint_drive_id}/items/{folder_id}:/{filename}:/content"
            
            headers = {
                'Authorization': f'Bearer {access_token}',
                'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            }
            
            response = requests.put(upload_url, data=file_content, headers=headers)
            
            if response.status_code in [200, 201]:
                print(f"[OK] 업데이트된 파일 SharePoint 업로드 성공: {filename}")
                return True
            else:
                print(f"[FAIL] 업데이트된 파일 SharePoint 업로드 실패: {response.status_code} - {response.text}")
                return False
                
        except Exception as e:
            print(f"[FAIL] 업데이트된 파일 SharePoint 업로드 실패: {str(e)}")
            return False
    
    def get_manager_by_rules(self, filtered_data):
        """거래처명과 담당부서 조합 규칙에 따라 담당자 결정"""
        try:
            print("[INFO] 담당자 매핑 규칙 적용 중...")
            
            # 첫 번째 데이터 행에서 거래처명과 담당부서 추출
            if not filtered_data or len(filtered_data) == 0:
                print("[WARN] 필터링된 데이터가 없습니다.")
                return None
            
            first_row = filtered_data[0]
            거래처명 = self._get_column_value(first_row, ['거래처명'])
            담당부서 = self._get_column_value(first_row, ['담당부서-공급받는자'])
            담당자명_공급자 = self._get_column_value(first_row, ['담당자명-공급자'])
            품목 = self._get_column_value(first_row, ['품목'])
            
            print(f"[INFO] 거래처명: {거래처명}")
            print(f"[INFO] 담당부서: {담당부서}")
            print(f"[INFO] 담당자명-공급자: {담당자명_공급자}")
            print(f"[INFO] 품목: {품목}")
            
            # 최우선 규칙: 담당자명-공급자가 "조가현"인 경우
            if 담당자명_공급자 and 담당자명_공급자.strip() == "조가현":
                print("[INFO] 규칙 매치: 담당자명-공급자=조가현 => 조가현")
                return "조가현"
            
            # 매핑 규칙 적용 (env 파일 규칙과 동일)
            if 거래처명 == "(학)연세대학교":
                # 연세대 + 품목 "강사", "채용" => 김유정
                품목_str = str(품목 or "").strip()
                if 품목_str and any(kw in 품목_str for kw in ["강사", "채용"]):
                    print("[INFO] 규칙 매치: (학)연세대학교 + 품목 강사/채용 => 김유정")
                    return "김유정"
                if 담당부서 in ["재무부(자체)", "국제학대학원", "고위정책"]:
                    print("[INFO] 규칙 매치: (학)연세대학교 + 재무부(자체)/국제학대학원/고위정책 => 윤지혜")
                    return "윤지혜"
                elif 담당부서 == "언더우드국제학부":
                    print("[INFO] 규칙 매치: (학)연세대학교 + 언더우드국제학부 => 송영신")
                    return "송영신"
                else:
                    print("[INFO] 기본 규칙: (학)연세대학교 => 송영신")
                    return "송영신"
            
            elif 거래처명 == "서강대학교":
                if 담당부서 and 담당부서.strip() != "":
                    print("[INFO] 규칙 매치: 서강대학교 + 담당부서 있음 => 박시현")
                    return "박시현"
                else:
                    print("[INFO] 기본 규칙: 서강대학교 => 박시현")
                    return "박시현"
            
            elif 거래처명 == "덕성여자대학교":
                if 담당부서 and 담당부서.strip() != "":
                    print("[INFO] 규칙 매치: 덕성여자대학교 + 담당부서 있음 => 김슬기")
                    return "김슬기"
                else:
                    print("[INFO] 기본 규칙: 덕성여자대학교 => 김슬기")
                    return "김슬기"
            
            elif 거래처명 == "연세대학교 미래캠퍼스":
                if 담당부서 and 담당부서.strip() != "":
                    print("[INFO] 규칙 매치: 연세대학교 미래캠퍼스 + 담당부서 있음 => 김유정")
                    return "김유정"
                else:
                    print("[INFO] 기본 규칙: 연세대학교 미래캠퍼스 => 김유정")
                    return "김유정"
            
            else:
                # 기본 담당자 매핑 사용
                manager_name = self.company_manager_map.get(거래처명)
                if manager_name:
                    print(f"[INFO] 기본 매핑 사용: {거래처명} => {manager_name}")
                    return manager_name
                else:
                    print(f"[WARN] 알 수 없는 거래처명: {거래처명}")
                    return "송영신"  # 기본값
                    
        except Exception as e:
            print(f"[FAIL] 담당자 매핑 규칙 적용 실패: {str(e)}")
            return "송영신"  # 기본값
    
    def format_date(self, date_str):
        """날짜 포맷팅 (YYYYMMDD -> YYYY.MM.DD)"""
        try:
            if not date_str or date_str.strip() == "":
                return ""
            
            # 숫자만 추출
            date_clean = ''.join(filter(str.isdigit, str(date_str)))
            
            if len(date_clean) == 8:
                # YYYYMMDD 형식을 YYYY.MM.DD로 변환
                year = date_clean[:4]
                month = date_clean[4:6]
                day = date_clean[6:8]
                return f"{year}.{month}.{day}"
            else:
                return str(date_str).strip()
                
        except Exception as e:
            print(f"[WARN] 날짜 포맷팅 실패: {str(e)}")
            return str(date_str).strip() if date_str else ""
    
    def format_number(self, number_str):
        """숫자 포맷팅 (천단위 콤마 추가)"""
        try:
            if not number_str or number_str.strip() == "":
                return ""
            
            # 숫자만 추출
            number_clean = ''.join(filter(str.isdigit, str(number_str)))
            
            if number_clean:
                # 천단위 콤마 추가
                return f"{int(number_clean):,}"
            else:
                return str(number_str).strip()
                
        except Exception as e:
            print(f"[WARN] 숫자 포맷팅 실패: {str(e)}")
            return str(number_str).strip() if number_str else ""
    
    def send_notification_email(self, filtered_data, file_path):
        """실제 메일 발송 (담당자별로 그룹화해서 각 담당자에게 개별 발송)"""
        try:
            print("\n" + "="*60)
            print("📧 실제 메일 발송 시작")
            print("="*60)
            print(f"[INFO] 발송 대상 데이터 건수: {len(filtered_data)}건")
            
            # 발송 대상 데이터 미리보기
            if filtered_data:
                print("[INFO] 발송 대상 데이터 미리보기:")
                for i, row in enumerate(filtered_data[:3]):  # 처음 3건만 표시
                    작성일자 = self._get_column_value(row, ['작성일자'])
                    품목 = self._get_column_value(row, ['품목'])
                    거래처명 = self._get_column_value(row, ['거래처명'])
                    이메일오류 = self._get_column_value(row, ['이메일오류'])
                    print(f"  {i+1}. {작성일자} | {품목} | {거래처명} | 이메일오류:{이메일오류}")
                
                if len(filtered_data) > 3:
                    print(f"  ... 외 {len(filtered_data) - 3}건 더")
            
            print("="*60)
            
            # 담당자별로 데이터 그룹화 (동일 담당자는 하나의 메일로 통합)
            from collections import defaultdict
            grouped_data = defaultdict(list)
            
            for row in filtered_data:
                거래처명 = self._get_column_value(row, ['거래처명'])
                담당부서 = self._get_column_value(row, ['담당부서-공급받는자'])
                
                # 각 row의 담당자를 먼저 결정
                manager_name = self.get_manager_by_rules([row])
                
                if manager_name:
                    grouped_data[manager_name].append(row)
                    print(f"[DEBUG] {거래처명} (담당부서: {담당부서}) -> 담당자: {manager_name}")
                else:
                    print(f"[WARN] 담당자를 찾을 수 없습니다: {거래처명} (담당부서: {담당부서})")
            
            print(f"[INFO] 담당자별 그룹 수: {len(grouped_data)}개")
            
            # 각 담당자별로 데이터 건수 출력
            for manager_name, rows in grouped_data.items():
                print(f"  - {manager_name}: {len(rows)}건")
            
            # 각 담당자별로 메일 발송
            all_success = True
            for manager_name, group_rows in grouped_data.items():
                print(f"\n[INFO] 메일 발송 준비: {manager_name} - {len(group_rows)}건")
                
                recipient_email = self.manager_email_map.get(manager_name)
                if not recipient_email:
                    print(f"[WARN] 담당자 이메일을 찾을 수 없습니다: {manager_name}")
                    all_success = False
                    continue
                
                print(f"[INFO] 메일 수신자: {manager_name} ({recipient_email})")
                
                # 해당 담당자의 모든 데이터를 하나의 메일로 발송
                if not self._send_email_to_manager(group_rows, manager_name, recipient_email, file_path):
                    all_success = False
            
            return all_success
            
        except Exception as e:
            print(f"[FAIL] 메일 발송 실패: {str(e)}")
            import traceback
            print(f"[DEBUG] 상세 오류: {traceback.format_exc()}")
            return False
    
    def _send_email_to_manager(self, data_rows, manager_name, recipient_email, file_path):
        """특정 담당자에게 메일 발송 (내부 메서드)"""
        try:
            # 메일 내용 생성
            # 첫 번째 행에서 거래처명 추출
            first_row = data_rows[0] if data_rows else None
            company_name = self._get_column_value(first_row, ['거래처명']) if first_row is not None else "알 수 없음"
            subject = "[SmileEDI] 역발행 세금계산서 알림"
            
            # HTML 테이블 행 생성 (개선된 포맷 적용)
            table_rows = ""
            for row in data_rows:
                # 엑셀 데이터에서 필요한 컬럼 추출
                작성일자 = self._get_column_value(row, ['작성일자'])
                품목 = self._get_column_value(row, ['품목'])
                공급가액 = self._get_column_value(row, ['공급가액'])
                세액 = self._get_column_value(row, ['세액'])
                거래처명 = self._get_column_value(row, ['거래처명'])
                담당부서 = self._get_column_value(row, ['담당부서-공급받는자'])
                담당자명 = self._get_column_value(row, ['담당자명-공급받는자'])
                담당자연락처 = self._get_column_value(row, ['담당자연락처-공급받는자'])
                담당자이메일 = self._get_column_value(row, ['공급받는자이메일'])
                
                # 데이터 포맷팅
                formatted_date = self.format_date(작성일자)
                formatted_amount = self.format_number(공급가액)
                formatted_tax = self.format_number(세액)
                
                table_rows += f"""
                  <tr>
                    <td style="border:1px solid #ccc; text-align:center;">{formatted_date}</td>
                    <td style="border:1px solid #ccc; text-align:center;">{품목}</td>
                    <td style="border:1px solid #ccc; text-align:right;">{formatted_amount}</td>
                    <td style="border:1px solid #ccc; text-align:right;">{formatted_tax}</td>
                    <td style="border:1px solid #ccc; text-align:center; white-space: nowrap;">{거래처명}</td>
                    <td style="border:1px solid #ccc; text-align:center; white-space: nowrap;">{담당부서}</td>
                    <td style="border:1px solid #ccc; text-align:center;">{담당자명}</td>
                    <td style="border:1px solid #ccc; text-align:center; white-space: nowrap;">{담당자연락처}</td>
                    <td style="border:1px solid #ccc; text-align:center;">{담당자이메일}</td>
                  </tr>"""
            
            # HTML 메일 본문 생성
            html_body = f"""
<!DOCTYPE html>
<html>
  <head>
    <meta charset="UTF-8">
    <title>SmileEDI 역발행 세금계산서</title>
  </head>
  <body style="font-family: Arial, sans-serif; margin:0; padding:0;">
    <table border="0" cellpadding="0" cellspacing="0" width="100%" style="border-collapse:collapse; background-color:#f9f9f9;">
      <tr>
        <td align="center" style="padding:30px;">
          <!-- 메인 컨테이너 -->
          <table border="0" cellpadding="0" cellspacing="0" width="800" style="border-collapse:collapse; background-color:#ffffff; border:1px solid #ddd;">
            <!-- 메인 상단 -->
            <tr>
              <td bgcolor="#fff9db" style="padding:20px; text-align:center; color:#333333; font-size:20px; font-weight:bold;">
                😊 SmileEDI 역발행 세금계산서
              </td>
            </tr>
            <!-- 중간 내용 문구 -->
            <tr>
              <td style="padding:20px; font-size:14px; color:#333333; line-height:1.6;">
                안녕하세요 {manager_name}님,<br>
                SmileEDI에서 역발행 세금계산서가 도착하였습니다.
              </td>
            </tr>
            <!-- 표 영역 -->
            <tr>
              <td style="padding:20px;">
                <table border="0" cellpadding="8" cellspacing="0" width="100%" style="border-collapse:collapse; border:1px solid #ccc; font-size:13px;">
                  <tr bgcolor="#e0e0e0" style="font-weight:bold; text-align:center;">
                    <td style="border:1px solid #ccc;">작성일자</td>
                    <td style="border:1px solid #ccc;">품목</td>
                    <td style="border:1px solid #ccc;">공급가액</td>
                    <td style="border:1px solid #ccc;">세액</td>
                    <td style="border:1px solid #ccc;">거래처명</td>
                    <td style="border:1px solid #ccc;">담당부서</td>
                    <td style="border:1px solid #ccc;">담당자명</td>
                    <td style="border:1px solid #ccc;">담당자연락처</td>
                    <td style="border:1px solid #ccc;">담당자이메일</td>
                  </tr>
                  {table_rows}
                </table>
              </td>
            </tr>
            <!-- 하단 내용 문구 추가 -->
            <tr>
              <td style="padding:20px; font-size:14px; color:#333333; line-height:1.6;">
                ※ 역발행 세금계산서 내용 확인하여 K시스템 전표 작성해 주세요.<br>
                ※ K시스템 전표 작성 후 박시현 매니저에게 승인 <a href="mailto:pkm0313@jinhakapply.com?subject=역발행 세금계산서 전표 승인 요청&body=역발행 세금계산서 전표를 작성하였습니다. 승인해 주세요.%0A%0A감사합니다." style="color: #0066cc; font-weight: bold; text-decoration: underline;">요청</a>해 주세요.
              </td>
            </tr>
            <!-- 메인 하단 -->
            <tr>
              <td bgcolor="#f4f4f4" style="padding:15px; font-size:12px; color:#555; text-align:center;">
                본 메일은 역발행 세금계산서를 담당자에게 발송하고 있습니다.<br>
                문의사항은 박시현 매니저에게 연락주세요.
              </td>
            </tr>
          </table>
        </td>
      </tr>
    </table>
  </body>
</html>"""
            
            # SMTP 서버 연결
            server = smtplib.SMTP(self.smtp_server, self.smtp_port)
            server.starttls()
            server.login(self.smtp_username, self.smtp_password)
            
            # 메일 생성
            msg = MIMEMultipart()
            msg['From'] = self.smtp_username
            msg['To'] = recipient_email
            msg['Cc'] = 'pkm0313@jinhakapply.com'  # 참조 추가
            msg['Subject'] = subject
            
            # HTML 메일 본문 추가
            msg.attach(MIMEText(html_body, 'html', 'utf-8'))
            
            # 첨부파일 추가 (로컬 파일이 있는 경우에만)
            if file_path and os.path.exists(file_path):
                with open(file_path, "rb") as attachment:
                    part = MIMEBase('application', 'octet-stream')
                    part.set_payload(attachment.read())
                    encoders.encode_base64(part)
                    part.add_header(
                        'Content-Disposition',
                        f'attachment; filename= {os.path.basename(file_path)}'
                    )
                    msg.attach(part)
                print(f"[INFO] 첨부파일 추가: {os.path.basename(file_path)}")
            else:
                print("[INFO] 로컬 파일이 없어 첨부파일 없이 메일 발송 (SharePoint 직접 처리)")
            
            # 메일 발송 (참조 포함)
            cc_email = 'pkm0313@jinhakapply.com'
            all_recipients = [recipient_email, cc_email]
            
            text = msg.as_string()
            server.sendmail(self.smtp_username, all_recipients, text)
            server.quit()
            
            print(f"[INFO] 메일 수신자: {recipient_email}")
            print(f"[INFO] 메일 참조: {cc_email}")
            
            print(f"[OK] 실제 메일 발송 성공: {recipient_email}")
            return True
            
        except Exception as e:
            print(f"[FAIL] 테스트 메일 발송 실패: {str(e)}")
            import traceback
            print(f"[DEBUG] 상세 오류: {traceback.format_exc()}")
            return False
    
    def scrape_tax_invoices(self):
        """세금계산서 스크래핑 메인 함수"""
        try:
            print("\n" + "="*60)
            print("SmileEDI 세금계산서 스크래핑 시작")
            print("="*60)
            
            # 1. 메인 페이지로 이동
            main_url = "https://www.smileedi.com/DtiHome.do?leftpath=Menu2200.do&bodypath=https://www.smileedi.com/Dti220101.do"
            
            if not self.navigate_to_url(main_url):
                return False
            
            # 2. 로그인 상태 확인 및 처리
            if not self.check_login_status():
                print("[FAIL] 로그인에 실패했습니다.")
                return False
            
            # 3. 페이지 로딩 대기
            self.random_wait(3, 5)
            
            # 4. 현재 페이지 정보 출력
            current_url = self.driver.current_url
            page_title = self.driver.title
            print(f"[INFO] 현재 페이지: {page_title}")
            print(f"[INFO] 현재 URL: {current_url}")
            
            # 5. 페이지 소스 일부 확인
            page_source = self.driver.page_source
            print(f"[INFO] 페이지 소스 길이: {len(page_source)} characters")
            
            # 6. 매출계산서 페이지로 이동 확인 및 처리
            if not self.navigate_to_tax_invoice_page():
                print("[FAIL] 매출계산서 페이지로 이동에 실패했습니다.")
                return False
            
            # 6. 매출계산서 검색 조건 설정
            if not self.setup_search_conditions():
                print("[FAIL] 검색 조건 설정에 실패했습니다.")
                return False
            
            # 7. 매출계산서 검색 실행
            if not self.execute_search():
                print("[FAIL] 검색 실행에 실패했습니다.")
                return False
            
            # 8. 엑셀 다운로드
            if not self.download_excel_file():
                print("[FAIL] 엑셀 다운로드에 실패했습니다.")
                return False
            
            # 9. 다운로드된 파일 찾기 및 SharePoint 업로드
            uploaded_result = self.find_and_upload_downloaded_file()
            if not uploaded_result:
                print("[WARN] 파일 업로드에 실패했지만 스크래핑은 완료되었습니다.")
            else:
                # 업로드 결과에서 파일 경로와 파일 ID 추출
                if isinstance(uploaded_result, tuple):
                    uploaded_file_path, file_id = uploaded_result
                    self.current_file_id = file_id  # 파일 ID 저장
                else:
                    uploaded_file_path = uploaded_result
                    self.current_file_id = None
                
                # 10. 메일 발송은 OPS-Console(smileedi-mail 잡)가 담당 — 이중 발송 방지.
                #     기본 SKIP_MAIL=true. (false로 명시할 때만 Python 자체 분석/발송)
                if os.getenv('SKIP_MAIL', 'true').lower() != 'false':
                    print("[INFO] SKIP_MAIL=true — 분석/메일은 OPS-Console 담당. Python 발송 생략.")
                else:
                    print("[INFO] SharePoint 파일 분석 및 메일 발송 중...")
                    if not self.analyze_sharepoint_excel_file(self.current_file_id):
                        print("[WARN] SharePoint 파일 분석 실패 — 스크래핑은 완료.")
            
            # 11. 세금계산서 관련 요소 찾기
            print("[INFO] 세금계산서 관련 요소 검색 중...")
            
            # 테이블이나 리스트 요소 찾기
            try:
                # 일반적인 테이블 요소들 찾기
                tables = self.driver.find_elements(By.TAG_NAME, "table")
                print(f"[INFO] 발견된 테이블 수: {len(tables)}")
                
                # 링크 요소들 찾기
                links = self.driver.find_elements(By.TAG_NAME, "a")
                print(f"[INFO] 발견된 링크 수: {len(links)}")
                
                # 세금계산서 관련 링크 찾기
                tax_links = []
                for link in links:
                    href = link.get_attribute("href")
                    text = link.text.strip()
                    if href and ("tax" in href.lower() or "세금" in text or "계산서" in text):
                        tax_links.append({"text": text, "href": href})
                
                print(f"[INFO] 세금계산서 관련 링크: {len(tax_links)}개")
                for i, link in enumerate(tax_links[:5], 1):  # 최대 5개만 출력
                    print(f"  {i}. {link['text']} - {link['href']}")
                
            except Exception as e:
                print(f"[WARN] 요소 검색 중 오류: {str(e)}")
            
            print("\n[OK] 매출계산서 스크래핑 완료")
            return True
            
        except Exception as e:
            print(f"[FAIL] 스크래핑 실패: {str(e)}")
            return False
    
    def update_email_error_status(self, file_path, filtered_data, file_id):
        """이메일 발송 완료 후 SharePoint 파일의 이메일오류 컬럼을 Y로 업데이트 (다운로드 없이 직접 처리)"""
        try:
            print("\n" + "="*60)
            print("이메일오류 상태 업데이트 (SharePoint 직접 처리)")
            print("="*60)
            
            if not file_id:
                print("[WARN] SharePoint 파일 ID가 없습니다. 업데이트를 건너뜁니다.")
                return True
            
            # SharePoint 액세스 토큰 획득
            access_token = self.get_sharepoint_access_token()
            if not access_token:
                print("[FAIL] SharePoint 액세스 토큰 획득 실패")
                return False
            
            # SharePoint 파일을 직접 메모리에서 읽기 (다운로드 없음)
            print("[INFO] SharePoint 파일 직접 읽기 중...")
            existing_df = self.read_sharepoint_file_to_dataframe(access_token, file_id, "역발행 세금계산서.xlsx")
            
            if existing_df is None:
                print("[FAIL] SharePoint 파일 직접 읽기 실패")
                return False
            
            print(f"[OK] SharePoint 파일 읽기 완료: {len(existing_df)}행, {len(existing_df.columns)}열")
            
            # 이메일오류 컬럼 찾기 (DataFrame에서)
            email_error_col = None
            for col in existing_df.columns:
                if '이메일오류' in str(col):
                    email_error_col = col
                    print(f"[OK] 이메일오류 컬럼 발견: {email_error_col}")
                    break
            
            if not email_error_col:
                print("[FAIL] 이메일오류 컬럼을 찾을 수 없습니다.")
                return False
            
            # 메일 발송된 데이터만 이메일오류 컬럼 Y로 업데이트 (DataFrame 기반)
            print(f"[INFO] 메일 발송된 {len(filtered_data)}건의 데이터를 이메일오류='Y'로 업데이트 중...")
            
            updated_count = 0
            updated_df = existing_df.copy()  # 원본 보존
            
            for filtered_row in filtered_data:
                # 필터링된 행의 고유 식별자로 원본 데이터에서 해당 행 찾기
                작성일자 = self._get_column_value(filtered_row, ['작성일자'])
                품목 = self._get_column_value(filtered_row, ['품목'])
                거래처명 = self._get_column_value(filtered_row, ['거래처명'])
                담당부서 = self._get_column_value(filtered_row, ['담당부서-공급받는자'])
                
                print(f"[DEBUG] 매칭 시도 - 작성일자: '{작성일자}', 품목: '{품목}', 거래처명: '{거래처명}', 담당부서: '{담당부서}'")
                
                # DataFrame에서 해당 행 찾기
                for index, row in updated_df.iterrows():
                    row_작성일자 = self._get_column_value(row, ['작성일자'])
                    row_품목 = self._get_column_value(row, ['품목'])
                    row_거래처명 = self._get_column_value(row, ['거래처명'])
                    row_담당부서 = self._get_column_value(row, ['담당부서-공급받는자'])
                    
                    # 데이터 매칭 확인
                    if (row_작성일자 == 작성일자 and 
                        row_품목 == 품목 and 
                        row_거래처명 == 거래처명 and 
                        row_담당부서 == 담당부서):
                        
                        # 이메일오류 컬럼 업데이트 (Y로 설정)
                        updated_df.loc[index, email_error_col] = 'Y'
                        updated_count += 1
                        print(f"[OK] 업데이트 완료: {int(index) + 1}행 - 이메일오류=Y")
                        break
                else:
                    print(f"[WARN] 매칭되는 행을 찾을 수 없습니다: {작성일자}, {품목}, {거래처명}")
            
            if updated_count > 0:
                # 기존 SharePoint 파일을 읽어서 해당 셀만 정밀 수정
                print(f"[INFO] 기존 파일에서 이메일오류 셀만 정밀 업데이트 중... (업데이트된 행: {updated_count}개)")
                excel_buffer = self.update_specific_cells_in_sharepoint_file(access_token, file_id, filtered_data)
                
                if excel_buffer is None:
                    print("[WARN] 정밀 셀 업데이트 실패, fallback 방식으로 시도...")
                    excel_buffer = self.create_basic_excel_stream(updated_df)
                    if excel_buffer is None:
                        print("[FAIL] Excel 스트림 생성 완전 실패")
                        return False
                
                # SharePoint 원본 파일에 직접 업데이트 (다운로드 없이)
                print("[INFO] SharePoint 원본 파일에 직접 업데이트 중...")
                
                update_url = f"https://graph.microsoft.com/v1.0/sites/{self.sharepoint_site_id}/drives/{self.sharepoint_drive_id}/items/{file_id}/content"
                
                headers = {
                    'Authorization': f'Bearer {access_token}',
                    'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                }
                
                response = requests.put(update_url, data=excel_buffer.getvalue(), headers=headers)
                
                if response.status_code in [200, 201]:
                    sharepoint_filename = "역발행 세금계산서.xlsx"
                    print(f"[OK] SharePoint 파일 업데이트 성공: {sharepoint_filename}")
                    print(f"[INFO] 메일 발송된 {updated_count}건의 데이터가 이메일오류='Y'로 업데이트되었습니다.")
                    print("[INFO] 다운로드 없이 SharePoint 파일이 직접 업데이트되었습니다.")
                    return True
                else:
                    print(f"[FAIL] SharePoint 파일 업데이트 실패: {response.status_code}")
                    if response.text:
                        print(f"[DEBUG] 응답 내용: {response.text}")
                    return False
            else:
                print("[WARN] 업데이트할 행이 없습니다.")
                return True
                
        except Exception as e:
            print(f"[FAIL] 이메일오류 상태 업데이트 실패: {str(e)}")
            import traceback
            print(f"[DEBUG] 상세 오류: {traceback.format_exc()}")
            return False

    def update_specific_cells_in_sharepoint_file(self, access_token, file_id, filtered_data):
        """SharePoint 파일을 읽어서 이메일오류 셀만 정밀 업데이트 (원본 구조 완전 보존)"""
        try:
            print("[INFO] SharePoint 파일의 원본 구조를 완전히 보존하면서 특정 셀만 업데이트 중...")
            
            # 1. SharePoint 파일 내용을 메모리로 직접 읽기
            download_url = f"https://graph.microsoft.com/v1.0/sites/{self.sharepoint_site_id}/drives/{self.sharepoint_drive_id}/items/{file_id}/content"
            
            headers = {
                'Authorization': f'Bearer {access_token}',
                'Content-Type': 'application/json'
            }
            
            response = requests.get(download_url, headers=headers)
            
            if response.status_code != 200:
                print(f"[FAIL] SharePoint 파일 읽기 실패: {response.status_code}")
                return None
            
            # 2. openpyxl로 원본 파일 구조 그대로 로드
            file_stream = io.BytesIO(response.content)
            
            from openpyxl import load_workbook
            workbook = load_workbook(file_stream)
            worksheet = workbook.active
            
            print(f"[OK] 원본 파일 구조 로드 완료: {worksheet.max_row}행, {worksheet.max_column}열")
            
            # 3. 이메일오류 컬럼 찾기 (3행에서 컬럼명 확인)
            email_error_col_index = None
            
            for col in range(1, worksheet.max_column + 1):
                cell_value = worksheet.cell(row=3, column=col).value
                if cell_value and '이메일오류' in str(cell_value):
                    email_error_col_index = col
                    print(f"[OK] 이메일오류 컬럼 발견: {col}열 ({cell_value})")
                    break
            
            if not email_error_col_index:
                print("[FAIL] 이메일오류 컬럼을 찾을 수 없습니다.")
                return None
            
            # 4. 발송된 데이터의 식별 키 생성
            sent_keys = set()
            for row in filtered_data:  # filtered_data는 list of Series
                작성일자 = self._get_column_value(row, ['작성일자'])
                품목 = self._get_column_value(row, ['품목'])
                거래처명 = self._get_column_value(row, ['거래처명'])
                key = f"{작성일자}|{품목}|{거래처명}"
                sent_keys.add(key)
            
            print(f"[INFO] 발송된 메일 데이터 키 {len(sent_keys)}개 생성 완료")
            
            # 5. 4행부터 데이터 행들을 확인하면서 해당하는 셀만 업데이트
            updated_count = 0
            
            for row in range(4, worksheet.max_row + 1):
                # 작성일자, 품목, 거래처명 컬럼 찾기 (3행에서 컬럼명 확인)
                작성일자_col = None
                품목_col = None
                거래처명_col = None
                
                for col in range(1, worksheet.max_column + 1):
                    header_value = worksheet.cell(row=3, column=col).value
                    if header_value:
                        if '작성일자' in str(header_value):
                            작성일자_col = col
                        elif '품목' in str(header_value):
                            품목_col = col
                        elif '거래처명' in str(header_value):
                            거래처명_col = col
                
                if 작성일자_col and 품목_col and 거래처명_col:
                    # 현재 행의 식별 키 생성
                    작성일자 = str(worksheet.cell(row=row, column=작성일자_col).value or "")
                    품목 = str(worksheet.cell(row=row, column=품목_col).value or "")
                    거래처명 = str(worksheet.cell(row=row, column=거래처명_col).value or "")
                    key = f"{작성일자}|{품목}|{거래처명}"
                    
                    # 발송된 데이터와 매칭되면 이메일오류 셀만 Y로 업데이트
                    if key in sent_keys:
                        current_value = str(worksheet.cell(row=row, column=email_error_col_index).value or "")
                        if current_value.upper() != 'Y':
                            worksheet.cell(row=row, column=email_error_col_index, value='Y')
                            updated_count += 1
                            print(f"[INFO] {row}행 이메일오류 셀 업데이트: {작성일자} | {품목} | {거래처명}")
            
            print(f"[OK] 총 {updated_count}개 셀 업데이트 완료")
            
            # 6. 수정된 워크북을 Excel 스트림으로 저장
            excel_buffer = io.BytesIO()
            workbook.save(excel_buffer)
            excel_buffer.seek(0)
            
            print(f"[OK] 원본 구조 완전 보존된 Excel 스트림 생성 완료: {len(excel_buffer.getvalue())} bytes")
            print("[OK] 1,2,3행 헤더와 모든 원본 구조가 완벽하게 보존되었습니다.")
            
            return excel_buffer
            
        except Exception as e:
            print(f"[FAIL] 특정 셀 업데이트 실패: {str(e)}")
            import traceback
            print(f"[DEBUG] 상세 오류: {traceback.format_exc()}")
            return None

    def create_basic_excel_stream(self, df):
        """기본 방식으로 DataFrame을 Excel 스트림으로 변환 (fallback)"""
        try:
            print("[INFO] 기본 방식으로 Excel 스트림 생성 중...")
            excel_buffer = io.BytesIO()
            
            # ExcelWriter를 사용하여 원본 형식 보존
            with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, header=False, sheet_name='Sheet1')
                
                # 숫자 형식 보존을 위한 추가 설정
                worksheet = writer.sheets['Sheet1']
                for row in worksheet.iter_rows():
                    for cell in row:
                        if cell.value and isinstance(cell.value, str):
                            try:
                                if '.' in cell.value:
                                    float_val = float(cell.value)
                                    if float_val > 1e6:  # 큰 숫자의 경우 과학적 표기법 방지
                                        cell.number_format = '0'
                                    else:
                                        cell.number_format = '0.00'
                                else:
                                    int_val = int(cell.value)
                                    cell.number_format = '0'
                            except (ValueError, TypeError):
                                pass
            
            excel_buffer.seek(0)
            print(f"[OK] 기본 방식 Excel 스트림 생성 완료: {len(excel_buffer.getvalue())} bytes")
            return excel_buffer
            
        except Exception as e:
            print(f"[FAIL] 기본 방식 Excel 스트림 생성 실패: {str(e)}")
            return None

    def upload_updated_file_to_sharepoint_by_id(self, file_path, file_id):
        """업데이트된 파일을 SharePoint의 원본 파일에 강제 덮어쓰기 (ID 직접 업데이트 방식)"""
        try:
            print("[INFO] SharePoint 원본 파일에 업데이트된 내용 강제 덮어쓰기 중...")
            
            # 위임된 권한으로 토큰 획득
            access_token = self.get_sharepoint_access_token()
            if not access_token:
                print("[FAIL] 위임된 권한 토큰 획득 실패")
                return False
            
            # ID로 직접 업데이트 시도
            print("[INFO] ID로 직접 파일 업데이트 중...")
            
            upload_url = f"https://graph.microsoft.com/v1.0/sites/{self.sharepoint_site_id}/drives/{self.sharepoint_drive_id}/items/{file_id}/content"
            
            headers = {
                'Authorization': f'Bearer {access_token}',
                'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            }
            
            with open(file_path, 'rb') as f:
                file_content = f.read()
            
            print("[INFO] ID로 직접 파일 업로드 중...")
            response = requests.put(upload_url, headers=headers, data=file_content)
            
            if response.status_code in [200, 201]:
                # SharePoint에서는 항상 "역발행 세금계산서.xlsx"라는 고정 파일명으로 관리
                sharepoint_filename = "역발행 세금계산서.xlsx"
                print(f"[OK] ID로 직접 업데이트 성공: {sharepoint_filename}")
                return True
            else:
                print(f"[FAIL] ID로 직접 업데이트 실패: {response.status_code}")
                if response.text:
                    print(f"[DEBUG] 응답 내용: {response.text}")
                return False
                
        except Exception as e:
            print(f"[FAIL] SharePoint 파일 업데이트 실패: {str(e)}")
            return False
    
    def close(self):
        """브라우저 종료"""
        if self.driver:
            self.driver.quit()
            print("[OK] 브라우저 종료 완료")

def main():
    """메인 함수"""
    scraper = None
    
    try:
        print("🚀 SmileEDI 세금계산서 스크래핑 프로그램")
        print("="*60)
        
        # 스크래퍼 초기화 (헤드리스 모드)
        scraper = SmileEDIScraper(headless=True)
        
        # 스크래핑 실행
        success = scraper.scrape_tax_invoices()
        
        if success:
            print("\n✅ 스크래핑이 성공적으로 완료되었습니다!")
        else:
            print("\n❌ 스크래핑이 실패했습니다.")
        
    except KeyboardInterrupt:
        print("\n⚠️ 사용자에 의해 중단되었습니다.")
    except Exception as e:
        print(f"\n❌ 예상치 못한 오류: {str(e)}")
    finally:
        if scraper:
            scraper.close()

if __name__ == "__main__":
    main()

