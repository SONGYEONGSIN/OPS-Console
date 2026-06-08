#!/usr/bin/env python3
"""Moa 서비스조회 스크래퍼 → OPS-Console /api/closing/ingest.

설계: .claude/plans/20260607-moa-closing-scrape.md (Phase 2).

흐름:
  격주 게이트(off주 exit 0) → Chrome 기동 → Moa 로그인 → SMS 2FA(baseline-diff 폴링)
  → ServiceSearch 학년도 오픈일 범위 검색 → '엑셀저장' 다운로드 → 11컬럼 파싱
  → 작성마감 < 스크래핑시각 필터 → ISO8601(+09:00) 직렬화 → POST 인제스트.

결정된 로직(격주 게이트/학년도/SMS 폴링/인제스트/필터)은 TS 단일 소스
(features/closing/{biweekly-gate,academic-year}.ts)와 동치로 구현했다.

╔══════════════════════════════════════════════════════════════════════╗
║ 확정됨 (공개 로그인 페이지 캡처):                                          ║
║   - MOA_LOGIN_URL = https://moa.jinhakapply.com/User/Login (SSR/ASP.NET) ║
║   - 로그인/2FA 셀렉터: #txtUserID #txtPassWord #txtSANum #btnLogin        ║
║   - #btnLogin 이중용도: 1차=SMS발송(Validate), 2차=인증확인               ║
║   - 폼 POST /User/LoginProcess + __RequestVerificationToken(CSRF)        ║
║                                                                        ║
║ ⚠️ CAPTCHA(#secCaptcha/#txtCaptchaCode, img /User/Captcha):              ║
║   평소 display:none, 로그인 실패(Status<0) 후 노출. 헤드리스가 1회라도      ║
║   실패하면 캡차로 잠김 → 첫 시도 성공 필수. 본 스크립트는 캡차 감지 시 abort. ║
║                                                                        ║
║ 확정됨 (라이브 디스커버리):                                                ║
║   - ServiceSearch = /Foundation/ServiceSearch                           ║
║   - 오픈일 input: #txtOpenFromTime/#txtOpenToTime, 포맷 'YYYY-MM-DD HH:MM' ║
║   - ⚠️ ddlManager(운영자)가 로그인 운영자로 기본 선택 → ''(전체)로 비워야 함  ║
║   - 엑셀저장 = JS GetUnivServiceListToExcel() (#searchForm POST)          ║
║   - 엑셀은 암호화(CDFV2) → MOA_EXCEL_PASSWORD + msoffcrypto 복호           ║
║                                                                        ║
║ ✅ 전체 플로우 라이브 검증됨 (2026-06-07):                                 ║
║   - 자동 2FA: 웹훅 baseline-diff → SMS '[…] 인증번호는 [123456] 입니다'     ║
║     → 대괄호 숫자 추출 → 로그인 성공 (수동 입력 없음).                        ║
║   - 엑셀 복호 = VelvetSweatshop(기본키). 14컬럼 전부 매핑, 430건 파싱 OK.   ║
║                                                                        ║
║ 운영 전 남은 것: GH Secrets 등록 + 워크플로 1회 실행 검증 + 인제스트 실전송.   ║
║ 검증용: CLOSING_DRY_RUN=true → 추출까지만(인제스트 미전송).                  ║
╚══════════════════════════════════════════════════════════════════════╝
"""

import os
import re
import sys
import time
import glob
import tempfile
from datetime import datetime, timedelta, timezone, date

import requests
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
import undetected_chromedriver as uc

KST = timezone(timedelta(hours=9))

# ── URL / 셀렉터 (전부 라이브 디스커버리로 확정) ──────────────────────────
# 로그인/ServiceSearch 모두 SSR(ASP.NET MVC).
MOA_LOGIN_URL = os.getenv("MOA_LOGIN_URL", "https://moa.jinhakapply.com/User/Login")
SERVICE_SEARCH_URL = os.getenv(
    "SERVICE_SEARCH_URL", "https://moa.jinhakapply.com/Foundation/ServiceSearch"
)

SELECTORS = {
    # 로그인/2FA — #btnLogin 이중용도(1차=SMS발송 Validate(), 2차=인증확인).
    "login_id": os.getenv("SEL_LOGIN_ID", "#txtUserID"),
    "login_pw": os.getenv("SEL_LOGIN_PW", "#txtPassWord"),
    "login_submit": os.getenv("SEL_LOGIN_SUBMIT", "#btnLogin"),
    "sms_code_input": os.getenv("SEL_SMS_INPUT", "#txtSANum"),
    "sms_submit": os.getenv("SEL_SMS_SUBMIT", "#btnLogin"),
    # 캡차 — 로그인 실패 후에만 노출(#secCaptcha display:none). 첫 시도 성공 시 불필요.
    "captcha_section": os.getenv("SEL_CAPTCHA_SECTION", "#secCaptcha"),
    # ServiceSearch — 오픈일 범위 input(flatpickr, 포맷 'YYYY-MM-DD HH:MM').
    "open_from": os.getenv("SEL_OPEN_FROM", "txtOpenFromTime"),  # id (JS value 주입)
    "open_to": os.getenv("SEL_OPEN_TO", "txtOpenToTime"),
    # ⚠️ 운영자/개발자/카테고리 드롭다운 — 로그인 운영자로 기본 선택됨(예: 송영신).
    #    전체 마감 서비스를 받으려면 반드시 ''(선택=전체)로 비워야 함.
    "ddl_manager": os.getenv("SEL_DDL_MANAGER", "ddlManager"),
    "ddl_developer": os.getenv("SEL_DDL_DEVELOPER", "ddlDeveloper"),
    "ddl_category": os.getenv("SEL_DDL_CATEGORY", "ddlMOACategoryName"),
}

# 엑셀 다운로드 = JS 함수 호출(searchForm POST → 암호화 xlsx URL). 버튼 클릭 대신 직접 호출.
EXCEL_DOWNLOAD_JS = "GetUnivServiceListToExcel()"

# 엑셀 헤더 → 내부 키. 실 다운로드 확인: 14컬럼 전부 적재.
EXCEL_COLUMN_MAP = {
    "접수구분": "admission_type",
    "서비스ID": "service_id",
    "대학명": "university_name",
    "지역": "region",
    "서비스명": "service_name",
    "대학구분": "university_type",
    "카테고리": "category",
    "운영자": "operator_name",
    "개발자": "developer_name",
    "작성시작": "write_start_at",
    "작성마감": "write_end_at",
    "결제시작": "pay_start_at",
    "결제마감": "pay_end_at",
    "단독여부": "solo",
}

# Moa SMS 포맷 확인: "[Web발신][내부관리자] 본인확인 인증번호는 [123456] 입니다."
# → 코드가 대괄호 안. \[(\d+)\]가 정확히 매치(다른 [한글] 브라켓은 숫자 아니라 불매치).
SMS_CODE_PATTERN = re.compile(os.getenv("MOA_SMS_CODE_REGEX", r"\[(\d+)\]"))


# ── 결정된 로직 (TS 단일 소스와 동치) ──────────────────────────────────


def should_run_this_week(now: datetime, anchor_monday: str) -> bool:
    """biweekly-gate.ts shouldRunThisWeek 동치. anchor 경과 주 패리티."""

    def monday_of(d: date) -> date:
        return d - timedelta(days=d.weekday())  # Mon=0

    now_kst = now.astimezone(KST).date()
    this_monday = monday_of(now_kst)
    anchor = monday_of(date.fromisoformat(anchor_monday))
    diff_weeks = round((this_monday - anchor).days / 7)
    return diff_weeks % 2 == 0


def academic_year_range(now: datetime):
    """academic-year.ts academicYearRangeKST 동치. {start,end} (date,time)."""
    now_kst = now.astimezone(KST)
    start_year = now_kst.year if now_kst.month >= 3 else now_kst.year - 1
    end_year = start_year + 1
    is_leap = (end_year % 4 == 0 and end_year % 100 != 0) or end_year % 400 == 0
    end_day = 29 if is_leap else 28
    return {
        "start": {"date": f"{start_year}-03-01", "time": "00:01"},
        "end": {"date": f"{end_year}-02-{end_day:02d}", "time": "23:59"},
    }


def fetch_sms_code(url: str) -> str | None:
    """make 웹훅 GET → 응답=SMS 본문. 정규식으로 인증번호 추출(없으면 None)."""
    try:
        res = requests.get(url, timeout=10)
        res.raise_for_status()
    except requests.RequestException as e:
        print(f"[WARN] SMS GET 실패: {e}")
        return None
    m = SMS_CODE_PATTERN.search(res.text or "")
    return m.group(1) if m else None


def poll_fresh_sms_code(url: str, baseline: str | None, timeout_sec: int, interval_sec: int) -> str:
    """baseline-diff 폴링 — baseline과 달라진 새 코드를 반환. 타임아웃 시 raise."""
    deadline = time.monotonic() + timeout_sec
    while time.monotonic() < deadline:
        code = fetch_sms_code(url)
        if code and code != baseline:
            masked = ("*" * (len(code) - 2) + code[-2:]) if len(code) > 2 else "**"
            print(f"[OK] 새 SMS 코드 수신 (…{masked})")
            return code
        time.sleep(interval_sec)
    raise RuntimeError(f"SMS 코드 폴링 타임아웃 ({timeout_sec}s) — baseline 미변경")


def post_ingest(base_url: str, secret: str, scraped_at: str, rows: list[dict]) -> None:
    """POST /api/closing/ingest (Bearer CRON_SECRET). 빈 배열은 호출하지 않는다."""
    url = f"{base_url.rstrip('/')}/api/closing/ingest"
    res = requests.post(
        url,
        headers={"Authorization": f"Bearer {secret}", "Content-Type": "application/json"},
        json={"scraped_at": scraped_at, "rows": rows},
        timeout=30,
    )
    res.raise_for_status()
    print(f"[OK] 인제스트 완료: {res.json()}")


def post_run_log(
    base_url: str, secret: str, status: str, service_count: int, message: str
) -> None:
    """POST /api/closing/run-log — 실행 결과 보고(best-effort).

    OPS 인스펙터에 실행 흔적(success/skipped/failed)을 남기기 위함. 보고 실패가
    스크랩 성패에 영향 주지 않도록 예외를 삼킨다. base_url/secret 없으면 skip.
    """
    if not base_url or not secret:
        return
    try:
        url = f"{base_url.rstrip('/')}/api/closing/run-log"
        requests.post(
            url,
            headers={
                "Authorization": f"Bearer {secret}",
                "Content-Type": "application/json",
            },
            json={
                "status": status,
                "service_count": service_count,
                "message": message[:1000],
            },
            timeout=15,
        )
        print(f"[OK] 실행기록 보고: {status} ({service_count})")
    except Exception as exc:  # noqa: BLE001 — best-effort 보고
        print(f"[WARN] 실행기록 보고 실패(무시): {exc}")


def _decrypt_if_needed(path: str):
    """Moa 엑셀은 서버 생성 시 표준 기본 키로 암호화(CDFV2/OLE). 복호 → BytesIO. 평문이면 path.

    기본 키 'VelvetSweatshop' = Excel이 "암호 없이 암호화"에 쓰는 공개 상수(비밀 아님).
    Moa 다운로드가 이 키를 사용함이 확인됨. 운영자별 비밀번호 아님.
    (다른 키로 바뀌면 MOA_EXCEL_PASSWORD env로 override.)
    """
    with open(path, "rb") as f:
        magic = f.read(4)
    if magic != b"\xd0\xcf\x11\xe0":  # OLE 매직 아님 → 평문 xlsx(zip)
        return path
    pw = os.getenv("MOA_EXCEL_PASSWORD") or "VelvetSweatshop"
    import io
    import msoffcrypto

    buf = io.BytesIO()
    with open(path, "rb") as f:
        office = msoffcrypto.OfficeFile(f)
        office.load_key(password=pw)
        office.decrypt(buf)
    buf.seek(0)
    return buf


def parse_excel(path: str, scraped_at_dt: datetime) -> list[dict]:
    """엑셀(암호화 복호 후) → 14컬럼 매핑 → 작성마감 < scraped_at 필터 → 인제스트 row."""
    from openpyxl import load_workbook

    wb = load_workbook(_decrypt_if_needed(path), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header = [str(c).strip() if c is not None else "" for c in next(rows_iter)]
    idx = {EXCEL_COLUMN_MAP[h]: i for i, h in enumerate(header) if h in EXCEL_COLUMN_MAP}

    missing = [k for k in EXCEL_COLUMN_MAP.values() if k not in idx]
    if missing:
        raise RuntimeError(f"엑셀 헤더 불일치 — 누락 컬럼: {missing} / 실제 헤더: {header}")

    out: list[dict] = []
    for raw in rows_iter:
        if raw is None or all(c is None for c in raw):
            continue
        get = lambda key: raw[idx[key]] if idx.get(key) is not None else None
        write_end = to_kst_iso(get("write_end_at"))
        if write_end is None:
            continue
        if datetime.fromisoformat(write_end) >= scraped_at_dt:
            continue  # 아직 마감 전
        out.append(
            {
                "service_id": int(get("service_id")),
                "university_name": str(get("university_name") or "").strip(),
                "region": _opt(get("region")),
                "service_name": str(get("service_name") or "").strip(),
                "university_type": _opt(get("university_type")),
                "category": _opt(get("category")),
                "admission_type": _opt(get("admission_type")),
                "operator_name": _opt(get("operator_name")),
                "developer_name": _opt(get("developer_name")),
                "write_start_at": to_kst_iso(get("write_start_at")),
                "write_end_at": write_end,
                "pay_start_at": to_kst_iso(get("pay_start_at")),
                "pay_end_at": to_kst_iso(get("pay_end_at")),
                "solo": _to_bool(get("solo")),
            }
        )
    return out


def _opt(v) -> str | None:
    s = str(v).strip() if v is not None else ""
    return s or None


def _to_bool(v) -> bool:
    return str(v).strip() in ("Y", "y", "true", "True", "1", "단독", "O")


def to_kst_iso(v) -> str | None:
    """Moa 표시값(datetime 또는 'YYYY-MM-DD HH:MM')을 KST(+09:00) ISO8601로."""
    if v is None or str(v).strip() == "":
        return None
    if isinstance(v, datetime):
        dt = v if v.tzinfo else v.replace(tzinfo=KST)
        return dt.astimezone(KST).isoformat()
    s = str(v).strip().replace("/", "-")
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).replace(tzinfo=KST).isoformat()
        except ValueError:
            continue
    raise ValueError(f"날짜 파싱 실패: {v!r}")


# ── 브라우저 / Moa DOM (LIVE DISCOVERY) ────────────────────────────────


def setup_driver(download_dir: str, headless: bool):
    """undetected-chromedriver로 Cloudflare 봇 탐지 회피.

    Cloudflare 'Just a moment' 챌린지는 헤드리스/자동화 시그니처를 잡으므로 uc를 쓴다.
    CI에서는 xvfb 가상 디스플레이 + non-headless(HEADLESS_MODE=false)가 통과율이 높다.
    """
    opts = uc.ChromeOptions()
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--window-size=1400,1000")
    opts.add_experimental_option(
        "prefs",
        {
            "download.default_directory": download_dir,
            "download.prompt_for_download": False,
        },
    )
    # setup-chrome이 알려준 바이너리/버전을 명시 — 미지정 시 uc가 최신 드라이버를
    # 받아 설치된 Chrome과 버전 불일치(SessionNotCreated)가 난다.
    chrome_bin = os.getenv("CHROME_BIN") or None
    cv = os.getenv("CHROME_VERSION", "")
    major = cv.split(".")[0] if cv else ""
    version_main = int(major) if major.isdigit() else None
    driver = uc.Chrome(
        options=opts,
        headless=headless,
        use_subprocess=True,
        browser_executable_path=chrome_bin,
        version_main=version_main,
    )
    # uc는 prefs의 download dir를 무시할 수 있어 CDP로 한 번 더 지정.
    try:
        driver.execute_cdp_cmd(
            "Page.setDownloadBehavior",
            {"behavior": "allow", "downloadPath": download_dir},
        )
    except Exception:  # noqa: BLE001
        pass
    return driver


def _dump_page(driver, label: str) -> None:
    """실패 진단용 — 현재 URL/타이틀 + 페이지 소스 힌트를 로그에 남기고,
    스크린샷·전체 HTML을 파일로 저장(워크플로가 artifact로 업로드).

    로그인 폼 미등장이 (1)셀렉터 변경인지 (2)차단/점검/캡차 페이지인지 즉시 판별한다.
    """
    try:
        out = os.getenv("CLOSING_DUMP_DIR", ".")
        print(f"[DUMP:{label}] url={driver.current_url} title={driver.title!r}")
        src = driver.page_source or ""
        print(f"[DUMP:{label}] page_source 길이={len(src)}")
        hints = [
            "txtUserID",
            "txtUserPwd",
            "secCaptcha",
            "점검",
            "차단",
            "Access Denied",
            "blocked",
            "Just a moment",
            "Cloudflare",
        ]
        found = [kw for kw in hints if kw.lower() in src.lower()]
        print(f"[DUMP:{label}] 힌트 포함: {found if found else '없음'}")
        try:
            driver.save_screenshot(os.path.join(out, f"fail-{label}.png"))
        except Exception as exc:  # noqa: BLE001
            print(f"[DUMP:{label}] 스크린샷 실패: {exc}")
        with open(os.path.join(out, f"fail-{label}.html"), "w", encoding="utf-8") as f:
            f.write(src[:300000])
        print(f"[DUMP:{label}] 저장: fail-{label}.png / fail-{label}.html")
    except Exception as exc:  # noqa: BLE001 — 덤프 실패는 무시
        print(f"[DUMP:{label}] 덤프 실패(무시): {exc}")


def _wait_cloudflare_clear(driver, timeout_sec: int = 45, poll_sec: int = 3) -> None:
    """Cloudflare 'Just a moment' JS 챌린지가 자동 통과되길 대기.

    uc(undetected-chromedriver)가 챌린지를 자동 처리하므로, JS 실행 시간을 주면
    타이틀이 'Just a moment...'에서 실제 로그인 페이지로 바뀐다. timeout 내 안 바뀌면
    그대로 진행(이후 login_id 대기가 최종 판정).
    """
    deadline = time.monotonic() + timeout_sec
    while time.monotonic() < deadline:
        title = (driver.title or "").lower()
        if "just a moment" not in title and "moment" not in title:
            return
        print("[INFO] Cloudflare 챌린지 통과 대기 중...")
        time.sleep(poll_sec)


def _open_login_page(driver, wait, attempts: int = 3) -> None:
    """로그인 페이지 진입 + (Cloudflare 챌린지 통과 대기) + 로그인 폼 등장 대기.

    자격증명 제출 '전' 단계라 일시적 지연/챌린지에는 안전하게 재시도한다. attempts회
    모두 실패하면 실패 페이지를 덤프하고 명확한 RuntimeError를 올린다.
    (SMS 미발송 단계이므로 재시도가 2FA 흐름을 깨지 않음.)
    """
    for i in range(attempts):
        try:
            driver.get(MOA_LOGIN_URL)
            _wait_cloudflare_clear(driver)
            wait.until(
                EC.presence_of_element_located(
                    (By.CSS_SELECTOR, SELECTORS["login_id"])
                )
            )
            return
        except TimeoutException:
            print(f"[WARN] 로그인 폼 미등장 (시도 {i + 1}/{attempts}) — 재시도")
            time.sleep(3)
    _dump_page(driver, "login")
    raise RuntimeError(
        f"로그인 폼(login_id={SELECTORS['login_id']}) {attempts}회 미등장 — "
        "Cloudflare 차단/페이지 변경 의심 (artifact의 fail-login.html/png 확인)"
    )


def login_and_2fa(driver, wait, env) -> None:
    """Moa 로그인 + SMS 2FA. #btnLogin 이중용도(1차 SMS발송 → 2차 인증확인).

    SMS 신선도: 로그인 제출 '직전' baseline 코드 저장 → 제출 후 폴링하며
    baseline과 달라지면 새 SMS로 간주(plan §결정1).
    캡차(#secCaptcha)가 보이면 자동 해결 불가 → 즉시 abort(첫 시도 성공 필수).
    """
    _open_login_page(driver, wait)
    driver.find_element(By.CSS_SELECTOR, SELECTORS["login_id"]).send_keys(env["username"])
    driver.find_element(By.CSS_SELECTOR, SELECTORS["login_pw"]).send_keys(env["password"])

    baseline = fetch_sms_code(env["sms_url"])  # 제출 전 baseline
    driver.find_element(By.CSS_SELECTOR, SELECTORS["login_submit"]).click()  # 1차 → SMS 발송
    _abort_if_captcha(driver)

    code = poll_fresh_sms_code(
        env["sms_url"], baseline, env["sms_timeout"], env["sms_interval"]
    )
    wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, SELECTORS["sms_code_input"])))
    driver.find_element(By.CSS_SELECTOR, SELECTORS["sms_code_input"]).send_keys(code)
    driver.find_element(By.CSS_SELECTOR, SELECTORS["sms_submit"]).click()  # 2차 → 인증확인
    time.sleep(2)
    _abort_if_captcha(driver)
    print("[OK] 로그인 + 2FA 완료")


def _abort_if_captcha(driver) -> None:
    els = driver.find_elements(By.CSS_SELECTOR, SELECTORS["captcha_section"])
    if els and els[0].is_displayed():
        raise RuntimeError(
            "캡차 노출 감지 — 직전 로그인 실패 추정. 자동 해결 불가, abort. "
            "(자격증명/셀렉터 확인 후 재시도)"
        )


def search_and_download(driver, wait, download_dir: str, ay: dict) -> str:
    """ServiceSearch 오픈일 범위 → 엑셀 다운로드(JS 호출) → 파일 경로.

    오픈일 포맷 'YYYY-MM-DD HH:MM'. 운영자/개발자/카테고리 드롭다운은 ''(전체)로 비운다
    — 로그인 운영자로 기본 선택돼 결과가 본인 담당으로 제한되는 것을 방지.
    엑셀 저장은 GetUnivServiceListToExcel()(searchForm POST)을 직접 호출.
    """
    driver.get(SERVICE_SEARCH_URL)
    wait.until(EC.presence_of_element_located((By.ID, SELECTORS["open_from"])))

    _set_value(driver, SELECTORS["open_from"], f'{ay["start"]["date"]} {ay["start"]["time"]}')
    _set_value(driver, SELECTORS["open_to"], f'{ay["end"]["date"]} {ay["end"]["time"]}')
    for ddl in (SELECTORS["ddl_manager"], SELECTORS["ddl_developer"], SELECTORS["ddl_category"]):
        _set_value(driver, ddl, "")  # '선택'(전체)

    before = set(glob.glob(os.path.join(download_dir, "*.xls*")))
    driver.execute_script(EXCEL_DOWNLOAD_JS)
    return _wait_download(download_dir, before, timeout=60)


def _set_value(driver, element_id: str, value: str) -> None:
    """flatpickr/select에 .value 직접 주입 + change 디스패치 (calendar UI 우회)."""
    driver.execute_script(
        "var el=document.getElementById(arguments[0]);"
        "if(el){el.value=arguments[1];"
        "el.dispatchEvent(new Event('change',{bubbles:true}));}",
        element_id,
        value,
    )


def _wait_download(download_dir: str, before: set, timeout: int) -> str:
    deadline = time.monotonic() + timeout
    while time.monotonic() < deadline:
        now = set(glob.glob(os.path.join(download_dir, "*.xls*")))
        new = [f for f in (now - before) if not f.endswith(".crdownload")]
        if new:
            print(f"[OK] 엑셀 다운로드: {os.path.basename(new[0])}")
            return new[0]
        time.sleep(1)
    raise RuntimeError("엑셀 다운로드 타임아웃")


# ── main ───────────────────────────────────────────────────────────────


def main() -> int:
    dry_run = os.getenv("CLOSING_DRY_RUN", "").lower() == "true"
    base_url = os.getenv("OPS_CONSOLE_BASE_URL", "")
    secret = os.getenv("CRON_SECRET", "")

    anchor = os.getenv("CLOSING_BIWEEKLY_ANCHOR", "2026-06-08")
    now = datetime.now(timezone.utc)
    if not should_run_this_week(now, anchor):
        print(f"[SKIP] 격주 비실행 주 (anchor={anchor}). 종료.")
        if not dry_run:
            post_run_log(base_url, secret, "skipped", 0, f"격주 off주 (anchor={anchor})")
        return 0

    env = {
        "username": os.getenv("MOA_USERNAME", ""),
        "password": os.getenv("MOA_PASSWORD", ""),
        "sms_url": os.getenv("MAKE_SMS_CODE_URL", ""),
        "sms_timeout": int(os.getenv("MOA_SMS_POLL_TIMEOUT_SEC", "90")),
        "sms_interval": int(os.getenv("MOA_SMS_POLL_INTERVAL_SEC", "3")),
        "wait_sec": int(os.getenv("MOA_WAIT_SEC", "40")),
        "base_url": base_url,
        "secret": secret,
    }
    missing = [k for k in ("username", "password", "sms_url") if not env[k]]
    if not dry_run:
        missing += [k for k in ("base_url", "secret") if not env[k]]
    if missing:
        print(f"[FAIL] 환경변수 누락: {missing}")
        if not dry_run:
            post_run_log(base_url, secret, "failed", 0, f"환경변수 누락: {missing}")
        return 1

    headless = os.getenv("HEADLESS_MODE", "true").lower() == "true"
    download_dir = tempfile.mkdtemp(prefix="moa-closing-")
    scraped_at_dt = datetime.now(KST)
    ay = academic_year_range(scraped_at_dt)
    print(f"[INFO] 학년도 검색 범위: {ay['start']['date']} ~ {ay['end']['date']}")

    try:
        driver = setup_driver(download_dir, headless)
        try:
            login_and_2fa(driver, WebDriverWait(driver, env["wait_sec"]), env)
            path = search_and_download(
                driver, WebDriverWait(driver, env["wait_sec"]), download_dir, ay
            )
        finally:
            driver.quit()

        rows = parse_excel(path, scraped_at_dt)
        print(
            f"[INFO] 마감 서비스 {len(rows)}건 추출 (scraped_at={scraped_at_dt.isoformat()})"
        )

        if dry_run:
            print("[DRY-RUN] 인제스트 미전송. 추출만 완료.")
            return 0
        if not rows:
            print("[INFO] 마감 0건 — 인제스트 미전송(빈 배열 거부 정책).")
            post_run_log(base_url, secret, "success", 0, "마감 0건 — 적재 없음")
            return 0

        post_ingest(env["base_url"], env["secret"], scraped_at_dt.isoformat(), rows)
        post_run_log(base_url, secret, "success", len(rows), f"적재 {len(rows)}건")
        return 0
    except Exception as exc:  # noqa: BLE001 — 실패도 실행기록에 보고 후 재전파
        if not dry_run:
            first = str(exc).strip().splitlines()
            detail = first[0] if first else ""
            msg = f"{type(exc).__name__}: {detail}".strip().rstrip(":").strip()
            post_run_log(base_url, secret, "failed", 0, msg[:300])
        raise


if __name__ == "__main__":
    sys.exit(main())
