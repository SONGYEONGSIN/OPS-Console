#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
SharePoint 폴더 탐색 및 최근 수정 파일 찾기
- 폴더 경로 조회
- 파일 목록 확인
- 최근 수정 파일 찾기
"""

import os
import sys
import requests
import re
import json
from typing import Dict, List, Optional, Any, Tuple
from datetime import datetime, timedelta
from urllib.parse import quote

# 현재 디렉토리를 Python 경로에 추가
current_dir = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, current_dir)

try:
    from dotenv import load_dotenv
    env_file = os.path.join(current_dir, "weekly_report_config.env")
    load_dotenv(env_file)
    print(f"[INFO] 환경 설정 파일 로드: {env_file}")
    
    # Teams 채팅 ID 확인
    teams_chat_id = os.getenv('TEAMS_CHAT_ID')
    if teams_chat_id:
        print(f"[OK] Teams 채팅 ID 설정됨: {teams_chat_id[:30]}...")
    else:
        print("[WARN] TEAMS_CHAT_ID가 설정되지 않았습니다.")
        
except ImportError:
    print("[FAIL] python-dotenv 패키지가 설치되지 않았습니다.")
    sys.exit(1)

try:
    import msal
except ImportError:
    print("[FAIL] msal 패키지가 설치되지 않았습니다.")
    sys.exit(1)

try:
    import openpyxl
    from openpyxl import styles as openpyxl_styles
    from openpyxl.utils import get_column_letter
except ImportError:
    print("[WARN] openpyxl 패키지가 설치되지 않았습니다. Excel 시트 조회 기능을 사용할 수 없습니다.")
    openpyxl = None
    openpyxl_styles = None  # type: ignore
    get_column_letter = None  # type: ignore

class SharePointFolderExplorer:
    """SharePoint 폴더 탐색 클래스"""
   
    def __init__(self, debug_mode=True):
        """초기화"""
        self.debug_mode = debug_mode
        self.access_token = None  # SharePoint용 토큰
        self.teams_access_token = None  # Teams용 토큰
        self.config = self.load_config()
        self.setup_authentication()
   
    def debug_print(self, message: str):
        """DEBUG 모드일 때만 출력"""
        if self.debug_mode:
            print(f"[DEBUG] {message}")
   
    def load_config(self) -> Dict:
        """설정 파일 로드"""
        config = {
            'tenant_id': os.getenv('SHAREPOINT_TENANT_ID'),
            'client_id': os.getenv('SHAREPOINT_CLIENT_ID'),
            'client_secret': os.getenv('SHAREPOINT_CLIENT_SECRET'),
            'site_url': os.getenv('SHAREPOINT_SITE_URL'),
            'site_id': None,  # 동적으로 가져올 예정
            'drive_id': None,  # 동적으로 가져올 예정
        }
       
        # 필수 설정 확인
        required_keys = ['tenant_id', 'client_id', 'client_secret', 'site_url']
        missing_keys = [key for key in required_keys if not config[key]]
       
        if missing_keys:
            print(f"[ERROR] 필수 설정이 누락되었습니다: {', '.join(missing_keys)}")
            raise ValueError(f"필수 설정이 누락되었습니다: {', '.join(missing_keys)}")
           
        return config
   
    def setup_authentication(self):
        """Microsoft Graph API 인증 설정"""
        try:
            # SharePoint는 항상 Client Credentials 방식 사용
            print("[INFO] SharePoint용 Client Credentials 인증 시도...")
            self.setup_client_credentials_authentication()
            
            # Teams 메시지 전송을 위한 Delegated 인증 (별도)
            teams_auth_type = os.getenv('TEAMS_AUTH_TYPE', 'client_credentials')
            if teams_auth_type == 'delegated':
                print("[INFO] Teams용 Delegated Permission 인증 시도...")
                self.setup_teams_delegated_authentication()
            else:
                print("[INFO] Teams도 Client Credentials 사용 (제한적)")
               
        except Exception as e:
            print(f"[FAIL] 인증 설정 실패: {str(e)}")
            raise
    
    def setup_client_credentials_authentication(self):
        """Client Credentials 방식 인증"""
        try:
            app = msal.ConfidentialClientApplication(
                self.config['client_id'],
                authority=f"https://login.microsoftonline.com/{self.config['tenant_id']}",
                client_credential=self.config['client_secret']
            )
           
            result = app.acquire_token_for_client(scopes=["https://graph.microsoft.com/.default"])
           
            if result and "access_token" in result:
                self.access_token = result["access_token"]
                print("[OK] Microsoft Graph API 인증 성공 (Client Credentials)")
                
                # Site ID와 Drive ID 동적으로 가져오기
                self.get_site_id()
                self.get_drive_id()
            else:
                raise Exception(f"인증 실패: {result.get('error_description', 'Unknown error')}")
               
        except Exception as e:
            print(f"[FAIL] Client Credentials 인증 실패: {str(e)}")
            raise
    
    def setup_teams_delegated_authentication(self):
        """Teams 전용 Delegated Permission 방식 인증 (Device Code Flow)"""
        try:
            print("[INFO] Teams용 Device Code Flow 인증 시작...")
            
            # 토큰 캐시 파일 경로
            cache_file = os.path.join(current_dir, "teams_token_cache.bin")
            
            # 토큰 캐시 로드
            cache = msal.SerializableTokenCache()
            if os.path.exists(cache_file):
                with open(cache_file, 'r') as f:
                    cache.deserialize(f.read())
                print("[INFO] 저장된 토큰 캐시를 찾았습니다.")
            
            app = msal.PublicClientApplication(
                self.config['client_id'],
                authority=f"https://login.microsoftonline.com/{self.config['tenant_id']}",
                token_cache=cache
            )
            
            # Teams 메시지 전송을 위한 권한 범위
            scopes = [
                "https://graph.microsoft.com/Chat.ReadWrite",
                "https://graph.microsoft.com/ChatMessage.Send"
            ]
            
            # 먼저 캐시된 토큰으로 시도
            accounts = app.get_accounts()
            if accounts:
                print("[INFO] 저장된 계정 정보를 찾았습니다. 자동 로그인 시도 중...")
                result = app.acquire_token_silent(scopes, account=accounts[0])
                
                if result and "access_token" in result:
                    self.teams_access_token = result["access_token"]
                    print("[OK] Teams용 자동 로그인 성공! (저장된 토큰 사용)")
                    return
                else:
                    print("[INFO] 저장된 토큰이 만료되었습니다. 재로그인이 필요합니다.")
            
            # 캐시된 토큰이 없거나 만료된 경우 Device Code Flow 시작
            flow = app.initiate_device_flow(scopes=scopes)
            
            if "user_code" not in flow:
                raise ValueError(f"Device Code Flow 시작 실패: {flow.get('error')}")
            
            print(f"\n{'='*60}")
            print("🔐 Teams 메시지 전송을 위한 로그인이 필요합니다")
            print('='*60)
            print(f"[STEP 1] 브라우저에서 다음 URL을 여세요:")
            print(f"         {flow['verification_uri']}")
            print(f"")
            print(f"[STEP 2] 다음 코드를 입력하세요:")
            print(f"         {flow['user_code']}")
            print(f"")
            print(f"[STEP 3] {os.getenv('TEAMS_USERNAME', 'ys1114@jinhakapply.com')} 계정으로 로그인")
            print(f"[STEP 4] Teams 권한 요청을 승인")
            print(f"[STEP 5] 완료 후 이 창으로 돌아오세요")
            print('='*60)
            print("\n[INFO] 사용자 인증을 기다리는 중...")
            
            # 사용자 인증 대기
            result = app.acquire_token_by_device_flow(flow)
            
            if result and "access_token" in result:
                self.teams_access_token = result["access_token"]
                print("[OK] Teams용 인증 성공 (Delegated Permission)")
                
                # 토큰 캐시 저장
                if cache.has_state_changed:
                    with open(cache_file, 'w') as f:
                        f.write(cache.serialize())
                    print(f"[OK] 토큰이 저장되었습니다: {cache_file}")
                    print("[INFO] 다음부터는 로그인 없이 자동으로 실행됩니다!")
            else:
                error_msg = result.get('error_description', 'Unknown error') if result else 'No result'
                print(f"[FAIL] Teams 인증 실패: {error_msg}")
                print("[WARN] Teams 메시지 전송 기능을 사용할 수 없습니다.")
               
        except Exception as e:
            print(f"[FAIL] Teams Delegated 인증 실패: {str(e)}")
            print("[WARN] Teams 메시지 전송 기능을 사용할 수 없습니다.")
    
    def get_site_id(self):
        """SharePoint 사이트 ID 획득"""
        try:
            print("[INFO] SharePoint 사이트 ID 획득 중...")
            
            # 사이트 URL에서 도메인과 경로 추출
            site_url = self.config['site_url']
            domain = site_url.split('//')[1].split('.sharepoint.com')[0]
            site_path = site_url.replace('https://', '').split('.sharepoint.com')[1]
            
            url = f"https://graph.microsoft.com/v1.0/sites/{domain}.sharepoint.com:{site_path}"
            self.debug_print(f"사이트 ID 조회 URL: {url}")
            
            response = requests.get(url, headers=self.get_headers())
            response.raise_for_status()
            
            site_data = response.json()
            self.config['site_id'] = site_data['id']
            print(f"[OK] 사이트 ID 획득: {self.config['site_id']}")
            
        except Exception as e:
            print(f"[FAIL] 사이트 ID 획득 실패: {str(e)}")
            raise
    
    def get_drive_id(self):
        """SharePoint 드라이브 ID 획득"""
        try:
            print("[INFO] SharePoint 드라이브 ID 획득 중...")
            
            url = f"https://graph.microsoft.com/v1.0/sites/{self.config['site_id']}/drives"
            self.debug_print(f"드라이브 ID 조회 URL: {url}")
            
            response = requests.get(url, headers=self.get_headers())
            response.raise_for_status()
            
            drives_data = response.json()
            drives = drives_data.get('value', [])
            
            if drives:
                # 기본 드라이브 (문서 라이브러리) 사용
                self.config['drive_id'] = drives[0]['id']
                print(f"[OK] 드라이브 ID 획득: {self.config['drive_id']}")
                print(f"[INFO] 드라이브 이름: {drives[0].get('name', 'Unknown')}")
            else:
                raise Exception("사용 가능한 드라이브가 없습니다")
            
        except Exception as e:
            print(f"[FAIL] 드라이브 ID 획득 실패: {str(e)}")
            raise
   
    def get_headers(self) -> Dict[str, str]:
        """API 요청 헤더 생성"""
        return {
            'Authorization': f'Bearer {self.access_token}',
            'Content-Type': 'application/json'
        }
   
    def list_root_folders(self) -> List[Dict]:
        """루트 폴더 목록 조회"""
        try:
            print("\n[STEP 1] 루트 폴더 조회")
            print("-" * 40)
           
            url = f"https://graph.microsoft.com/v1.0/sites/{self.config['site_id']}/drives/{self.config['drive_id']}/root/children"
            self.debug_print(f"루트 폴더 조회 URL: {url}")
           
            response = requests.get(url, headers=self.get_headers())
            response.raise_for_status()
           
            data = response.json()
            items = data.get('value', [])
           
            folders = []
            files = []
           
            for item in items:
                if item.get('folder'):
                    folders.append(item)
                else:
                    files.append(item)
           
            print(f"[INFO] 루트에서 발견된 항목: {len(items)}개")
            print(f"[INFO] 폴더: {len(folders)}개, 파일: {len(files)}개")
           
            print("\n[FOLDERS] 폴더 목록:")
            for i, folder in enumerate(folders, 1):
                folder_name = folder.get('name', 'Unknown')
                print(f"  {i}. 📁 {folder_name}")
           
            print("\n[FILES] 파일 목록 (처음 5개):")
            for i, file in enumerate(files[:5], 1):
                file_name = file.get('name', 'Unknown')
                file_size = file.get('size', 0)
                print(f"  {i}. 📄 {file_name} ({file_size} bytes)")
           
            return folders
           
        except Exception as e:
            print(f"[FAIL] 루트 폴더 조회 실패: {str(e)}")
            return []
   
    def explore_folder(self, folder_path: str) -> List[Dict]:
        """특정 폴더 내용 조회"""
        try:
            print(f"\n[STEP 2] 폴더 탐색: {folder_path}")
            print("-" * 40)
           
            # URL 인코딩 적용
            encoded_path = quote(folder_path, safe='')
            url = f"https://graph.microsoft.com/v1.0/sites/{self.config['site_id']}/drives/{self.config['drive_id']}/root:/{encoded_path}:/children"
           
            self.debug_print(f"폴더 조회 URL: {url}")
           
            response = requests.get(url, headers=self.get_headers())
           
            if response.status_code == 200:
                data = response.json()
                items = data.get('value', [])
               
                folders = []
                files = []
               
                for item in items:
                    if item.get('folder'):
                        folders.append(item)
                    else:
                        files.append(item)
               
                print(f"[OK] 폴더 발견: {folder_path}")
                print(f"[INFO] 항목 수: {len(items)}개 (폴더: {len(folders)}개, 파일: {len(files)}개)")
               
                # 폴더 목록
                if folders:
                    print("\n[FOLDERS] 하위 폴더:")
                    for i, folder in enumerate(folders, 1):
                        folder_name = folder.get('name', 'Unknown')
                        print(f"  {i}. 📁 {folder_name}")
               
                # 파일 목록
                if files:
                    print("\n[FILES] 파일 목록:")
                    for i, file in enumerate(files, 1):
                        file_name = file.get('name', 'Unknown')
                        file_size = file.get('size', 0)
                        modified = file.get('lastModifiedDateTime', 'Unknown')
                        print(f"  {i}. 📄 {file_name}")
                        print(f"      크기: {file_size} bytes, 수정일: {modified}")
               
                return files
               
            elif response.status_code == 404:
                print(f"[FAIL] 폴더를 찾을 수 없습니다: {folder_path}")
                return []
            else:
                print(f"[ERROR] 폴더 조회 실패: {response.status_code}")
                print(f"[ERROR] 응답: {response.text}")
                return []
               
        except Exception as e:
            print(f"[FAIL] 폴더 탐색 실패: {str(e)}")
            return []
   
    def find_latest_file_in_folder(self, folder_path: str, file_pattern: str = None) -> Optional[Dict]:
        """폴더에서 최근 수정된 파일 찾기"""
        try:
            print(f"\n[STEP 3] 최근 수정 파일 검색")
            print("-" * 40)
           
            files = self.explore_folder(folder_path)
           
            if not files:
                print("[FAIL] 폴더에 파일이 없습니다.")
                return None
           
            # 파일 필터링 (패턴이 있으면 적용)
            if file_pattern:
                filtered_files = [f for f in files if file_pattern in f.get('name', '')]
                print(f"[INFO] '{file_pattern}' 패턴으로 필터링: {len(filtered_files)}개 파일")
                files = filtered_files
           
            if not files:
                print(f"[FAIL] '{file_pattern}' 패턴에 맞는 파일이 없습니다.")
                return None
           
            # 최근 수정일 기준으로 정렬
            latest_file = max(files, key=lambda x: x.get('lastModifiedDateTime', ''))
           
            file_name = latest_file.get('name', 'Unknown')
            file_size = latest_file.get('size', 0)
            modified_time = latest_file.get('lastModifiedDateTime', 'Unknown')
            file_id = latest_file.get('id', 'Unknown')
           
            print(f"[OK] 최근 수정된 파일 발견:")
            print(f"  📄 파일명: {file_name}")
            print(f"  📏 크기: {file_size} bytes")
            print(f"  📅 수정일: {modified_time}")
            print(f"  🆔 파일 ID: {file_id}")
           
            return latest_file
           
        except Exception as e:
            print(f"[FAIL] 최근 파일 검색 실패: {str(e)}")
            return None
   
    def search_files_by_name(self, search_term: str) -> List[Dict]:
        """파일명으로 검색"""
        try:
            print(f"\n[STEP 4] 파일명 검색: '{search_term}'")
            print("-" * 40)
           
            search_url = f"https://graph.microsoft.com/v1.0/sites/{self.config['site_id']}/drives/{self.config['drive_id']}/root/search(q='{search_term}')"
            self.debug_print(f"검색 URL: {search_url}")
           
            response = requests.get(search_url, headers=self.get_headers())
            response.raise_for_status()
           
            data = response.json()
            files = data.get('value', [])
           
            print(f"[INFO] 검색 결과: {len(files)}개 파일 발견")
           
            # 파일 목록 출력
            for i, file in enumerate(files[:10], 1):  # 최대 10개만
                file_name = file.get('name', 'Unknown')
                file_path = file.get('parentReference', {}).get('path', 'Unknown')
                modified_time = file.get('lastModifiedDateTime', 'Unknown')
                print(f"  {i}. 📄 {file_name}")
                print(f"      경로: {file_path}")
                print(f"      수정일: {modified_time}")
           
            return files
           
        except Exception as e:
            print(f"[FAIL] 파일 검색 실패: {str(e)}")
            return []
   
    def download_file(self, file_id: str) -> Optional[bytes]:
        """파일 다운로드"""
        try:
            url = f"https://graph.microsoft.com/v1.0/sites/{self.config['site_id']}/drives/{self.config['drive_id']}/items/{file_id}/content"
           
            self.debug_print(f"파일 다운로드 URL: {url}")
           
            response = requests.get(url, headers=self.get_headers())
            response.raise_for_status()
           
            # 응답 내용 확인
            content_type = response.headers.get('content-type', '')
            content_length = len(response.content)
           
            self.debug_print(f"다운로드된 파일 정보:")
            self.debug_print(f"  Content-Type: {content_type}")
            self.debug_print(f"  Content-Length: {content_length} bytes")
           
            # Excel 파일 시그니처 확인
            if response.content.startswith(b'PK'):
                self.debug_print("Excel 파일 형식 확인됨 (ZIP 기반)")
            else:
                self.debug_print("Excel 파일 형식이 아닐 수 있습니다")
           
            return response.content
           
        except Exception as e:
            print(f"[FAIL] 파일 다운로드 실패: {str(e)}")
            return None
   
    def get_excel_sheets(self, file_content: bytes) -> List[str]:
        """Excel 파일의 시트 목록 조회"""
        try:
            if openpyxl is None:
                print("[FAIL] openpyxl 패키지가 설치되지 않았습니다.")
                return []
           
            print(f"[INFO] Excel 파일 처리 시작: {len(file_content)} bytes")
           
            # 파일 내용을 BytesIO로 변환
            import io
            file_stream = io.BytesIO(file_content)
           
            # Excel 파일 열기
            self.debug_print("Excel 파일 로드 중...")
            workbook = openpyxl.load_workbook(file_stream)
           
            sheet_names = workbook.sheetnames
            print(f"[OK] Excel 파일 로드 완료. 시트 수: {len(sheet_names)}")
           
            return sheet_names
           
        except Exception as e:
            print(f"[FAIL] Excel 시트 조회 실패: {str(e)}")
            return []
   
    def analyze_latest_file_sheets(self, file_info: Dict) -> Optional[Dict]:
        """최근 파일의 시트 정보 분석"""
        try:
            print(f"\n[STEP 3] Excel 시트 분석")
            print("-" * 40)
           
            file_name = file_info.get('name', 'Unknown')
            file_id = file_info.get('id', '')
           
            print(f"[INFO] 분석 대상 파일: {file_name}")
            print(f"[INFO] 파일 ID: {file_id}")
           
            # 파일 다운로드
            print("[INFO] 파일 다운로드 중...")
            file_content = self.download_file(file_id)
            if not file_content:
                print("[FAIL] 파일 다운로드 실패")
                return None
           
            # 시트 목록 조회
            print("[INFO] 시트 목록 조회 중...")
            sheet_names = self.get_excel_sheets(file_content)
            if not sheet_names:
                print("[FAIL] 시트 목록 조회 실패")
                return None
           
            # 주간업무보고서 관련 시트 찾기
            # 규칙: 첫 번째 시트가 항상 최신 시트
            report_sheets = []
            for sheet_name in sheet_names:
                # "YYYY년 MM월 N주차" 패턴을 가진 시트 찾기
                if re.search(r'\d{4}년\s*\d+월\s*\d+주차', sheet_name):
                    report_sheets.append(sheet_name)
           
            # 가장 최근 시트는 첫 번째 시트
            latest_sheet = None
            if report_sheets:
                print(f"[INFO] 발견된 주간보고 시트: {len(report_sheets)}개")
                
                # 첫 번째 시트가 최신 시트
                latest_sheet = report_sheets[0]
                print(f"[OK] 최근 시트 선택 (첫 번째 시트): {latest_sheet}")
            else:
                print("[WARN] 주간업무보고서 관련 시트를 찾을 수 없습니다.")
                # 패턴 매칭 실패시에도 첫 번째 시트 사용
                latest_sheet = sheet_names[0] if sheet_names else None
                if latest_sheet:
                    print(f"[INFO] 첫 번째 시트를 사용: {latest_sheet}")
           
            # 결과 정리
            result = {
                'file_info': file_info,
                'sheet_names': sheet_names,
                'report_sheets': report_sheets,
                'latest_sheet': latest_sheet,
                'total_sheets': len(sheet_names)
            }
           
            return result
           
        except Exception as e:
            print(f"[FAIL] 시트 분석 실패: {str(e)}")
            return None
   
    def get_iso_month_weeks_count(self, year: int, month: int) -> int:
        """ISO 8601 기준으로 해당 월의 주차 수 동적 계산
        규칙: 목요일이 속한 월이 해당 주의 월을 결정, 주는 월요일 시작
        """
        # 해당 월의 첫 번째 목요일 찾기
        first_day = datetime(year, month, 1)
        days_to_thursday = (3 - first_day.weekday()) % 7
        first_thursday = first_day + timedelta(days=days_to_thursday)
        
        # 해당 월의 마지막 목요일 찾기
        if month == 12:
            next_month_first = datetime(year + 1, 1, 1)
        else:
            next_month_first = datetime(year, month + 1, 1)
        last_day = next_month_first - timedelta(days=1)
        days_from_thursday = (last_day.weekday() - 3) % 7
        last_thursday = last_day - timedelta(days=days_from_thursday)
        
        # 주차 수 = (마지막 목요일 - 첫째 목요일) / 7 + 1
        total_weeks = ((last_thursday - first_thursday).days // 7) + 1
        
        self.debug_print(f"ISO 주차 계산: {year}년 {month}월 = {total_weeks}주차")
        return total_weeks
    
    def calculate_next_week_filename(self, current_filename: str) -> str:
        """현재 파일명에서 다음 주차 파일명 계산"""
        try:
            # 패턴: "주간업무보고서_진학어플라이본부_2025_9월3주차.xlsx"
            pattern = r'(주간업무보고서_진학어플라이본부)_(\d{4})_(\d+)월(\d+)주차(\.xlsx)'
            match = re.search(pattern, current_filename)
           
            if match:
                prefix = match.group(1)  # "주간업무보고서_진학어플라이본부"
                year = int(match.group(2))  # 2025
                month = int(match.group(3))  # 9
                week = int(match.group(4))   # 3
                extension = match.group(5)  # ".xlsx"
               
                # 다음 주차 계산
                next_week = week + 1
               
                # ISO 8601 기준으로 해당 월의 최대 주차 수 동적 계산
                max_weeks = self.get_iso_month_weeks_count(year, month)
                if next_week > max_weeks:
                    next_week = 1
                    month += 1
                    if month > 12:
                        month = 1
                        year += 1
               
                # 새로운 파일명 생성
                next_filename = f"{prefix}_{year}_{month}월{next_week}주차{extension}"
                print(f"[INFO] 파일명 변환: {current_filename} → {next_filename}")
                return next_filename
           
            print(f"[WARN] 파일명 패턴을 찾을 수 없습니다: {current_filename}")
            return current_filename
           
        except Exception as e:
            print(f"[FAIL] 파일명 계산 실패: {str(e)}")
            return current_filename
   
    def calculate_next_week_sheetname(self, current_sheetname: str) -> str:
        """현재 시트명에서 다음 주차 시트명 계산"""
        try:
            # 패턴: "2025년 9월 3주차"
            pattern = r'(\d{4})년\s*(\d+)월\s*(\d+)주차'
            match = re.search(pattern, current_sheetname)
           
            if match:
                year = int(match.group(1))  # 2025
                month = int(match.group(2))  # 9
                week = int(match.group(3))   # 3
               
                # 다음 주차 계산
                next_week = week + 1
               
                # ISO 8601 기준으로 해당 월의 최대 주차 수 동적 계산
                max_weeks = self.get_iso_month_weeks_count(year, month)
                if next_week > max_weeks:
                    next_week = 1
                    month += 1
                    if month > 12:
                        month = 1
                        year += 1
               
                # 새로운 시트명 생성
                next_sheetname = f"{year}년 {month}월 {next_week}주차"
                print(f"[INFO] 시트명 변환: {current_sheetname} → {next_sheetname}")
                return next_sheetname
           
            print(f"[WARN] 시트명 패턴을 찾을 수 없습니다: {current_sheetname}")
            return current_sheetname
           
        except Exception as e:
            print(f"[FAIL] 시트명 계산 실패: {str(e)}")
            return current_sheetname
    
    def get_week_dates_from_sheetname(self, sheetname: str) -> Tuple[Optional[datetime], Optional[datetime]]:
        """시트명에서 해당 주차의 평일 날짜 범위 계산 (월~금)"""
        try:
            # 패턴: "2026년 1월 1주차"
            pattern = r'(\d{4})년\s*(\d+)월\s*(\d+)주차'
            match = re.search(pattern, sheetname)
            
            if not match:
                return None, None
            
            year = int(match.group(1))
            month = int(match.group(2))
            week = int(match.group(3))
            
            # ISO 8601: 해당 월의 첫 번째 목요일 찾기
            first_day = datetime(year, month, 1)
            days_to_thursday = (3 - first_day.weekday()) % 7
            first_thursday = first_day + timedelta(days=days_to_thursday)
            
            # N주차의 목요일 계산
            target_thursday = first_thursday + timedelta(weeks=week - 1)
            
            # 해당 주의 월요일과 금요일
            week_monday = target_thursday - timedelta(days=3)  # 목요일 - 3일 = 월요일
            week_friday = target_thursday + timedelta(days=1)  # 목요일 + 1일 = 금요일
            
            return week_monday, week_friday
            
        except Exception as e:
            print(f"[WARN] 주차 날짜 계산 실패: {str(e)}")
            return None, None
    
    def format_date_range(self, start_date: datetime, end_date: datetime) -> str:
        """날짜 범위를 "M/D~M/D" 형식으로 포맷"""
        try:
            start_str = f"{start_date.month}/{start_date.day}"
            end_str = f"{end_date.month}/{end_date.day}"
            return f"{start_str}~{end_str}"
        except Exception as e:
            print(f"[WARN] 날짜 포맷 실패: {str(e)}")
            return ""
    
    def update_sheet_cell_contents(self, worksheet: Any, new_sheetname: str, prev_h3_value: str = None) -> bool:
        """시트 내 셀 내용 자동 업데이트 (B2, B3, H3)"""
        try:
            print(f"[INFO] 셀 내용 자동 업데이트 중...")
            
            # B2 셀: 주차 텍스트 업데이트
            b2_cell = worksheet['B2']
            if b2_cell.value:
                old_b2_value = str(b2_cell.value)
                print(f"[INFO] B2 원본 값: {old_b2_value}")
                
                # "YYYY년 M월 N주차" 패턴 찾아서 새 시트명으로 교체
                pattern = r'\d{4}년\s*\d+월\s*\d+주차'
                if re.search(pattern, old_b2_value):
                    new_b2_value = re.sub(pattern, new_sheetname, old_b2_value)
                    b2_cell.value = new_b2_value
                    print(f"[OK] B2 업데이트: {old_b2_value} → {new_b2_value}")
                else:
                    print(f"[WARN] B2에서 주차 패턴을 찾을 수 없습니다")
            
            # 새 시트의 현재 주차 날짜 계산
            this_week_start, this_week_end = self.get_week_dates_from_sheetname(new_sheetname)
            
            if this_week_start and this_week_end:
                # B3 셀: 주간 업무 (이전 주차) - 원본 시트의 H3 값 사용
                b3_cell = worksheet['B3']
                if b3_cell.value and prev_h3_value:
                    old_b3_value = str(b3_cell.value)
                    print(f"[INFO] B3 원본 값: {old_b3_value}")
                    print(f"[INFO] 이전 시트의 H3 값: {prev_h3_value}")
                    
                    # "M/D~M/D" 패턴 찾아서 원본의 H3 날짜로 교체
                    date_pattern = r'\d+/\d+~\d+/\d+'
                    prev_h3_date_match = re.search(date_pattern, prev_h3_value)
                    
                    if re.search(date_pattern, old_b3_value) and prev_h3_date_match:
                        prev_date_range = prev_h3_date_match.group(0)
                        new_b3_value = re.sub(date_pattern, prev_date_range, old_b3_value)
                        b3_cell.value = new_b3_value
                        print(f"[OK] B3 업데이트: {old_b3_value} → {new_b3_value}")
                    else:
                        print(f"[WARN] B3 또는 이전 H3에서 날짜 패턴을 찾을 수 없습니다")
                
                # H3 셀: 차주 업무 (다음 주차) - 현재 주차 + 7일
                h3_cell = worksheet['H3']
                if h3_cell.value:
                    old_h3_value = str(h3_cell.value)
                    print(f"[INFO] H3 원본 값: {old_h3_value}")
                    
                    # "M/D~M/D" 패턴 찾아서 다음 주차 날짜로 교체
                    date_pattern = r'\d+/\d+~\d+/\d+'
                    if re.search(date_pattern, old_h3_value):
                        # 차주는 현재 주 + 7일
                        next_week_start = this_week_start + timedelta(days=7)
                        next_week_end = this_week_end + timedelta(days=7)
                        new_date_range = self.format_date_range(next_week_start, next_week_end)
                        new_h3_value = re.sub(date_pattern, new_date_range, old_h3_value)
                        h3_cell.value = new_h3_value
                        print(f"[OK] H3 업데이트: {old_h3_value} → {new_h3_value}")
                    else:
                        print(f"[WARN] H3에서 날짜 패턴을 찾을 수 없습니다")
            else:
                print(f"[WARN] 주차 날짜 계산 실패, B3/H3 업데이트 건너뜀")
            
            return True
            
        except Exception as e:
            print(f"[FAIL] 셀 내용 업데이트 실패: {str(e)}")
            import traceback
            print(f"[DEBUG] 상세 오류: {traceback.format_exc()}")
            return False
   
    def copy_file_with_next_week(self, source_file: Dict) -> Optional[Dict]:
        """파일을 다음 주차로 복사"""
        try:
            print(f"\n[STEP 4] 파일 복사 (다음 주차)")
            print("-" * 40)
           
            source_name = source_file.get('name', '')
            print(f"[INFO] 원본 파일: {source_name}")
           
            # 다음 주차 파일명 생성 (올바른 형식으로)
            next_week_name = self.calculate_next_week_filename(source_name)
            print(f"[INFO] 복사할 파일명: {next_week_name}")
           
            # 파일 다운로드
            print("[INFO] 원본 파일 다운로드 중...")
            file_content = self.download_file(source_file.get('id', ''))
            if not file_content:
                print("[FAIL] 파일 다운로드 실패")
                return None
           
            # 파일 크기 확인
            print(f"[INFO] 다운로드된 파일 크기: {len(file_content)} bytes")
           
            # Excel 파일 시그니처 확인
            if file_content.startswith(b'PK'):
                print("[OK] Excel 파일 형식 확인됨")
            else:
                print("[WARN] Excel 파일 형식이 아닐 수 있습니다")
           
            # 새 파일 업로드
            print("[INFO] 새 파일 업로드 중...")
            uploaded_file = self.upload_new_file(file_content, next_week_name, source_file)
            if not uploaded_file:
                print("[FAIL] 파일 업로드 실패")
                return None
           
            print(f"[OK] 파일 복사 완료: {next_week_name}")
            return uploaded_file
           
        except Exception as e:
            print(f"[FAIL] 파일 복사 실패: {str(e)}")
            return None
   
    def upload_new_file(self, file_content: bytes, filename: str, source_file: Dict) -> Optional[Dict]:
        """새 파일 업로드"""
        try:
            parent_ref = source_file.get('parentReference', {})
            parent_path = parent_ref.get('path', '')
           
            # URL 경로 수정: root: 대신 올바른 형식 사용
            if not parent_path:
                # 루트에 업로드
                encoded_filename = quote(filename, safe='')
                upload_url = f"https://graph.microsoft.com/v1.0/sites/{self.config['site_id']}/drives/{self.config['drive_id']}/root:/{encoded_filename}:/content"
            else:
                # 특정 폴더에 업로드
                encoded_path = quote(parent_path, safe='')
                encoded_filename = quote(filename, safe='')
                upload_url = f"https://graph.microsoft.com/v1.0/sites/{self.config['site_id']}/drives/{self.config['drive_id']}/root:/{encoded_path}/{encoded_filename}:/content"
           
            self.debug_print(f"업로드 URL: {upload_url}")
           
            # Content-Type 헤더 제거 (파일 업로드 시)
            headers = self.get_headers()
            if 'Content-Type' in headers:
                del headers['Content-Type']
           
            response = requests.put(upload_url, data=file_content, headers=headers)
           
            if response.status_code in [200, 201]:
                uploaded_file = response.json()
                print(f"[OK] 파일 업로드 성공: {uploaded_file.get('name', 'Unknown')}")
               
                # 업로드 후 즉시 확인
                print("[INFO] 업로드된 파일 확인 중...")
                uploaded_id = uploaded_file.get('id', '')
                if uploaded_id:
                    # 업로드된 파일을 다시 다운로드하여 확인
                    verification_content = self.download_file(uploaded_id)
                    if verification_content:
                        print(f"[OK] 업로드된 파일 확인 완료: {len(verification_content)} bytes")
                       
                        # Excel 파일인지 확인
                        if verification_content.startswith(b'PK'):
                            print("[OK] 업로드된 파일이 Excel 형식입니다.")
                        else:
                            print("[WARN] 업로드된 파일이 Excel 형식이 아닐 수 있습니다.")
                    else:
                        print("[WARN] 업로드된 파일 확인 실패")
               
                return uploaded_file
            else:
                print(f"[FAIL] 파일 업로드 실패: {response.status_code}")
                print(f"[ERROR] 응답: {response.text}")
               
                # 대안 방법: 폴더 ID를 사용한 업로드
                return self.upload_file_alternative(file_content, filename, source_file)
               
        except Exception as e:
            print(f"[FAIL] 파일 업로드 실패: {str(e)}")
            return None
   
    def upload_file_alternative(self, file_content: bytes, filename: str, source_file: Dict) -> Optional[Dict]:
        """대안 업로드 방법 (폴더 ID 사용)"""
        try:
            print("[INFO] 대안 업로드 방법 시도 중...")
           
            parent_ref = source_file.get('parentReference', {})
            parent_id = parent_ref.get('id', '')
           
            if not parent_id:
                print("[FAIL] 부모 폴더 ID를 찾을 수 없습니다.")
                return None
           
            # 폴더 ID를 사용한 업로드 (올바른 형식)
            encoded_filename = quote(filename, safe='')
            upload_url = f"https://graph.microsoft.com/v1.0/sites/{self.config['site_id']}/drives/{self.config['drive_id']}/items/{parent_id}:/{encoded_filename}:/content"
           
            self.debug_print(f"대안 업로드 URL: {upload_url}")
           
            headers = self.get_headers()
            if 'Content-Type' in headers:
                del headers['Content-Type']
           
            response = requests.put(upload_url, data=file_content, headers=headers)
           
            if response.status_code in [200, 201]:
                uploaded_file = response.json()
                print(f"[OK] 대안 업로드 성공: {uploaded_file.get('name', 'Unknown')}")
                return uploaded_file
            else:
                print(f"[FAIL] 대안 업로드 실패: {response.status_code}")
                print(f"[ERROR] 응답: {response.text}")
               
                # 최종 대안: 간단한 폴더 경로 사용
                return self.upload_file_simple(file_content, filename, source_file)
               
        except Exception as e:
            print(f"[FAIL] 대안 업로드 실패: {str(e)}")
            return None
   
    def upload_file_simple(self, file_content: bytes, filename: str, source_file: Dict) -> Optional[Dict]:
        """간단한 업로드 방법 (폴더 경로 직접 사용)"""
        try:
            print("[INFO] 간단한 업로드 방법 시도 중...")
           
            # 폴더 경로를 직접 사용
            folder_path = "General/주간업무보고서_진학어플라이본부"
            encoded_path = quote(folder_path, safe='')
            encoded_filename = quote(filename, safe='')
           
            upload_url = f"https://graph.microsoft.com/v1.0/sites/{self.config['site_id']}/drives/{self.config['drive_id']}/root:/{encoded_path}/{encoded_filename}:/content"
           
            self.debug_print(f"간단한 업로드 URL: {upload_url}")
           
            headers = self.get_headers()
            if 'Content-Type' in headers:
                del headers['Content-Type']
           
            response = requests.put(upload_url, data=file_content, headers=headers)
           
            if response.status_code in [200, 201]:
                uploaded_file = response.json()
                print(f"[OK] 간단한 업로드 성공: {uploaded_file.get('name', 'Unknown')}")
                return uploaded_file
            else:
                print(f"[FAIL] 간단한 업로드 실패: {response.status_code}")
                print(f"[ERROR] 응답: {response.text}")
                return None
               
        except Exception as e:
            print(f"[FAIL] 간단한 업로드 실패: {str(e)}")
            return None
   
    def copy_excel_sheet_with_next_week(self, file_info: Dict, source_sheet: str) -> bool:
        """Excel 시트를 다음 주차로 복사"""
        try:
            print(f"\n[STEP 5] Excel 시트 복사 (다음 주차)")
            print("-" * 40)
           
            if openpyxl is None:
                print("[FAIL] openpyxl 패키지가 설치되지 않았습니다.")
                return False
           
            file_name = file_info.get('name', '')
            file_id = file_info.get('id', '')
           
            print(f"[INFO] 대상 파일: {file_name}")
            print(f"[INFO] 원본 시트: {source_sheet}")
           
            # 다음 주차 시트명 생성 (올바른 형식으로)
            next_week_sheet = self.calculate_next_week_sheetname(source_sheet)
            print(f"[INFO] 복사할 시트명: {next_week_sheet}")
           
            # 파일 다운로드
            print("[INFO] 파일 다운로드 중...")
            file_content = self.download_file(file_id)
            if not file_content:
                print("[FAIL] 파일 다운로드 실패")
                return False
           
            print(f"[INFO] 다운로드된 파일 크기: {len(file_content)} bytes")
           
            # Excel 파일 처리
            print("[INFO] Excel 파일 처리 중...")
            import io
            file_stream = io.BytesIO(file_content)
           
            try:
                workbook = openpyxl.load_workbook(file_stream)
                print(f"[OK] Excel 파일 로드 성공. 시트 수: {len(workbook.sheetnames)}")
                print(f"[INFO] 현재 시트 목록: {workbook.sheetnames}")
            except Exception as e:
                print(f"[FAIL] Excel 파일 로드 실패: {str(e)}")
                return False
           
            # 원본 시트 확인
            if source_sheet not in workbook.sheetnames:
                print(f"[FAIL] 원본 시트를 찾을 수 없습니다: {source_sheet}")
                print(f"[INFO] 사용 가능한 시트: {workbook.sheetnames}")
                return False
           
            # 새 시트명이 이미 존재하는지 확인
            if next_week_sheet in workbook.sheetnames:
                print(f"[WARN] 시트 '{next_week_sheet}'가 이미 존재합니다. 삭제 후 복사합니다.")
                workbook.remove(workbook[next_week_sheet])
           
            # 시트 복사 (완전한 복사 방법)
            print(f"[INFO] 시트 복사 중: {source_sheet} → {next_week_sheet}")
            try:
                source_worksheet = workbook[source_sheet]
                
                # 새 시트 생성 (제목을 먼저 설정)
                print(f"[INFO] 새 시트 생성: {next_week_sheet}")
                new_worksheet = workbook.create_sheet(title=next_week_sheet)
                
                # 모든 셀 복사 (값, 스타일, 병합, 테두리 등 완전 복사)
                print("[INFO] 셀 복사 중 (완전한 스타일 포함)...")
                for row_num, row in enumerate(source_worksheet.iter_rows(), 1):
                    for col_num, cell in enumerate(row, 1):
                        new_cell = new_worksheet.cell(row=row_num, column=col_num)
                        
                        # 값 복사
                        if cell.value is not None:
                            new_cell.value = cell.value
                        
                        # 완전한 스타일 복사
                        try:
                            if cell.has_style:
                                # 폰트 복사
                                if cell.font and openpyxl_styles:
                                    new_cell.font = openpyxl_styles.Font(
                                        name=cell.font.name,
                                        size=cell.font.size,
                                        bold=cell.font.bold,
                                        italic=cell.font.italic,
                                        vertAlign=cell.font.vertAlign,
                                        underline=cell.font.underline,
                                        strike=cell.font.strike,
                                        color=cell.font.color
                                    )
                                
                                # 정렬 복사
                                if cell.alignment and openpyxl_styles:
                                    new_cell.alignment = openpyxl_styles.Alignment(
                                        horizontal=cell.alignment.horizontal,
                                        vertical=cell.alignment.vertical,
                                        text_rotation=cell.alignment.text_rotation,
                                        wrap_text=cell.alignment.wrap_text,
                                        shrink_to_fit=cell.alignment.shrink_to_fit,
                                        indent=cell.alignment.indent
                                    )
                                
                                # 테두리 복사
                                if cell.border and openpyxl_styles:
                                    new_cell.border = openpyxl_styles.Border(
                                        left=cell.border.left,
                                        right=cell.border.right,
                                        top=cell.border.top,
                                        bottom=cell.border.bottom,
                                        diagonal=cell.border.diagonal
                                    )
                                
                                # 배경색 복사
                                if cell.fill and openpyxl_styles:
                                    new_cell.fill = openpyxl_styles.PatternFill(
                                        fill_type=cell.fill.fill_type,
                                        start_color=cell.fill.start_color,
                                        end_color=cell.fill.end_color
                                    )
                                
                                # 숫자 형식 복사
                                if cell.number_format:
                                    new_cell.number_format = cell.number_format
                                    
                        except Exception as style_error:
                            # 스타일 복사 실패해도 계속 진행
                            print(f"[WARN] 셀 [{row_num},{col_num}] 스타일 복사 실패: {str(style_error)}")
                            pass
                
                # 병합된 셀 복사
                print("[INFO] 병합된 셀 복사 중...")
                for merged_range in source_worksheet.merged_cells.ranges:
                    try:
                        # 병합 범위를 새 시트에 적용
                        new_worksheet.merge_cells(str(merged_range))
                        print(f"[INFO] 병합된 셀 복사: {merged_range}")
                    except Exception as merge_error:
                        print(f"[WARN] 병합된 셀 복사 실패 {merged_range}: {str(merge_error)}")
                
                # 행 높이와 열 너비 복사
                print("[INFO] 행 높이와 열 너비 복사 중...")
                for row_num in range(1, source_worksheet.max_row + 1):
                    if source_worksheet.row_dimensions[row_num].height:
                        new_worksheet.row_dimensions[row_num].height = source_worksheet.row_dimensions[row_num].height
                
                if get_column_letter:
                    for col_num in range(1, source_worksheet.max_column + 1):
                        col_letter = get_column_letter(col_num)
                        if source_worksheet.column_dimensions[col_letter].width:
                            new_worksheet.column_dimensions[col_letter].width = source_worksheet.column_dimensions[col_letter].width
                
                # 시트 제목 재확인 및 수정
                if new_worksheet.title != next_week_sheet:
                    print(f"[WARN] 시트 제목이 예상과 다름: {new_worksheet.title} → {next_week_sheet}")
                    new_worksheet.title = next_week_sheet
                    print(f"[INFO] 시트 제목 수정 완료: {new_worksheet.title}")
                
                # 원본 시트의 H3 값 가져오기 (이전 주차의 차주 업무)
                prev_h3_value = None
                try:
                    prev_h3_cell = source_worksheet['H3']
                    if prev_h3_cell.value:
                        prev_h3_value = str(prev_h3_cell.value)
                        print(f"[INFO] 원본 시트의 H3 값: {prev_h3_value}")
                except Exception as h3_error:
                    print(f"[WARN] 원본 H3 값 가져오기 실패: {str(h3_error)}")
                
                # 셀 내용 자동 업데이트 (B2, B3, H3)
                print(f"\n[INFO] 셀 내용 자동 업데이트 시작...")
                print("-" * 40)
                self.update_sheet_cell_contents(new_worksheet, next_week_sheet, prev_h3_value)
                print("-" * 40)
                
                # 새 시트를 맨 앞으로 이동
                print("[INFO] 새 시트를 맨 앞으로 이동 중...")
                try:
                    # 시트 순서 변경: 새 시트를 첫 번째 위치로 이동
                    # move_sheet()를 사용하여 시트를 맨 앞으로 이동
                    workbook.move_sheet(next_week_sheet, offset=-(len(workbook.sheetnames) - 1))
                    print(f"[OK] 시트 '{next_week_sheet}'가 맨 앞으로 이동됨")
                except Exception as move_error:
                    print(f"[WARN] 시트 이동 실패: {str(move_error)}")
                
                print(f"[OK] 시트 복사 완료: {source_sheet} → {new_worksheet.title}")
                
            except Exception as copy_error:
                print(f"[FAIL] 시트 복사 실패: {str(copy_error)}")
                import traceback
                print(f"[DEBUG] 상세 오류: {traceback.format_exc()}")
                return False
           
            # 복사 후 시트 목록 확인 및 검증
            print(f"[INFO] 복사 후 시트 목록: {workbook.sheetnames}")
            
            # 새로 생성된 시트가 실제로 존재하는지 확인
            if next_week_sheet not in workbook.sheetnames:
                print(f"[FAIL] 시트 '{next_week_sheet}'가 워크북에 없습니다!")
                print(f"[INFO] 현재 시트 목록: {workbook.sheetnames}")
                return False
            
            # 새 시트의 내용 확인
            try:
                new_sheet = workbook[next_week_sheet]
                cell_count = 0
                for row in new_sheet.iter_rows():
                    for cell in row:
                        if cell.value is not None:
                            cell_count += 1
                
                print(f"[INFO] 새 시트 '{next_week_sheet}'의 데이터 셀 수: {cell_count}")
                
                if cell_count == 0:
                    print(f"[WARN] 새 시트에 데이터가 없습니다.")
                else:
                    print(f"[OK] 새 시트에 데이터가 있습니다.")
                    
            except Exception as sheet_check_error:
                print(f"[WARN] 새 시트 확인 중 오류: {str(sheet_check_error)}")
           
            # 파일 저장 (더 안전한 방법)
            print("[INFO] 파일 저장 중...")
            try:
                output_stream = io.BytesIO()
                workbook.save(output_stream)
                output_stream.seek(0)
                updated_content = output_stream.getvalue()
               
                print(f"[INFO] 저장된 파일 크기: {len(updated_content)} bytes")
               
                # 메모리에서 직접 검증 (임시 파일 없이)
                print("[INFO] 저장된 파일 검증 중...")
                try:
                    # 저장된 내용을 다시 BytesIO로 로드하여 검증
                    verification_stream = io.BytesIO(updated_content)
                    verification_workbook = openpyxl.load_workbook(verification_stream)
                    verification_sheets = verification_workbook.sheetnames
                    print(f"[INFO] 저장된 파일의 시트 목록 (처음 10개): {verification_sheets[:10]}")
                    
                    # 새 시트가 있는지 확인
                    if next_week_sheet in verification_sheets:
                        print(f"[OK] 시트 '{next_week_sheet}'가 저장된 파일에 존재합니다.")
                        
                        # 시트 내용도 간단히 확인
                        try:
                            target_sheet = verification_workbook[next_week_sheet]
                            cell_count = 0
                            for row in target_sheet.iter_rows():
                                for cell in row:
                                    if cell.value is not None:
                                        cell_count += 1
                           
                            print(f"[INFO] 저장된 시트 '{next_week_sheet}'의 데이터 셀 수: {cell_count}")
                           
                            if cell_count > 0:
                                print(f"[OK] 저장된 시트에 데이터가 있습니다.")
                            else:
                                print(f"[WARN] 저장된 시트에 데이터가 없습니다.")
                               
                        except Exception as content_error:
                            print(f"[WARN] 저장된 시트 내용 확인 중 오류: {str(content_error)}")
                       
                        verification_success = True
                    else:
                        print(f"[FAIL] 시트 '{next_week_sheet}'가 저장된 파일에 없습니다!")
                        print(f"[INFO] 저장된 파일의 모든 시트: {verification_sheets}")
                        verification_success = False
                   
                    # 워크북 닫기
                    verification_workbook.close()
                   
                except Exception as verification_error:
                    print(f"[FAIL] 저장된 파일 검증 실패: {str(verification_error)}")
                    verification_success = False
               
                if not verification_success:
                    return False
               
            except Exception as save_error:
                print(f"[FAIL] 파일 저장 실패: {str(save_error)}")
                return False
           
            # 업로드 전 최종 확인
            print("[INFO] 업로드 전 최종 시트 확인...")
            final_check_stream = io.BytesIO(updated_content)
            final_check_workbook = openpyxl.load_workbook(final_check_stream)
            final_sheets = final_check_workbook.sheetnames
            print(f"[INFO] 업로드할 파일의 시트 목록: {final_sheets}")
           
            if next_week_sheet not in final_sheets:
                print(f"[FAIL] 업로드 전 확인: 시트 '{next_week_sheet}'가 없습니다!")
                final_check_workbook.close()
                return False
           
            print(f"[OK] 업로드 전 확인: 시트 '{next_week_sheet}' 존재 확인")
            final_check_workbook.close()
           
            # 업로드
            print("[INFO] 업데이트된 파일 업로드 중...")
            uploaded_file = self.upload_new_file(updated_content, file_name, file_info)
            if uploaded_file:
                print(f"[OK] 시트 복사 및 파일 업데이트 완료")
               
                # 업로드 후 즉시 검증 (잠시 대기 후)
                print("[INFO] 업로드 완료 후 잠시 대기 중...")
                import time
                time.sleep(3)  # 3초 대기 (더 길게)
               
                print("[INFO] 업로드된 파일 검증 중...")
                verification_result = self.verify_sheet_copy(uploaded_file, next_week_sheet)
                if verification_result:
                    print(f"[OK] 시트 복사 검증 완료: {next_week_sheet}")
                   
                    # 추가 검증: 파일 크기 비교
                    original_size = len(file_content)
                    uploaded_content = self.download_file(uploaded_file.get('id', ''))
                    if uploaded_content:
                        uploaded_size = len(uploaded_content)
                        print(f"[INFO] 파일 크기 비교: 원본 {original_size} bytes, 업로드 후 {uploaded_size} bytes")
                       
                        if abs(original_size - uploaded_size) > 1000:  # 1KB 이상 차이나면 의심
                            print(f"[WARN] 파일 크기가 크게 다릅니다. 업로드에 문제가 있을 수 있습니다.")
                       
                        # 업로드된 파일의 시트도 확인
                        try:
                            uploaded_stream = io.BytesIO(uploaded_content)
                            uploaded_workbook = openpyxl.load_workbook(uploaded_stream)
                            uploaded_sheets = uploaded_workbook.sheetnames
                            print(f"[INFO] 업로드된 파일의 실제 시트 목록: {uploaded_sheets}")
                           
                            if next_week_sheet in uploaded_sheets:
                                print(f"[OK] 업로드된 파일에 시트 '{next_week_sheet}'가 실제로 존재합니다.")
                            else:
                                print(f"[FAIL] 업로드된 파일에 시트 '{next_week_sheet}'가 없습니다!")
                                print("[INFO] 업로드 과정에서 시트가 손실되었을 수 있습니다.")
                                uploaded_workbook.close()
                                return False
                           
                            uploaded_workbook.close()
                           
                        except Exception as sheet_check_error:
                            print(f"[WARN] 업로드된 파일 시트 확인 중 오류: {str(sheet_check_error)}")
                   
                    return True
                else:
                    print(f"[FAIL] 시트 복사 검증 실패: {next_week_sheet}")
                    print("[INFO] 실제 파일을 확인해보세요. 시트가 복사되지 않았을 수 있습니다.")
                    return False
            else:
                print("[FAIL] 파일 업데이트 실패")
                return False
               
        except Exception as e:
            print(f"[FAIL] 시트 복사 실패: {str(e)}")
            import traceback
            print(f"[DEBUG] 상세 오류: {traceback.format_exc()}")
            return False
   
    def verify_sheet_copy(self, file_info: Dict, expected_sheet: str) -> bool:
        """시트 복사 검증 (실제 파일 다운로드하여 검증)"""
        try:
            print(f"[INFO] 시트 복사 검증: {expected_sheet}")
            print("=" * 50)
           
            file_id = file_info.get('id', '')
            file_name = file_info.get('name', '')
           
            print(f"[INFO] 검증 대상 파일: {file_name}")
            print(f"[INFO] 파일 ID: {file_id}")
            print(f"[INFO] 찾을 시트명: '{expected_sheet}'")
           
            # 실제 파일 다운로드
            print("[INFO] 실제 파일 다운로드 중...")
            file_content = self.download_file(file_id)
            if not file_content:
                print("[FAIL] 검증용 파일 다운로드 실패")
                return False
           
            print(f"[INFO] 다운로드된 파일 크기: {len(file_content)} bytes")
           
            # Excel 파일 열기
            import io
            file_stream = io.BytesIO(file_content)
            workbook = openpyxl.load_workbook(file_stream)
           
            all_sheets = workbook.sheetnames
            print(f"[INFO] 업로드된 파일의 모든 시트 ({len(all_sheets)}개):")
            for i, sheet in enumerate(all_sheets, 1):
                print(f"  {i}. '{sheet}'")
           
            # 정확한 시트명 매칭
            exact_match = expected_sheet in all_sheets
            print(f"[INFO] 정확한 시트명 매칭: {exact_match}")
           
            if exact_match:
                print(f"[OK] 시트 '{expected_sheet}' 검증 성공!")
               
                # 시트 내용 상세 확인
                try:
                    target_sheet = workbook[expected_sheet]
                    print(f"[INFO] 시트 '{expected_sheet}' 상세 분석:")
                   
                    # 셀 데이터 확인
                    cell_count = 0
                    non_empty_rows = 0
                    max_col = 0
                   
                    for row_num, row in enumerate(target_sheet.iter_rows(), 1):
                        row_has_data = False
                        for col_num, cell in enumerate(row, 1):
                            if cell.value is not None:
                                cell_count += 1
                                row_has_data = True
                                max_col = max(max_col, col_num)
                       
                        if row_has_data:
                            non_empty_rows += 1
                   
                    print(f"  - 데이터가 있는 셀 수: {cell_count}")
                    print(f"  - 데이터가 있는 행 수: {non_empty_rows}")
                    print(f"  - 최대 열 번호: {max_col}")
                   
                    if cell_count > 0:
                        print(f"[OK] 시트에 실제 데이터가 있습니다.")
                       
                        # 첫 번째 몇 개 셀 값 확인
                        print(f"[INFO] 첫 번째 셀 값들:")
                        for row_num in range(1, min(4, non_empty_rows + 1)):
                            for col_num in range(1, min(4, max_col + 1)):
                                cell = target_sheet.cell(row=row_num, column=col_num)
                                if cell.value is not None:
                                    print(f"  [{row_num},{col_num}]: {str(cell.value)[:50]}")
                       
                    else:
                        print(f"[WARN] 시트에 데이터가 없습니다.")
                       
                except Exception as content_error:
                    print(f"[WARN] 시트 내용 확인 중 오류: {str(content_error)}")
               
                workbook.close()
                return True
               
            else:
                print(f"[FAIL] 시트 '{expected_sheet}' 검증 실패!")
                print(f"[INFO] 예상 시트명: '{expected_sheet}'")
                print(f"[INFO] 실제 시트 목록: {all_sheets}")
               
                # 유사한 시트명 찾기
                similar_sheets = []
                for sheet in all_sheets:
                    # 공백 제거 후 비교
                    clean_expected = expected_sheet.replace(' ', '').replace('년', '').replace('월', '').replace('주차', '')
                    clean_sheet = sheet.replace(' ', '').replace('년', '').replace('월', '').replace('주차', '')
                   
                    if clean_expected in clean_sheet or clean_sheet in clean_expected:
                        similar_sheets.append(sheet)
               
                if similar_sheets:
                    print(f"[INFO] 유사한 시트명 발견: {similar_sheets}")
                else:
                    print(f"[INFO] 유사한 시트명이 없습니다.")
               
                workbook.close()
                return False
               
        except Exception as e:
            print(f"[FAIL] 시트 검증 실패: {str(e)}")
            import traceback
            print(f"[DEBUG] 검증 오류 상세: {traceback.format_exc()}")
            return False
    
    def get_file_share_link(self, file_info: Dict) -> Optional[str]:
        """SharePoint 파일의 공유 링크 생성 (여러 방법 시도)"""
        try:
            print(f"[INFO] SharePoint 파일 공유 링크 생성 중...")
            
            file_id = file_info.get('id', '')
            file_name = file_info.get('name', '')
            
            print(f"[INFO] 파일명: {file_name}")
            print(f"[INFO] 파일 ID: {file_id}")
            
            # 방법 1: 파일의 기존 공유 권한에서 링크 조회 (최우선)
            print(f"[INFO] 기존 공유 권한 조회 중...")
            permissions_url = f"https://graph.microsoft.com/v1.0/sites/{self.config['site_id']}/drives/{self.config['drive_id']}/items/{file_id}/permissions"
            headers = self.get_headers()
            
            try:
                perm_response = requests.get(permissions_url, headers=headers)
                if perm_response.status_code == 200:
                    permissions = perm_response.json().get('value', [])
                    self.debug_print(f"권한 목록: {len(permissions)}개")
                    
                    # 조직 공유 링크 찾기
                    for perm in permissions:
                        link = perm.get('link', {})
                        if link and link.get('scope') in ['organization', 'users']:
                            share_url = link.get('webUrl', '')
                            if share_url and '/:x:/' in share_url:
                                print(f"[OK] 기존 공유 링크 발견!")
                                print(f"[INFO] 링크: {share_url}")
                                return share_url
                    
                    print(f"[INFO] 기존 공유 링크 없음. 새로 생성합니다.")
            except Exception as perm_error:
                self.debug_print(f"권한 조회 실패: {str(perm_error)}")
            
            # 방법 2: Graph API로 새 공유 링크 생성
            print(f"[INFO] Graph API로 공유 링크 생성 중...")
            create_url = f"https://graph.microsoft.com/v1.0/sites/{self.config['site_id']}/drives/{self.config['drive_id']}/items/{file_id}/createLink"
            
            # 링크 스코프 설정 (환경 변수에서 가져오기)
            # organization: 조직 내부용 (?e= 없을 수 있음)
            # anonymous: 외부 공유 가능 (?e= 포함됨)
            link_scope = os.getenv('SHAREPOINT_LINK_SCOPE', 'organization')
            
            payload = {
                "type": "view",
                "scope": link_scope
            }
            
            print(f"[INFO] 링크 스코프: {link_scope}")
            
            self.debug_print(f"공유 링크 생성 URL: {create_url}")
            
            response = requests.post(create_url, headers=headers, json=payload)
            
            if response.status_code in [200, 201]:
                data = response.json()
                
                # 디버그: 전체 응답 확인
                self.debug_print(f"공유 링크 API 응답: {json.dumps(data, indent=2, ensure_ascii=False)}")
                
                # link 객체에서 webUrl 추출
                link_data = data.get('link', {})
                share_link = link_data.get('webUrl', '')
                
                if share_link:
                    print(f"[OK] 공유 링크 생성 완료 (Graph API)")
                    print(f"[INFO] 링크: {share_link}")
                    
                    # 짧은 링크 형식인지 확인
                    if '/:x:/' in share_link and '?e=' in share_link:
                        print(f"[OK] 미리보기 지원 링크입니다!")
                    elif '/:x:/' in share_link:
                        print(f"[OK] 짧은 링크 형식입니다.")
                    else:
                        print(f"[WARN] Doc.aspx 형식 링크입니다. 미리보기는 작동하지만 링크가 깁니다.")
                    
                    return share_link
                else:
                    print(f"[WARN] 공유 링크가 응답에 없습니다.")
            else:
                print(f"[WARN] 공유 링크 생성 실패: {response.status_code}")
                self.debug_print(f"응답: {response.text}")
            
            # 방법 3: 파일의 webUrl 사용
            file_web_url = file_info.get('webUrl', '')
            if file_web_url:
                print(f"[INFO] 파일의 webUrl 사용 (대체)")
                print(f"[INFO] 링크: {file_web_url}")
                return file_web_url
            
            # 방법 4: 직접 경로 링크로 대체
            print(f"[INFO] 직접 경로 링크로 대체합니다.")
            return self.get_file_direct_link(file_info)
            
        except Exception as e:
            print(f"[FAIL] 공유 링크 생성 실패: {str(e)}")
            print(f"[INFO] 직접 경로 링크로 대체합니다.")
            return self.get_file_direct_link(file_info)
    
    def get_file_direct_link(self, file_info: Dict) -> Optional[str]:
        """SharePoint 파일의 직접 경로 링크 생성 (대체 방법)"""
        try:
            file_name = file_info.get('name', '')
            
            # 사이트 URL에서 기본 경로 추출
            site_url = self.config['site_url']
            base_url = site_url.replace('/sites/ApplyService2', '')
            
            # 파일 경로 구성
            folder_path = "General/주간업무보고서_진학어플라이본부"
            encoded_folder = quote(folder_path, safe='')
            encoded_filename = quote(file_name, safe='')
            
            # SharePoint 직접 링크 생성
            direct_link = f"{base_url}/sites/ApplyService2/Shared%20Documents/{encoded_folder}/{encoded_filename}"
            
            print(f"[OK] 직접 경로 링크 생성 완료")
            print(f"[INFO] 링크: {direct_link}")
            
            return direct_link
            
        except Exception as e:
            print(f"[FAIL] 직접 링크 생성 실패: {str(e)}")
            return None
    
    def get_sender_by_week(self, month: int, week: int) -> str:
        """월 주차별 발송자 결정 (순환)"""
        try:
            # 발송자 목록 (순환 순서)
            senders = [
                "임형섭 부장님",
                "전성대 부장님",
                "허승철 부장님"
            ]
            
            # 기준점: 2026년 1월 5주차 = 임형섭 부장님 (인덱스 0) - 실제 발송 이력 기준
            # 2026년 1월 1주차부터 누적 주차 계산
            
            # ISO 8601 기준으로 각 월의 주차 수를 동적 계산
            # (목요일이 속한 월이 해당 주의 월을 결정)
            
            # 1월부터 현재 월 이전까지의 누적 주차 수
            total_weeks = 0
            for m in range(1, month):
                total_weeks += self.get_iso_month_weeks_count(2026, m)
            
            # 현재 월의 주차 추가
            total_weeks += (week - 1)  # 1주차는 0, 2주차는 1...
            
            # 1월 5주차 = 누적 4주차 (0부터 시작) = 임형섭 부장님 (인덱스 0)
            # 역산: (base_offset + 4) % 3 = 0 → base_offset = -4 % 3 = 2
            base_offset = 2
            
            # 순환 인덱스 계산
            cycle_index = (base_offset + total_weeks) % len(senders)
            
            selected_sender = senders[cycle_index]
            
            print(f"[INFO] 발송자 선택: {month}월 {week}주차 → {selected_sender}")
            print(f"[INFO] 2026년 기준 누적 주차: {total_weeks}")
            print(f"[INFO] 순환 인덱스: {cycle_index} (총 {len(senders)}명)")
            
            return selected_sender
            
        except Exception as e:
            print(f"[FAIL] 발송자 선택 실패: {str(e)}")
            return "허승철 부장님"  # 기본값
    
    def create_teams_message(self, file_info: Dict, month: int, week: int) -> Optional[str]:
        """Teams 메시지 내용 생성 (HTML 형식)"""
        try:
            print(f"[INFO] Teams 메시지 생성 중...")
            
            # 파일 공유 링크 생성
            share_link = self.get_file_share_link(file_info)
            if not share_link:
                print("[FAIL] 공유 링크 생성 실패")
                return None
            
            # 발송자 선택
            sender = self.get_sender_by_week(month, week)
            
            # 파일명 추출
            file_name = file_info.get('name', '주간업무보고서')
            
            # HTML 형식의 메시지 생성 (링크 클릭 가능)
            message = f"""안녕하세요.<br>
{month}월 {week}주차 주간보고 공유드립니다.<br>
작성 후 좋아요 눌러주세요.<br>
<br>
발송자 : {sender}<br>
<br>
<a href="{share_link}">{file_name}</a>"""
            
            print(f"[OK] Teams 메시지 생성 완료")
            print(f"[INFO] 메시지 내용:")
            print("-" * 50)
            # 디버그 출력용 (텍스트 형식으로)
            print(f"안녕하세요.")
            print(f"{month}월 {week}주차 주간보고 공유드립니다.")
            print(f"작성 후 좋아요 눌러주세요.")
            print(f"")
            print(f"발송자 : {sender}")
            print(f"{file_name}: {share_link}")
            print("-" * 50)
            
            return message
            
        except Exception as e:
            print(f"[FAIL] Teams 메시지 생성 실패: {str(e)}")
            return None
    
    def list_teams_chats(self) -> List[Dict]:
        """사용 가능한 Teams 채팅 목록 조회"""
        try:
            print(f"[INFO] Teams 채팅 목록 조회 중...")
            
            url = "https://graph.microsoft.com/v1.0/chats"
            headers = self.get_headers()
            
            response = requests.get(url, headers=headers)
            
            if response.status_code == 200:
                data = response.json()
                chats = data.get('value', [])
                
                print(f"[OK] {len(chats)}개의 채팅을 찾았습니다:")
                print("-" * 60)
                
                for i, chat in enumerate(chats, 1):
                    chat_id = chat.get('id', 'Unknown')
                    topic = chat.get('topic', '제목 없음')
                    chat_type = chat.get('chatType', 'Unknown')
                    
                    print(f"{i}. 채팅 ID: {chat_id}")
                    print(f"   제목: {topic}")
                    print(f"   타입: {chat_type}")
                    print()
                
                return chats
            else:
                print(f"[FAIL] 채팅 목록 조회 실패: {response.status_code}")
                print(f"[ERROR] 응답: {response.text}")
                return []
                
        except Exception as e:
            print(f"[FAIL] 채팅 목록 조회 실패: {str(e)}")
            return []
    
    def get_teams_chat_id(self) -> Optional[str]:
        """Teams 그룹 채팅 ID 조회"""
        try:
            print(f"[INFO] Teams 그룹 채팅 ID 조회 중...")
            
            # 환경변수에서 채팅 ID 가져오기
            chat_id = os.getenv('TEAMS_CHAT_ID')
            
            if not chat_id:
                print("[FAIL] TEAMS_CHAT_ID 환경변수가 설정되지 않았습니다.")
                print("[INFO] 사용 가능한 채팅 목록을 조회합니다...")
                
                # 채팅 목록 조회
                chats = self.list_teams_chats()
                
                if chats:
                    print("\n[INFO] 환경변수 파일에 원하는 채팅 ID를 설정하세요:")
                    print("TEAMS_CHAT_ID=19:chat_id_here@thread.v2")
                
                return None
            
            print(f"[OK] Teams 채팅 ID 확인: {chat_id[:30]}...")
            return chat_id
            
        except Exception as e:
            print(f"[FAIL] Teams 채팅 ID 조회 실패: {str(e)}")
            return None
    
    def send_teams_message(self, message: str) -> bool:
        """Teams 그룹 채팅 메시지 전송"""
        try:
            print(f"[INFO] Teams 그룹 채팅 메시지 전송 중...")
            
            # Teams용 토큰 확인
            if not self.teams_access_token:
                print("[FAIL] Teams 인증 토큰이 없습니다.")
                print("[INFO] TEAMS_AUTH_TYPE=delegated로 설정하고 다시 실행하세요.")
                return False
            
            # Teams 채팅 ID 조회
            chat_id = self.get_teams_chat_id()
            if not chat_id:
                return False
            
            # Graph API URL 구성
            url = f"https://graph.microsoft.com/v1.0/chats/{chat_id}/messages"
            
            # 메시지 페이로드 구성 (HTML 형식으로 링크 클릭 가능)
            payload = {
                "body": {
                    "contentType": "html",
                    "content": message
                }
            }
            
            # HTTP 헤더 (Teams 전용 토큰 사용)
            headers = {
                'Authorization': f'Bearer {self.teams_access_token}',
                'Content-Type': 'application/json'
            }
            
            print(f"[INFO] Graph API URL: {url}")
            print(f"[INFO] 메시지 내용: {message[:100]}...")
            
            # 메시지 전송
            response = requests.post(
                url,
                headers=headers,
                data=json.dumps(payload),
                timeout=30
            )
            
            if response.status_code == 201:
                print(f"[OK] Teams 그룹 채팅 메시지 전송 성공!")
                response_data = response.json()
                message_id = response_data.get('id', 'Unknown')
                print(f"[INFO] 메시지 ID: {message_id}")
                return True
            else:
                print(f"[FAIL] Teams 메시지 전송 실패: {response.status_code}")
                print(f"[ERROR] 응답: {response.text}")
                
                # 상세 오류 정보 출력
                try:
                    error_data = response.json()
                    error_code = error_data.get('error', {}).get('code', 'Unknown')
                    error_message = error_data.get('error', {}).get('message', 'Unknown')
                    print(f"[ERROR] 오류 코드: {error_code}")
                    print(f"[ERROR] 오류 메시지: {error_message}")
                    
                    # 권한 관련 오류인 경우 안내 메시지
                    if response.status_code == 401:
                        print("\n[INFO] Teams 그룹 채팅 메시지 전송 해결 방법:")
                        print("1. 환경 설정 파일에서 TEAMS_AUTH_TYPE=delegated로 변경")
                        print("2. Teams 그룹 채팅에 메시지를 보낼 수 있는 사용자 계정 설정:")
                        print("   TEAMS_USERNAME=your_email@domain.com")
                        print("   TEAMS_PASSWORD=your_password")
                        print("3. Azure AD 앱 등록에서 Delegated Permission 추가:")
                        print("   - Chat.ReadWrite (사용자 위임 권한)")
                        print("4. 사용자 동의 부여")
                        
                except:
                    pass
                
                return False
                
        except Exception as e:
            print(f"[FAIL] Teams 메시지 전송 실패: {str(e)}")
            import traceback
            print(f"[DEBUG] 상세 오류: {traceback.format_exc()}")
            return False
    

def main():
    """메인 함수"""
    print("🔍 SharePoint 폴더 탐색기")
    print("=" * 50)
   
    try:
        # 탐색기 초기화
        explorer = SharePointFolderExplorer(debug_mode=True)
       
        # 1. 먼저 루트 폴더 구조 확인
        print(f"\n{'='*60}")
        print("SharePoint 루트 폴더 구조 확인")
        print('='*60)
        
        root_folders = explorer.list_root_folders()
        
        # 2. 가능한 폴더 경로들 시도
        possible_paths = [
            "General/General",
            "주간업무보고서_진학어플라이본부", 
            "General/주간업무보고서",
            "주간업무보고서",
            "진학어플라이본부",
            "업무보고서"
        ]
        
        working_folder = None
        for path in possible_paths:
            print(f"\n{'='*60}")
            print(f"폴더 경로 시도: {path}")
            print('='*60)
            
            files = explorer.explore_folder(path)
            if files:
                working_folder = path
                print(f"[SUCCESS] 폴더 발견: {path}")
                break
            else:
                print(f"[FAIL] 폴더 없음: {path}")
       
        if working_folder:
            print(f"[SUCCESS] 작업 폴더 발견: {working_folder}")
        else:
            print(f"[FAIL] 모든 폴더 경로 시도 실패")
       
        # 2. 최근 수정 파일 찾기
        latest_file = None
        if working_folder:
            print(f"\n{'='*60}")
            print(f"최근 수정 파일 검색: {working_folder}")
            print('='*60)
           
            latest_file = explorer.find_latest_file_in_folder(
                working_folder,
                "주간업무보고서_진학어플라이본부"
            )
       
        # 3. Excel 시트 분석
        sheet_analysis = None
        if latest_file:
            print(f"\n{'='*60}")
            print("Excel 시트 분석")
            print('='*60)
           
            sheet_analysis = explorer.analyze_latest_file_sheets(latest_file)
       
        # 4. 파일 복사 (다음 주차)
        copied_file = None
        if latest_file:
            print(f"\n{'='*60}")
            print("파일 복사 (다음 주차)")
            print('='*60)
           
            copied_file = explorer.copy_file_with_next_week(latest_file)
       
        # 5. Excel 시트 복사 (다음 주차)
        sheet_copied = False
        if copied_file and sheet_analysis and sheet_analysis.get('latest_sheet'):
            print(f"\n{'='*60}")
            print("Excel 시트 복사 (다음 주차)")
            print('='*60)
           
            sheet_copied = explorer.copy_excel_sheet_with_next_week(
                copied_file,
                sheet_analysis['latest_sheet']
            )
        
        # 6. Teams 메시지 전송
        teams_sent = False
        if copied_file and sheet_copied:
            print(f"\n{'='*60}")
            print("Teams 메시지 전송")
            print('='*60)
            
            # 파일명에서 월과 주차 추출
            file_name = copied_file.get('name', '')
            month_match = re.search(r'(\d+)월(\d+)주차', file_name)
            
            if month_match:
                month = int(month_match.group(1))
                week = int(month_match.group(2))
                
                # Teams 메시지 생성
                teams_message = explorer.create_teams_message(copied_file, month, week)
                
                if teams_message:
                    # Teams 메시지 전송
                    teams_sent = explorer.send_teams_message(teams_message)
                else:
                    print("[FAIL] Teams 메시지 생성 실패")
            else:
                print(f"[FAIL] 파일명에서 월/주차 정보를 추출할 수 없습니다: {file_name}")
       
        # 결과 요약
        print(f"\n{'='*60}")
        print("📋 탐색 결과 요약")
        print('='*60)
       
        if working_folder:
            print(f"✅ 작업 폴더: {working_folder}")
        else:
            print("❌ 작업 폴더: 없음")
       
        if latest_file:
            print(f"✅ 최근 파일: {latest_file.get('name', 'Unknown')}")
            print(f"✅ 파일 ID: {latest_file.get('id', 'Unknown')}")
            print(f"✅ 수정일: {latest_file.get('lastModifiedDateTime', 'Unknown')}")
        else:
            print("❌ 최근 파일: 없음")
       
        if sheet_analysis:
            print(f"\n📊 Excel 시트 정보:")
            print(f"✅ 총 시트 수: {sheet_analysis['total_sheets']}")
            if sheet_analysis['latest_sheet']:
                print(f"✅ 최근 시트: {sheet_analysis['latest_sheet']}")
        else:
            print("\n❌ Excel 시트 정보: 없음")
       
        if copied_file:
            print(f"\n📄 파일 복사 결과:")
            print(f"✅ 복사된 파일: {copied_file.get('name', 'Unknown')}")
            print(f"✅ 파일 ID: {copied_file.get('id', 'Unknown')}")
        else:
            print("\n❌ 파일 복사: 실패")
       
        if sheet_copied:
            print(f"\n📋 시트 복사 결과:")
            print(f"✅ 시트 복사: 성공")
        else:
            print("\n❌ 시트 복사: 실패")
        
        # Teams 메시지 전송 결과
        if teams_sent:
            print(f"\n📱 Teams 메시지 전송 결과:")
            print(f"✅ Teams 메시지: 전송 성공")
        else:
            print("\n❌ Teams 메시지: 전송 실패")
       
    except Exception as e:
        print(f"\n[ERROR] 전체 프로세스 실패: {str(e)}")

if __name__ == "__main__":
    main()